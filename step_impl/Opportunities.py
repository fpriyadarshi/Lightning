from decimal import Decimal
from getgauge.python import step, Messages, after_suite, after_spec, before_suite, after_step
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import uuid
import json
import os
from pathlib import Path
import sys
import xlsxwriter
from getgauge.python import data_store
from datetime import date
from time import sleep
import random
from random import randint
from step_impl import Drivers
from step_impl import Common_Steps
from step_impl import Utils
from selenium.webdriver.support.select import Select
from getgauge.python import DataStoreFactory, Screenshots
from datetime import datetime
import shutil
import yaml
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook.workbook import Workbook
import pdb
import pandas as pd
from itertools import product
import logging

# from itertools import product


class Opportunities():

    # --------------------------
    # Gauge step implementations
    # --------------------------
    def get_Current_Fiscal_Year(self):
        try:
            currentDate = date.today()
            calendar_date_time_obj = datetime.strftime(currentDate, '%Y-%m-%d')
            minDateQuery = "SELECT Fiscal_Year__c FROM Calendar__c where Calendar_Date__c = " + calendar_date_time_obj + " limit 1"
            queryResult = Drivers.sf.query_all(query=minDateQuery)
            rec = queryResult['records']
            currentFiscalYear = rec[0]['Fiscal_Year__c']
            print("Current Fiscal Year: ", rec[0]['Fiscal_Year__c'])                
            return currentFiscalYear
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
            Messages.write_message(f"{exc_type}, {fname}, {exc_tb.tb_lineno}")

    def get_FiscalYearList(self):
        try:
            inStoreCycle = []
            currentFiscalYear = int(self.get_Current_Fiscal_Year())
            previous2Year = currentFiscalYear - 2
            previous1Year = currentFiscalYear - 1
            Next1Year = currentFiscalYear + 1
            Next2Year = currentFiscalYear + 2
            years = [previous2Year, previous1Year, currentFiscalYear, Next1Year, Next2Year]

            startPoint = 0
            endPoint = 0
            for year in years:
                if year == previous2Year:
                    startPoint = 7
                    endPoint = 14
                elif (year == Next1Year) or (year == previous1Year) or (year == currentFiscalYear):
                    startPoint = 1
                    endPoint = 14
                elif year == Next2Year:
                    startPoint = 1
                    endPoint = 7
                    
                for x in range(startPoint,endPoint,1):
                    if len(str(x)) == 1:
                        fiscalYear = str(year) + "0" + str(x)
                    else:
                        fiscalYear = str(year) + str(x)  
                    inStoreCycle.append(fiscalYear)
            logging.info(f"Cycles List: {inStoreCycle}")
            print(f"Cycles List: {inStoreCycle}")
            return inStoreCycle
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
            Messages.write_message(f"{exc_type}, {fname}, {exc_tb.tb_lineno}")
            
    @step("Select Opportunity Type <productType>")
    def open_tab(self, productType):
        Drivers.driverWait.until(
            EC.frame_to_be_available_and_switch_to_it((By.XPATH, f"//iframe[@title='accessibility title']")))
        
        xPath = "//div[@id='rtDivId']/select"
        Drivers.driverWait.until(
            EC.visibility_of_element_located((By.XPATH, xPath)))        
        dropDownProductLine = Select(
            Drivers.driver.find_element_by_xpath(xPath))
        dropDownProductLine.select_by_value(productType)
        
        xPath = f"//input[contains(@value,'Next')]"
        buttonToClick = Drivers.driver.find_element_by_xpath(xPath)
        # Drivers.driver.execute_script("scroll(250, 0)")
        buttonToClick.click()
        sleep(5)

    @step("Create <productType> Opportunity And Save Details to <fileName> <table>")
    def create_opportunity(self, productType, fileName, table):
        data_store.spec.clear()
        opportunityDictionary = {}
        opportunitiesDetails = {}
        Drivers.driverWait.until(
            EC.visibility_of_element_located((By.XPATH, "//input[@value='Next']")))
        fieldsData = {}
        # currentDate = date.today()
        # currentDate = currentDate.strftime("%m/%d/%Y")
        dt = datetime.today()
        print(dt.month, " ", dt.day, " ", dt.year)
        year = int(dt.year) + 2
        currentDate = f"{str(dt.month)}/{str(dt.day)}/{str(year)}"
        napiDate = ""
        if productType == "FSI_US" or productType == "FSI_CA":
            napiDate = Utils.getNAPIDate()
            currentDate = napiDate

        colHeaderMap = {}
        colMasterHeaderMap = {}
        columnIndex = 0
        rows = table.rows
        for row in rows:
            row0 = str(row[0]).strip()
            row1 = str(row[1]).strip()
            if row0 == "Name":
                fieldsData["Name"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Name", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Name")
            elif row0 == "Stage":
                fieldsData["Stage"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Stage", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Stage")
            elif row0 == "Probability":
                fieldsData["Probability"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Probability", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Probability")
            elif row0 == "Account ID":
                fieldsData["Account ID"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Account ID", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Account ID")
            elif row0 == "Freedom":
                fieldsData["Freedom"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Freedom Enabled", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Freedom Enabled")
            elif row0 == "Insert Date/ Start Date":
                fieldsData["Insert Date/ Start Date"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Insert Date/ Start Date", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Insert Date/ Start Date")
            elif row0 == "Execution End Date":
                fieldsData["Execution End Date"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Execution End Date", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Execution End Date")
            elif row0 == "Expected close date":
                fieldsData["Expected close date"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Expected close date", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Expected close date")
            elif row0 == "Contract Date":
                fieldsData["Contract Date"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Contract Date", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Contract Date")
            elif row0 == "Buisness Type":
                fieldsData["Buisness Type"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Buisness Type", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Buisness Type")
            elif row0 == "Projected Average CPM":
                fieldsData["Projected Average CPM"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Projected Average CPM", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Projected Average CPM")
            elif row0 == "Projected Circulation (000)":
                fieldsData["Projected Circulation (000)"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Projected Circulation (000)", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Projected Circulation (000)")
            elif row0 == "ILoc Asset Due Date":
                fieldsData["ILoc Asset Due Date"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "ILoc Asset Due Date", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "ILoc Asset Due Date")
            elif row0 == "Billed based on Actual Execution":
                fieldsData["Billed based on Actual Execution"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Billed based on Actual Execution", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Billed based on Actual Execution")

            # #SSMG FIELDS
            # elif row0 == "ILoc Market List Due Date":
            #     fieldsData["ILoc Market List Due Date"] = row1
            # elif row0 == "ILoc Material Due Date":
            #     fieldsData["ILoc Material Due Date"] = row1

            # FSI FIELDS
            elif row0 == "FSI - Insert Date/Start Date":
                fieldsData["FSI - Insert Date/Start Date"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "FSI - Insert Date/Start Date", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "FSI - Insert Date/Start Date")

            # MERCHANDISING FIELDS
            elif row0 == "Client List Due Date":
                fieldsData["Client List Due Date"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Client List Due Date", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Client List Due Date")
            elif row0 == "Artwork Due Date":
                fieldsData["Artwork Due Date"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Artwork Due Date", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Artwork Due Date")
            elif row0 == "ILoc Format":
                fieldsData["ILoc Format"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "ILoc Format", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "ILoc Format")
            elif row0 == "ILoc Targeting":
                fieldsData["ILoc Targeting"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "ILoc Targeting", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "ILoc Targeting")
            elif row0 == "ILoc Artwork Due Date":
                fieldsData["ILoc Artwork Due Date"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "ILoc Artwork Due Date", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "ILoc Artwork Due Date")
            elif row0 == "ILoc Sample Due Date":
                fieldsData["ILoc Sample Due Date"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "ILoc Sample Due Date", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "ILoc Sample Due Date")
            elif row0 == "ILoc Creative Due Date":
                fieldsData["ILoc Creative Due Date"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "ILoc Creative Due Date", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "ILoc Creative Due Date")
            elif row0 == "Opportunity Type":
                fieldsData["Opportunity Type"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Opportunity Type", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Opportunity Type")
            elif row0 == "ILoc Signed LOC/Store List Due Date":
                fieldsData["ILoc Signed LOC/Store List Due Date"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "ILoc Signed LOC/Store List Due Date", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "ILoc Signed LOC/Store List Due Date")
            elif row0 == "ILoc Program Document Due Date":
                fieldsData["ILoc Program Document Due Date"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "ILoc Program Document Due Date", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "ILoc Program Document Due Date")
            elif row0 == "ILoc Material Due Date":
                fieldsData["ILoc Material Due Date"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "ILoc Material Due Date", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "ILoc Material Due Date")
            elif row0 == "Material Due To":
                fieldsData["Material Due To"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Material Due To", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Material Due To")
            elif row0 == "POS Materials Produced":
                fieldsData["POS Materials Produced"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "POS Materials Produced", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "POS Materials Produced")
            elif row0 == "POS Material Disposition":
                fieldsData["POS Material Disposition"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "POS Material Disposition", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "POS Material Disposition")
            elif row0 == "Image Request Details":
                fieldsData["Image Request Details"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Image Request Details", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Image Request Details")
            elif row0 == "Image Subset List Count":
                fieldsData["Image Subset List Count"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Image Subset List Count", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Image Subset List Count")
            elif row0 == "Network Retailer List":
                fieldsData["Network Retailer List"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Network Retailer List", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Network Retailer List")
            elif row0 == "Non Network Retailer List":
                fieldsData["Non Network Retailer List"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Non Network Retailer List", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Non Network Retailer List")
            elif row0 == "Job Description/Comments":
                fieldsData["Job Description/Comments"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Job Description/Comments", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Job Description/Comments")

            # INSTORE FIELDS
            elif row0 == "InStore Cycle":
                fieldsData["InStore Cycle"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "InStore Cycle", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "InStore Cycle")
            elif row0 == "Projected Average CPS(Space)":
                fieldsData["Projected Average CPS(Space)"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Projected Average CPS(Space)", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Projected Average CPS(Space)")
            elif row0 == "Projected Store Count":
                fieldsData["Projected Store Count"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Projected Store Count", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Projected Store Count")
            elif row0 == "Expiration Date":
                fieldsData["Expiration Date"] = row1
                # Utils.setExcelColumnHeaders(self, productType, "Expiration Date", opportunitiesDetailsFileName, ws1, wb)
#                 colHeaderMap, columnIndex = Utils.setColumnDetails(self, productType, ws1, columnIndex, colHeaderMap, wb, "Expiration Date")
        Messages.write_message(fieldsData)
        fieldType = None
#         wb.save(filename=opportunitiesDetailsFileName)

        print(colHeaderMap)
#         colMasterHeaderMap[productType] = colHeaderMap
#         opportunityHeaderFilePath = str(opportunitiesDetailsPath) + "\\Data\\" + os.getenv("OPPORTUNITY_COLUMN_HEADER_FILE")
#         Utils.set_Opportunities_Header_Details(opportunityHeaderFilePath, colMasterHeaderMap)
        # with open(os.getenv("OPPORTUNITY_COLUMN_HEADER_FILE"), 'w') as yaml_file:
        #             yaml.dump(colHeaderMap, yaml_file, default_flow_style=False)

        # rowNum = len(ws1['A']) + 1
#         rowNum = int(ws1.max_row) + 1
        if "Name" in fieldsData:
            today = date.today()
            oppDateToday = today.strftime("%Y%m%d")

            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='Name']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='Name']")
            oppName = fieldsData["Name"] + "_" + oppDateToday
            fieldType.send_keys(oppName)
            Messages.write_message("Opportunity Name: " + oppName)
            sleep(0.5)
            opportunitiesDetails["NAME"] = oppName
            # Utils.setDataInXlsx(self, wb, ws1, oppName, "Name", productType, opportunitiesDetailsFileName, rowNum)

        if "Stage" in fieldsData:
            Drivers.driverWait.until(
                EC.visibility_of_element_located((By.XPATH, "//select[@style='StageName']")))
            xPath = "//select[@style='StageName']"
            fieldType = Select(
                Drivers.driver.find_element_by_xpath(xPath))
            fieldType.select_by_value(fieldsData["Stage"])
            Messages.write_message("Stage Selected: " + fieldsData["Stage"])
            sleep(0.5)
            opportunitiesDetails["STAGE"] = fieldsData["Stage"]
            # Utils.setDataInXlsx(self, wb, ws1, fieldsData["Stage"], "Stage", productType, opportunitiesDetailsFileName, rowNum)

        if "Probability" in fieldsData:
            Drivers.driverWait.until(
                EC.visibility_of_element_located((By.XPATH, "//select[@style='Probability__c']")))
            xPath = "//select[@style='Probability__c']"
            fieldType = Select(
                Drivers.driver.find_element_by_xpath(xPath))
            fieldType.select_by_value(fieldsData["Probability"])
            Messages.write_message(
                "Probability Selected: " + fieldsData["Probability"])
            sleep(0.5)
            opportunitiesDetails["PROBABILITY"] = fieldsData["Probability"]
            # Utils.setDataInXlsx(self, wb, ws1, fieldsData["Probability"], "Probability", productType, opportunitiesDetailsFileName, rowNum)

        if "Insert Date/ Start Date" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='Insert_Date__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='Insert_Date__c']")
            fieldType.send_keys(currentDate)
            Messages.write_message(
                "Insert Date/ Start Date Entered: " + currentDate)
            sleep(0.5)
            opportunitiesDetails["INSERT DATE START DATE"] = currentDate
            # Utils.setDataInXlsx(self, wb, ws1, currentDate, "Insert Date/ Start Date", productType, opportunitiesDetailsFileName, rowNum)

        if "Execution End Date" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='End_Date__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='End_Date__c']")
            fieldType.send_keys(currentDate)
            Messages.write_message(
                "Execution End Date Entered: " + currentDate)
            sleep(0.5)
            opportunitiesDetails["EXECUTION END DATE"] = currentDate
            # Utils.setDataInXlsx(self, wb, ws1, currentDate, "Execution End Date", productType, opportunitiesDetailsFileName, rowNum)

        if "Expected close date" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='CloseDate']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='CloseDate']")
            if productType == "FSI US" or productType == "FSI CA":
                currentDate = Utils.getNAPIDate()
            fieldType.send_keys(currentDate)
            Messages.write_message(
                "Expected close Date Entered: " + currentDate)
            sleep(0.5)
            opportunitiesDetails["EXPECTED CLOSE DATE"] = currentDate
            # Utils.setDataInXlsx(self, wb, ws1, currentDate, "Expected close date", productType, opportunitiesDetailsFileName, rowNum)

        if "Contract Date" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='Contract_Date__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='Contract_Date__c']")
            fieldType.send_keys(currentDate)
            Messages.write_message("Contract Date Entered: " + currentDate)
            sleep(0.5)
            opportunitiesDetails["CONTRACT DATE"] = currentDate
            # Utils.setDataInXlsx(self, wb, ws1, currentDate, "Contract Date", productType, opportunitiesDetailsFileName, rowNum)

        if "Buisness Type" in fieldsData:
            Drivers.driverWait.until(
                EC.visibility_of_element_located((By.XPATH, "//select[@style='Business_Type__c']")))
            xPath = "//select[@style='Business_Type__c']"
            fieldType = Select(
                Drivers.driver.find_element_by_xpath(xPath))
            fieldType.select_by_value(fieldsData["Buisness Type"])
            Messages.write_message(
                "Buisness Type Selected: " + fieldsData["Buisness Type"])
            sleep(0.5)
            opportunitiesDetails["BUSINESS TYPE"] = fieldsData["Buisness Type"]
            # Utils.setDataInXlsx(self, wb, ws1, fieldsData["Buisness Type"], "Buisness Type", productType, opportunitiesDetailsFileName, rowNum)

        if "ILoc Asset Due Date" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='ILoc_Assest_Due_Date__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='ILoc_Assest_Due_Date__c']")
            fieldType.send_keys(currentDate)
            Messages.write_message(
                "ILoc Asset Due Date Entered: " + currentDate)
            sleep(0.5)
            opportunitiesDetails["ILOC ASSET DUE DATE"] = currentDate
            # Utils.setDataInXlsx(self, wb, ws1, fieldsData["ILoc Asset Due Date"], "ILoc Asset Due Date", productType, opportunitiesDetailsFileName, rowNum)

        if "Artwork Due Date" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='Artwork_Due_Date__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='Artwork_Due_Date__c']")
            fieldType.send_keys(currentDate)
            Messages.write_message(
                "Artwork Due Date Entered: " + currentDate)
            sleep(0.5)
            opportunitiesDetails["ARTWORK DUE DATE"] = currentDate
            # Utils.setDataInXlsx(self, wb, ws1, currentDate, "Artwork Due Date", productType, opportunitiesDetailsFileName, rowNum)

        if "Client List Due Date" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='ClientListDueDate__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='ClientListDueDate__c']")
            fieldType.send_keys(currentDate)
            Messages.write_message(
                "Client List Due Date Entered: " + currentDate)
            sleep(0.5)
            opportunitiesDetails["CLIENT LIST DUE DATE"] = currentDate
            # Utils.setDataInXlsx(self, wb, ws1, currentDate, "Client List Due Date", productType, opportunitiesDetailsFileName, rowNum)

        if "Billed based on Actual Execution" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='Billed_based_on_Actual_Execution__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='Billed_based_on_Actual_Execution__c']")
            if fieldsData["Billed based on Actual Execution"] == "Checked":
                fieldType.click()
                Messages.write_message(
                    "Billed based on Actual Execution Selected: " + fieldsData["Billed based on Actual Execution"])
            sleep(0.5)
            opportunitiesDetails["BILLED BASED ON ACTUAL EXECUTION"] = fieldsData["Billed based on Actual Execution"]
            # Utils.setDataInXlsx(self, wb, ws1, fieldsData["Billed based on Actual Execution"], "Billed based on Actual Execution", productType, opportunitiesDetailsFileName, rowNum)

        if "Freedom" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='Freedom_Enabled__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='Freedom_Enabled__c']")
            if fieldsData["Freedom"] == "Checked":
                fieldType.click()
                Messages.write_message(
                    "Freedom Selected: " + fieldsData["Freedom"])
            sleep(0.5)
            opportunitiesDetails["FREEDOM"] = fieldsData["Freedom"]
            # Utils.setDataInXlsx(self, wb, ws1, fieldsData["Freedom Enabled"], "Freedom Enabled", productType, opportunitiesDetailsFileName, rowNum)

        if "Account ID" in fieldsData:
            accountName = ''
            if fieldsData["Account ID"] == "FROM PROPERTY":
                if "US" in productType:
                    accountName = os.getenv("US_SALES_REPRESENTATIVE_ACCOUNT")
                elif "CA" in productType:
                    accountName = os.getenv("CA_SALES_REPRESENTATIVE_ACCOUNT")
            else:
                accountName = fieldsData["Account ID"]

            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//img[@title='Account Name Lookup (New Window)']")))

            fieldType = Drivers.driver.find_element_by_xpath(
                "//img[@title='Account Name Lookup (New Window)']")

            fieldType.click()
            sleep(0.2)

            Drivers.driver.switch_to.window(Drivers.driver.window_handles[1])
            sleep(0.2)

            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.ID, "searchFrame")))
            sleep(0.2)

            Drivers.driver.switch_to.frame("searchFrame")
            # Drivers.driver.switch_to_frame("searchFrame")
            sleep(0.2)

            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.ID, "lksrch")))
            sleep(0.2)
            txtBoxSearch = Drivers.driver.find_element_by_id("lksrch")
            txtBoxSearch.send_keys(accountName)
            sleep(0.2)

            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.NAME, "go")))
            sleep(0.2)
            buttonSearch = Drivers.driver.find_element_by_name("go")
            buttonSearch.click()
            sleep(0.2)

            Drivers.driver.switch_to.window(Drivers.driver.window_handles[1])
            sleep(0.2)

            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.ID, "resultsFrame")))
            sleep(0.2)

            Drivers.driver.switch_to.frame("resultsFrame")
            sleep(0.2)

            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.LINK_TEXT, accountName)))
            sleep(0.2)

            linkAccountName = Drivers.driver.find_element_by_link_text(
                accountName)
            linkAccountName.click()
            sleep(0.2)

            Drivers.driver.switch_to.window(Drivers.driver.window_handles[0])
            sleep(3)
            opportunitiesDetails["ACCOUNT ID"] = accountName
            # Utils.setDataInXlsx(self, wb, ws1, accountName, "Account ID", productType, opportunitiesDetailsFileName, rowNum)

        # SSD FIELDS
        if "ILoc Format" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='ILoc_Format__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='ILoc_Format__c']")
            fieldType.send_keys(fieldsData["ILoc Format"])
            Messages.write_message(
                "ILoc Format Entered: " + fieldsData["ILoc Format"])
            sleep(0.5)
            opportunitiesDetails["ILOC FORMAT"] = fieldsData["ILoc Format"]
            # Utils.setDataInXlsx(self, wb, ws1, fieldsData["ILoc Format"], "ILoc Format", productType, opportunitiesDetailsFileName, rowNum)

        if "ILoc Targeting" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='ILoc_Targeting__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='ILoc_Targeting__c']")
            fieldType.send_keys(fieldsData["ILoc Targeting"])
            Messages.write_message(
                "ILoc Targeting Entered: " + fieldsData["ILoc Targeting"])
            sleep(0.5)
            opportunitiesDetails["ILOC TARGETING"] = fieldsData["ILoc Targeting"]
            # Utils.setDataInXlsx(self, wb, ws1, fieldsData["ILoc Targeting"] , "ILoc Targeting", productType, opportunitiesDetailsFileName, rowNum)

        if "ILoc Artwork Due Date" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='ILocArtworkDueDate__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='ILocArtworkDueDate__c']")
            fieldType.send_keys(currentDate)
            Messages.write_message(
                "ILoc Artwork Due Date: " + currentDate)
            sleep(0.5)
            opportunitiesDetails["ILOC ARTWORK DUE DATE"] = currentDate
            # Utils.setDataInXlsx(self, wb, ws1, currentDate, "ILoc Artwork Due Date", productType, opportunitiesDetailsFileName, rowNum)

        if "ILoc Creative Due Date" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='ILoc_Creative_Due_Date__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='ILoc_Creative_Due_Date__c']")
            fieldType.send_keys(currentDate)
            Messages.write_message(
                "ILoc Creative Due Date Entered: " + currentDate)
            sleep(0.5)
            opportunitiesDetails["ILOC CREATIVE DUE DATE"] = currentDate
            # Utils.setDataInXlsx(self, wb, ws1, currentDate, "ILoc Creative Due Date", productType, opportunitiesDetailsFileName, rowNum)

        if "ILoc Sample Due Date" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='ILoc_Sample_Due_Date__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='ILoc_Sample_Due_Date__c']")
            fieldType.send_keys(currentDate)
            Messages.write_message(
                "ILoc Sample Due Date Entered: " + currentDate)
            sleep(0.5)
            opportunitiesDetails["ILOC SAMPLE DUE DATE"] = currentDate
            # Utils.setDataInXlsx(self, wb, ws1, currentDate, "ILoc Sample Due Date", productType, opportunitiesDetailsFileName, rowNum)

        # MERC FIELDS
        if "Opportunity Type" in fieldsData:
            Drivers.driverWait.until(
                EC.visibility_of_element_located((By.XPATH, "//select[@style='Type']")))
            xPath = "//select[@style='Type']"
            optionsXpath = "//select[@style='Type']/option"
            dropDownProductLine = Select(
                Drivers.driver.find_element_by_xpath(xPath))
            dropDownOptionCnt = str(
                len(Drivers.driver.find_elements_by_xpath(optionsXpath)))
            if fieldsData["Opportunity Type"] == "RANDOM":
                dropDownProductLine.select_by_index(
                    randint(2, (int(dropDownOptionCnt) - 1)))
            else:
                dropDownProductLine.select_by_value(
                    fieldsData["Opportunity Type"])
            selectedPartner = dropDownProductLine.first_selected_option.text
            Messages.write_message(
                "Opportunity Type Selected: " + selectedPartner)
            sleep(0.5)
            opportunitiesDetails["OPPORTUNITY TYPE"] = selectedPartner
            # Utils.setDataInXlsx(self, wb, ws1, selectedPartner, "Opportunity Type", productType, opportunitiesDetailsFileName, rowNum)

        if "Material Due To" in fieldsData:
            Drivers.driverWait.until(
                EC.visibility_of_element_located((By.XPATH, "//select[@style='Client_Supplied_Materials_Due_To__c']")))
            xPath = "//select[@style='Client_Supplied_Materials_Due_To__c']"
            optionsXpath = "//select[@style='Client_Supplied_Materials_Due_To__c']/option"
            dropDownProductLine = Select(
                Drivers.driver.find_element_by_xpath(xPath))
            dropDownOptionCnt = str(
                len(Drivers.driver.find_elements_by_xpath(optionsXpath)))
            if fieldsData["Material Due To"] == "RANDOM":
                dropDownProductLine.select_by_index(
                    randint(2, (int(dropDownOptionCnt) - 1)))
            else:
                dropDownProductLine.select_by_value(
                    fieldsData["Material Due To"])
            selectedPartner = dropDownProductLine.first_selected_option.text
            Messages.write_message(
                "Material Due To Selected: " + selectedPartner)
            sleep(0.5)
            opportunitiesDetails["MATERIAL DUE TO"] = selectedPartner
            # Utils.setDataInXlsx(self, wb, ws1, selectedPartner, "Material Due To", productType, opportunitiesDetailsFileName, rowNum)

        if "POS Materials Produced" in fieldsData:
            Drivers.driverWait.until(
                EC.visibility_of_element_located((By.XPATH, "//select[@style='POS_Materials_Produced_If_yes_above__c']")))
            xPath = "//select[@style='POS_Materials_Produced_If_yes_above__c']"
            optionsXpath = "//select[@style='POS_Materials_Produced_If_yes_above__c']/option"
            dropDownProductLine = Select(
                Drivers.driver.find_element_by_xpath(xPath))
            dropDownOptionCnt = str(
                len(Drivers.driver.find_elements_by_xpath(optionsXpath)))
            if fieldsData["POS Materials Produced"] == "RANDOM":
                dropDownProductLine.select_by_index(
                    randint(2, (int(dropDownOptionCnt) - 1)))
            else:
                dropDownProductLine.select_by_value(
                    fieldsData["POS Materials Produced"])
            selectedPartner = dropDownProductLine.first_selected_option.text
            Messages.write_message(
                "POS Materials Produced Selected: " + selectedPartner)
            sleep(0.5)
            opportunitiesDetails["POS MATERIALS PRODUCED"] = selectedPartner
            # Utils.setDataInXlsx(self, wb, ws1, selectedPartner, "POS Materials Produced", productType, opportunitiesDetailsFileName, rowNum)

        if "POS Material Disposition" in fieldsData:
            Drivers.driverWait.until(
                EC.visibility_of_element_located((By.XPATH, "//select[@style='POS_Material_Disposition__c']")))
            xPath = "//select[@style='POS_Material_Disposition__c']"
            optionsXpath = "//select[@style='POS_Material_Disposition__c']/option"
            dropDownProductLine = Select(
                Drivers.driver.find_element_by_xpath(xPath))
            dropDownOptionCnt = str(
                len(Drivers.driver.find_elements_by_xpath(optionsXpath)))
            if fieldsData["POS Material Disposition"] == "RANDOM":
                dropDownProductLine.select_by_index(
                    randint(2, (int(dropDownOptionCnt) - 1)))
            else:
                dropDownProductLine.select_by_value(
                    fieldsData["POS Material Disposition"])
            selectedPartner = dropDownProductLine.first_selected_option.text
            Messages.write_message(
                "POS Material Disposition Selected: " + selectedPartner)
            sleep(0.5)
            opportunitiesDetails["POS MATERIAL DISPOSITION"] = selectedPartner
            # Utils.setDataInXlsx(self, wb, ws1, selectedPartner, "POS Material Disposition", productType, opportunitiesDetailsFileName, rowNum)

        if "Image Request Details" in fieldsData:
            Drivers.driverWait.until(
                EC.visibility_of_element_located((By.XPATH, "//select[@style='Image_Request_Details__c']")))
            xPath = "//select[@style='Image_Request_Details__c']"
            optionsXpath = "//select[@style='Image_Request_Details__c']/option"
            dropDownProductLine = Select(
                Drivers.driver.find_element_by_xpath(xPath))
            dropDownOptionCnt = str(
                len(Drivers.driver.find_elements_by_xpath(optionsXpath)))
            if fieldsData["Image Request Details"] == "RANDOM":
                dropDownProductLine.select_by_index(
                    randint(2, (int(dropDownOptionCnt) - 1)))
            else:
                dropDownProductLine.select_by_value(
                    fieldsData["Image Request Details"])
            selectedPartner = dropDownProductLine.first_selected_option.text
            Messages.write_message(
                "Image Request Details Selected: " + selectedPartner)
            sleep(0.5)
            opportunitiesDetails["IMAGE REQUEST DEATAILS"] = selectedPartner
            # Utils.setDataInXlsx(self, wb, ws1, selectedPartner, "Image Request Details", productType, opportunitiesDetailsFileName, rowNum)

        if "Image Subset List Count" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='Subset_List_Count__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='Subset_List_Count__c']")
            if fieldsData["Image Subset List Count"] == "RANDOM":
                imageSubsetListCount = randint(1, 100)
            else:
                imageSubsetListCount = fieldsData["Image Subset List Count"]
            fieldType.send_keys(str(imageSubsetListCount))
            Messages.write_message(
                "Image Subset List Count: " + str(imageSubsetListCount))
            sleep(0.5)
            opportunitiesDetails["IMAGE SUBSET LIST COUNT"] = imageSubsetListCount
            # Utils.setDataInXlsx(self, wb, ws1, imageSubsetListCount, "Image Subset List Count", productType, opportunitiesDetailsFileName, rowNum)

        if "Network Retailer List" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//textarea[@style='Network_Retailer_List__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//textarea[@style='Network_Retailer_List__c']")
            fieldType.send_keys(fieldsData["Network Retailer List"])
            Messages.write_message(
                "Network Retailer List: " + fieldsData["Network Retailer List"])
            sleep(0.5)
            opportunitiesDetails["NETWORK RETAILER LIST"] = fieldsData["Network Retailer List"]
            # Utils.setDataInXlsx(self, wb, ws1, fieldsData["Network Retailer List"], "Network Retailer List", productType, opportunitiesDetailsFileName, rowNum)

        if "Non Network Retailer List" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//textarea[@style='Non_Network_Retailer_List__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//textarea[@style='Non_Network_Retailer_List__c']")
            fieldType.send_keys(fieldsData["Non Network Retailer List"])
            Messages.write_message(
                "Non Network Retailer List: " + fieldsData["Non Network Retailer List"])
            sleep(0.5)
            opportunitiesDetails["NETWORK RETAILER LIST"] = fieldsData["Non Network Retailer List"]
            # Utils.setDataInXlsx(self, wb, ws1, fieldsData["Non Network Retailer List"], "Non Network Retailer List", productType, opportunitiesDetailsFileName, rowNum)

        if "Job Description/Comments" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//textarea[@style='Job_Description_Comments__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//textarea[@style='Job_Description_Comments__c']")
            fieldType.send_keys(fieldsData["Job Description/Comments"])
            Messages.write_message(
                "Job Description/Comments: " + fieldsData["Job Description/Comments"])
            sleep(0.5)
            opportunitiesDetails["JOB DESCRIPTION/COMMENTS"] = fieldsData["Job Description/Comments"]
            # Utils.setDataInXlsx(self, wb, ws1, fieldsData["Job Description/Comments"], "Job Description/Comments", productType, opportunitiesDetailsFileName, rowNum)

        if "ILoc Signed LOC/Store List Due Date" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='ILocSignedLOCORStoreListDueDt__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='ILocSignedLOCORStoreListDueDt__c']")
            fieldType.send_keys(currentDate)
            Messages.write_message(
                "ILoc Signed LOC/Store List Due Date Entered: " + currentDate)
            sleep(0.5)
            opportunitiesDetails["ILOC SIGNED LOC/STORE LIST DUE DATE"] = currentDate
            # Utils.setDataInXlsx(self, wb, ws1, currentDate, "ILoc Signed LOC/Store List Due Date", productType, opportunitiesDetailsFileName, rowNum)

        if "ILoc Program Document Due Date" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='ILocProgramDocumentDueDate__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='ILocProgramDocumentDueDate__c']")
            fieldType.send_keys(currentDate)
            Messages.write_message(
                "ILoc Program Document Due Date Entered: " + currentDate)
            sleep(0.5)
            opportunitiesDetails["ILOC PROGRAM DOCUMENT DUE DATE"] = currentDate
            # Utils.setDataInXlsx(self, wb, ws1, currentDate, "ILoc Program Document Due Date", productType, opportunitiesDetailsFileName, rowNum)

        if "ILoc Material Due Date" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='ILoc_Material_Due_Date__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='ILoc_Material_Due_Date__c']")
            fieldType.send_keys(currentDate)
            Messages.write_message(
                "ILoc Material Due Date Entered: " + currentDate)
            sleep(0.5)
            opportunitiesDetails["ILOC MATERIAL DUE DATE"] = currentDate
            # Utils.setDataInXlsx(self, wb, ws1, currentDate , "ILoc Material Due Date", productType, opportunitiesDetailsFileName, rowNum)
        
        # INSTORE FIELDS
        if "InStore Cycle" in fieldsData:
            inStoreCycle = ""
            current_month = datetime.now().strftime('%m')
            current_year_full = datetime.now().strftime('%Y')
            # current_day = datetime.now().strftime('%d')
            # if (int(current_month) < 10):
            #     current_month = "0" + str(int(current_month)+2)

            currentMonth = int(current_month)
            currentYear = int(self.get_Current_Fiscal_Year())
            if (currentMonth + 2) > 13:
                currentYear = currentYear + 1
                currentMonth = "01"
                inStoreCycle = f"{currentYear}{currentMonth}"
            else:
                currentMonth = currentMonth + 2
            if currentMonth <= 9:
                currentMonth = "0" + str(currentMonth)
                inStoreCycle = f"{currentYear}{currentMonth}"
            else:
                inStoreCycle = f"{currentYear}{currentMonth}"
            print(f"Instore Cycle {currentYear}{currentMonth}")
            Messages.write_message(f"Instore Cycle {currentYear}{currentMonth}")
                        
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//img[@title='InStore Cycle Lookup (New Window)']")))

            fieldType = Drivers.driver.find_element_by_xpath(
                "//img[@title='InStore Cycle Lookup (New Window)']")            
            fieldType.click()
            sleep(0.5)

            Drivers.driver.switch_to.window(Drivers.driver.window_handles[-1])
            sleep(0.5)

            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.ID, "searchFrame")))
            sleep(0.5)

            Drivers.driver.switch_to.frame("searchFrame")
            sleep(0.5)

            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.ID, "lksrch")))
            sleep(0.2)
            txtBoxSearch = Drivers.driver.find_element_by_id("lksrch")
            txtBoxSearch.send_keys(inStoreCycle)
            sleep(0.2)

            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.NAME, "go")))
            sleep(0.2)
            buttonSearch = Drivers.driver.find_element_by_name("go")
            buttonSearch.click()
            sleep(0.2)

            Drivers.driver.switch_to.window(Drivers.driver.window_handles[-1])
            sleep(0.2)

            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.ID, "resultsFrame")))
            sleep(0.2)

            Drivers.driver.switch_to.frame("resultsFrame")
            sleep(0.2)

            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.LINK_TEXT, inStoreCycle)))
            sleep(0.2)

            linkAccountName = Drivers.driver.find_element_by_link_text(
                inStoreCycle)
            linkAccountName.click()
            sleep(0.2)

            Drivers.driver.switch_to.window(Drivers.driver.window_handles[-1])
            sleep(3)
            opportunitiesDetails["INSTORE CYCLE"] = inStoreCycle
            # Utils.setDataInXlsx(self, wb, ws1, inStoreCycle, "InStore Cycle", productType, opportunitiesDetailsFileName, rowNum)

        if "Expiration Date" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='Expiration_Date__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='Expiration_Date__c']")
            if productType == "FSI US" or productType == "FSI CA":
                currentDate = Utils.getNAPIDate()
                print(currentDate)
            fieldType.send_keys(currentDate)
            Messages.write_message("Expiration Date: " + currentDate)
            sleep(0.5)
            opportunitiesDetails["EXPIRATION DATE"] = currentDate
            # Utils.setDataInXlsx(self, wb, ws1, currentDate, "Expiration Date", productType, opportunitiesDetailsFileName, rowNum)

        if "Projected Average CPS(Space)" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='Estimated_Average_CPS__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='Estimated_Average_CPS__c']")
            if fieldsData["Projected Average CPS(Space)"] == "RANDOM":
                projectedAvgSpace = randint(1, 100)
            else:
                projectedAvgSpace = fieldsData["Projected Average CPS(Space)"]
            fieldType.send_keys(str(projectedAvgSpace))
            Messages.write_message(
                "Projected Average CPS(Space): " + str(projectedAvgSpace))
            sleep(0.5)
            opportunitiesDetails["PROJECTED AVERAGE CPS(SPACE)"] = projectedAvgSpace
            # Utils.setDataInXlsx(self, wb, ws1, projectedAvgSpace, "Projected Average CPS(Space)", productType, opportunitiesDetailsFileName, rowNum)

        if "Projected Store Count" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='Estimated_Store_Count__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='Estimated_Store_Count__c']")
            if fieldsData["Projected Store Count"] == "RANDOM":
                projectedStoreCount = randint(1, 100)
            else:
                projectedStoreCount = fieldsData["Projected Store Count"]
            fieldType.send_keys(str(projectedStoreCount))
            Messages.write_message(
                "Projected Store Count: " + str(projectedStoreCount))
            sleep(0.5)
            opportunitiesDetails["PROJECTED STORE COUNT"] = projectedStoreCount
            # Utils.setDataInXlsx(self, wb, ws1, projectedStoreCount, "Projected Store Count", productType, opportunitiesDetailsFileName, rowNum)

        # FSI Fields
        if "Projected Average CPM" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='Estimated_Average_CPM__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='Estimated_Average_CPM__c']")
            if fieldsData["Projected Average CPM"] == "RANDOM":
                projectedAvgCPM = randint(1, 100)
            else:
                projectedAvgCPM = fieldsData["Projected Average CPM"]
            fieldType.send_keys(str(projectedAvgCPM))
            Messages.write_message(
                "Projected Average CPM: " + str(projectedAvgCPM))
            sleep(0.5)
            opportunitiesDetails["PROJECTED AVERAGE CPM"] = projectedAvgCPM
            # Utils.setDataInXlsx(self, wb, ws1, projectedAvgCPM, "Projected Average CPM", productType, opportunitiesDetailsFileName, rowNum)

        if "Projected Circulation (000)" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='Expected_Circulation__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='Expected_Circulation__c']")
            if fieldsData["Projected Circulation (000)"] == "RANDOM":
                projectedCirculation = randint(1, 100)
            else:
                projectedCirculation = fieldsData["Projected Circulation (000)"]
            fieldType.send_keys(str(projectedCirculation))
            Messages.write_message(
                "Projected Circulation (000): " + str(projectedCirculation))
            sleep(0.5)
            opportunitiesDetails["PROJECTED CIRCULATION (000)"] = projectedCirculation
            # Utils.setDataInXlsx(self, wb, ws1, projectedCirculation, "Projected Circulation (000)", productType, opportunitiesDetailsFileName, rowNum)

        if "FSI - Insert Date/Start Date" in fieldsData:
            # napiDate = Utils.getNAPIDate()
            print(napiDate)
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//img[@title='FSI - Insert Date/Start Date Lookup (New Window)']")))

            fieldType = Drivers.driver.find_element_by_xpath(
                "//img[@title='FSI - Insert Date/Start Date Lookup (New Window)']")

            handle_first = Drivers.driver.window_handles[0]
            fieldType.click()
            sleep(0.2)

            Drivers.driver.switch_to.window(Drivers.driver.window_handles[1])
            sleep(0.2)

            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.ID, "searchFrame")))
            sleep(0.2)

            Drivers.driver.switch_to.frame("searchFrame")
            sleep(0.2)

            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.ID, "lksrch")))
            sleep(0.2)
            txtBoxSearch = Drivers.driver.find_element_by_id("lksrch")
            txtBoxSearch.send_keys(napiDate)
            sleep(0.2)

            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.NAME, "go")))
            sleep(0.2)
            buttonSearch = Drivers.driver.find_element_by_name("go")
            buttonSearch.click()
            sleep(0.2)

            Drivers.driver.switch_to.window(Drivers.driver.window_handles[1])
            sleep(0.2)

            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.ID, "resultsFrame")))
            sleep(0.2)

            Drivers.driver.switch_to.frame("resultsFrame")
            sleep(0.2)

#             pdb.set_trace()
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.LINK_TEXT, napiDate)))
            sleep(0.2)

            linkAccountName = Drivers.driver.find_element_by_link_text(
                napiDate)
            linkAccountName.click()
            sleep(0.2)

            Drivers.driver.switch_to.window(Drivers.driver.window_handles[0])
            sleep(3)
            opportunitiesDetails["INSTORE CYCLE"] = napiDate
#             #Utils.setDataInXlsx(self, wb, ws1, napiDate, "FSI - Insert Date/Start Date", productType, opportunitiesDetailsFileName, rowNum)

        # SSMG FIELDS
        if "ILoc Market List Due Date" in fieldsData:
            Drivers.driverWait.until(EC.visibility_of_element_located(
                (By.XPATH, "//input[@style='ILocMarketListDueDate__c']")))
            fieldType = Drivers.driver.find_element_by_xpath(
                "//input[@style='ILocMarketListDueDate__c']")
            fieldType.send_keys(currentDate)
            Messages.write_message("ILoc Market List Due Date: " + currentDate)
            sleep(0.5)
            opportunitiesDetails["ILOC MARKET LIST DUE DATE"] = currentDate
            # Utils.setDataInXlsx(self, wb, ws1, currentDate, "ILoc Market List Due Date", productType, opportunitiesDetailsFileName, rowNum)

#         wb.save(filename=opportunitiesDetailsFileName)
        print("\n", "Opportunity Detail: ")
        print(opportunitiesDetails)
        Messages.write_message(opportunitiesDetails)
        data_store.spec[productType] = opportunitiesDetails
#         column_count = ws1.max_column
#         row_count = ws1.max_row
#         Messages.write_message("Column Count: " + str(column_count))
#         Messages.write_message("Row Count: " + str(row_count))
#         opportunityDictionary[productType] = opportunitiesDetails
#         Utils.set_Opportunities_Details(
#             'OpportunityData.yaml', opportunityDictionary)

    @step("Select <productName> for <productLine> Opportunity")
    def select_product(self, productName, productLine):

        opportunitiesDetails = {}
        opportunitiesDetails = data_store.spec.get(productLine)
        Drivers.driverWait.until(
            EC.visibility_of_element_located((By.XPATH, "//input[@value='Next']")))
        productTable = Drivers.driver.find_element_by_xpath(
            "//table[contains(@class,'slds-table')]/tbody")
        tableRows = productTable.find_elements_by_tag_name("tr")
#         pdb.set_trace()
        if productName == "Product":
            cnt = randint(1, len(tableRows))
            productXpath = "//table[contains(@class,'slds-table')]/tbody/tr[" + str(
                cnt) + "]/td[2]"
            radioButtonXpath = "//table[contains(@class,'slds-table')]/tbody/tr[" + str(
                cnt) + "]/td[1]/span/input[@id='radio']"
            radioButton = Drivers.driver.find_element_by_xpath(
                radioButtonXpath)
            radioButton.click()
        else:
            inct = 1
            for row in tableRows:
                Messages.write_message(f"Product {inct}: {row.text}")
                # print("--", row.find_element_by_xpath("//td[@scope='row'][2]").text.strip(), "--")
                productXpath = "//table[contains(@class,'slds-table')]/tbody/tr[" + str(inct) + "]/td[2]"                
                product = Drivers.driver.find_element_by_xpath(productXpath).text        
                print(product)
                if str(productName).upper() == str(product).upper():
                    productXpath = "//table[contains(@class,'slds-table')]/tbody/tr[" + str(
                        inct) + "]/td[2]"
                    radioButtonXpath = "//table[contains(@class,'slds-table')]/tbody/tr[" + str(
                        inct) + "]/td[1]/span/input[@id='radio']"
                    radioButton = Drivers.driver.find_element_by_xpath(
                        radioButtonXpath)
                    radioButton.click()
                    # row.find_element_by_xpath(
                    #     "//child::td[1]/span/input").click()
                    break
                inct = inct + 1
        productName = Drivers.driver.find_element_by_xpath(productXpath).text

        opportunitiesDetails.update({"PRODUCT NAME": productName})
        print("\n", "Product Detail: ")
        print(opportunitiesDetails)

        Messages.write_message("Selected Product: " + productName)
#         productDetails["PRODUCT NAME"] = productName
#         Utils.set_Opportunities_Details('OpportunityData.yaml', productDetails)
        data_store.spec[productLine] = opportunitiesDetails
        
        xPath = f"//input[contains(@value,'{buttonName}')]"
        buttonToClick = Drivers.driver.find_element_by_xpath(xPath)
        buttonToClick.click()
        sleep(5)

#     @step("Select <productName> from Product Details page")
#     def select_product(self, productName):
# #         productDetails = {}
#         Drivers.driverWait.until(
#             EC.visibility_of_element_located((By.XPATH, "//input[@value='Next']")))
#         productTable = Drivers.driver.find_element_by_xpath(
#             "//table[contains(@class,'slds-table')]/tbody")
#         tableRows = productTable.find_elements_by_tag_name("tr")
#
#         if productName == "Product":
#             cnt = randint(1, len(tableRows))
#             productXpath = "//table[contains(@class,'slds-table')]/tbody/tr[" + str(
#                 cnt) + "]/td[2]"
#             radioButtonXpath = "//table[contains(@class,'slds-table')]/tbody/tr[" + str(
#                 cnt) + "]/td[1]/span/input[@id='radio']"
#             radioButton = Drivers.driver.find_element_by_xpath(
#                 radioButtonXpath)
#             radioButton.click()
#         else:
#             inct = 1
#             for row in tableRows:
#                 print(row.text)
#                 # print("--", row.find_element_by_xpath("//td[@scope='row'][2]").text.strip(), "--")
#                 if productName in row.text:
#                     productXpath = "//table[contains(@class,'slds-table')]/tbody/tr[" + str(
#                     inct) + "]/td[2]"
#                     radioButtonXpath = "//table[contains(@class,'slds-table')]/tbody/tr[" + str(
#                         inct) + "]/td[1]/span/input[@id='radio']"
#                     radioButton = Drivers.driver.find_element_by_xpath(
#                         radioButtonXpath)
#                     radioButton.click()
#                     # row.find_element_by_xpath(
#                     #     "//child::td[1]/span/input").click()
#                     break
#                 inct = inct + 1
#         productName = Drivers.driver.find_element_by_xpath(productXpath).text
#
#         Utils.setExcelColumnHeaders(self, productType, "Product Name", opportunitiesDetailsFileName, ws1, wb)
#
#         Messages.write_message("Selected Product: " + productName)
# #         productDetails["PRODUCT NAME"] = productName
# #         Utils.set_Opportunities_Details('OpportunityData.yaml', productDetails)

    @step("Add <lineItemCount> Charge Description for <productName> Product on Product Details page")
    def enter_pricing_details(self, lineItemCount, productName):
        Drivers.driverWait.until(
            EC.visibility_of_element_located((By.ID, "SelectedProductType")))
#         if productName == "CO51 FREEDOM US" or productName == "CO51 FREEDOM CA" or productName == "CO51 NON FREEDOM US" or productName == "CO51 NON FREEDOM CA":
        if productName == "CO51_US" or productName == "CO51_CA":
            Common_Steps.CommonSteps.pricing_details(self,
                lineItemCount, productName)

        if productName == "DIGITAL_US" or productName == "DIGITAL_CA":
            Common_Steps.CommonSteps.pricing_details(self,
                lineItemCount, productName)

        if productName == "SMARTSOURCE_DIRECT_US" or productName == "SMARTSOURCE_DIRECT_CA":
            Common_Steps.CommonSteps.pricing_details(self,
                lineItemCount, productName)

        if productName == "MERCHANDISING_US" or productName == "MERCHANDISING_CA":
            Common_Steps.CommonSteps.pricing_details(self,
                lineItemCount, productName)

        if productName == "INSTORE_US" or productName == "INSTORE_CA":
            Common_Steps.CommonSteps.pricing_details(self,
                lineItemCount, productName)

        if productName == "FSI_US" or productName == "FSI_CA":
            Common_Steps.CommonSteps.pricing_details(self,
                lineItemCount, productName)

        if productName == "SSMG_US" or productName == "SSMG_CA":
            Common_Steps.CommonSteps.pricing_details(self,
                lineItemCount, productName)

#     @step("Add <lineItemCount> Charge Description for <productName> Product on Product Details page")
#     def enter_pricing_details(self, lineItemCount, productName):
#         Drivers.driverWait.until(
#             EC.visibility_of_element_located((By.ID, "SelectedProductType")))
# #         if productName == "CO51 FREEDOM US" or productName == "CO51 FREEDOM CA" or productName == "CO51 NON FREEDOM US" or productName == "CO51 NON FREEDOM CA":
#         if productName == "CO51 US" or productName == "CO51 CA":
#             Common_Steps.CommonSteps.pricing_details(
#                 self, lineItemCount, productName)
#
#         if productName == "DIGITAL US" or productName == "DIGITAL CA":
#             Common_Steps.CommonSteps.pricing_details(
#                 self, lineItemCount, productName)
#
#         if productName == "SMARTSOURCE DIRECT US" or productName == "SMARTSOURCE DIRECT CA":
#             Common_Steps.CommonSteps.pricing_details(
#                 self, lineItemCount, productName)
#
#         if productName == "MERCHANDISING US" or productName == "MERCHANDISING CA":
#             Common_Steps.CommonSteps.pricing_details(
#                 self, lineItemCount, productName)
#
#         if productName == "INSTORE US" or productName == "INSTORE CA":
#             Common_Steps.CommonSteps.pricing_details(
#                 self, lineItemCount, productName)
#
#         if productName == "FSI US" or productName == "FSI CA" or productName == "FSI US REMNANT" or productName == "FSI CA REMNANT":
#             Common_Steps.CommonSteps.pricing_details(
#                 self, lineItemCount, productName)
#
#         if productName == "SSMG US" or productName == "SSMG CA":
#             Common_Steps.CommonSteps.pricing_details(
#                 self, lineItemCount, productName)

    @step("Add Charge Description for <productName> Product on Product Details page <table>")
    def enter_co51_pricing_details(self, productName, table):
        # Drivers.driverWait.until(
        #     EC.visibility_of_element_located((By.ID, "SelectedProductType")))
        Common_Steps.CommonSteps.pricing_details_table(self,
            productName, table)

    @step("Update <productName> Opportunity Fields Data <table>")
    def update_opportunity_fields_data(self, productName, table):
        sleep(5)
        opportunitiesDetails = {}
        oppDetails = {}
        oppDetails = data_store.spec.get(productName)

        oppName = oppDetails["NAME"]
        soql = f"SELECT Id FROM Opportunity WHERE Name = '{oppName}'"
        queryResult = Drivers.sf.query_all(query=soql)
        recDetails = queryResult['records']
        Messages.write_message("Opportunity:" + str(recDetails[0]["Id"]))

#         pdb.set_trace()
        fieldsData = {}
        dt = datetime.today()
        print(dt.month, " ", dt.day, " ", dt.year)
        year = int(dt.year) + 2
        currentDate = f"{str(dt.month)}/{str(dt.day)}/{str(year)}"

        insertDate = currentDate
        insert_date_time_obj = datetime.strptime(insertDate, '%m/%d/%Y')
        insDate = insert_date_time_obj.date()

        rows = table.rows
        for row in rows:
            row0 = str(row[0]).strip()
            row1 = str(row[1]).strip()
            # FSI FIELDS
            if row0 == "ILOC Ad Description":
                fieldsData["ILOC Ad Description"] = row1
            elif row0 == "ILoc Circulation Charges":
                fieldsData["ILoc Circulation Charges"] = row1
            elif row0 == "ILOC Production Charges":
                fieldsData["ILOC Production Charges"] = row1
            elif row0 == "ILoc Other Charges":
                fieldsData["ILoc Other Charges"] = row1
            elif row0 == "ILOC Total Program Fee":
                fieldsData["ILOC Total Program Fee"] = row1
            elif row0 == "ILoc Artwork Due Date":
                fieldsData["ILoc Artwork Due Date"] = row1
            elif row0 == "ILoc Market List Due Date":
                fieldsData["ILoc Market List Due Date"] = row1
            elif row0 == "ILoc Material Due Date":
                fieldsData["ILoc Material Due Date"] = row1
            elif row0 == "ILoc Category":
                fieldsData["ILoc Category"] = row1
            elif row0 == "Status":
                fieldsData["Status"] = row1
            elif row0 == "ILOC Remnant Estimated Cost":
                fieldsData["ILOC Remnant Estimated Cost"] = row1
            elif row0 == "ILOC Remnant Insert Date":
                fieldsData["ILOC Remnant Insert Date"] = row1
            elif row0 == "ILOC Total Remnant Charge Detail":
                fieldsData["ILOC Total Remnant Charge Detail"] = row1
            elif row0 == "Add Remnant Guidelines":
                fieldsData["Add Remnant Guidelines"] = row1
            elif row0 == "Billed based on Actual Execution":
                fieldsData["Billed based on Actual Execution"] = row1

            # IS FIELDS
            elif row0 == "ILoc Store Count":
                fieldsData["ILoc Store Count"] = row1
            elif row0 == "ILoc Trade Class":
                fieldsData["ILoc Trade Class"] = row1
            elif row0 == "ILoc Brand":
                fieldsData["ILoc Brand"] = row1
            elif row0 == "ILoc Loc Type":
                fieldsData["ILoc Loc Type"] = row1
            elif row0 == "ILoc Program":
                fieldsData["ILoc Program"] = row1
            elif row0 == "ILoc Geography":
                fieldsData["ILoc Geography"] = row1
            elif row0 == "ILoc Asset Due Date":
                fieldsData["ILoc Asset Due Date"] = row1

            # DG FIELDS
            elif row0 == "Artwork Due Date":
                fieldsData["Artwork Due Date"] = row1
            elif row0 == "Client List Due Date":
                fieldsData["Client List Due Date"] = row
            elif row0 == "ILOC Client processing fee":
                fieldsData["ILOC Client processing fee"] = row
            elif row0 == "ILOC Client clip rate":
                fieldsData["ILOC Client clip rate"] = row

            # SSD FIELDS
            elif row0 == "ILoc Format":
                fieldsData["ILoc Format"] = row1
            elif row0 == "ILoc Targeting":
                fieldsData["ILoc Targeting"] = row1
            elif row0 == "ILoc Sample Due Date":
                fieldsData["ILoc Sample Due Date"] = row1
            elif row0 == "ILoc Creative Due Date":
                fieldsData["ILoc Creative Due Date"] = row1

        # FSI FIELDS
        if "ILOC Ad Description" in fieldsData:
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILOCAdDescription__c': fieldsData["ILOC Ad Description"]})
            opportunitiesDetails["ILOC Ad Description".upper(
            )] = fieldsData["ILOC Ad Description"]

        if "ILoc Circulation Charges" in fieldsData:
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILocCirculationCharges__c': fieldsData["ILoc Circulation Charges"]})
            opportunitiesDetails["ILoc Circulation Charges".upper(
            )] = fieldsData["ILoc Circulation Charges"]

        if "ILoc Other Charges" in fieldsData:
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILocOtherCharges__c': fieldsData["ILOC Production Charges"]})
            opportunitiesDetails["ILoc Other Charges".upper(
            )] = fieldsData["ILOC Production Charges"]
            sleep(1)
        if "ILOC Production Charges" in fieldsData:
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILOCProductionCharges__c': fieldsData["ILOC Production Charges"]})
            opportunitiesDetails["ILOC Production Charges".upper(
            )] = fieldsData["ILOC Production Charges"]

        if "ILOC Total Program Fee" in fieldsData:
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILOCTotalProgramFee__c': fieldsData["ILOC Total Program Fee"]})
            opportunitiesDetails["ILOC Total Program Fee".upper(
            )] = fieldsData["ILOC Total Program Fee"]

        if "ILoc Artwork Due Date" in fieldsData:
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILocArtworkDueDate__c': str(insDate)})
            opportunitiesDetails["ILoc Artwork Due Date".upper()] = currentDate

        if "ILoc Market List Due Date" in fieldsData:
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILocMarketListDueDate__c': str(insDate)})
            opportunitiesDetails["ILoc Market List Due Date".upper()
                                 ] = currentDate

        if "ILoc Material Due Date" in fieldsData:
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILoc_Material_Due_Date__c': str(insDate)})
            opportunitiesDetails["ILoc Material Due Date".upper()
                                 ] = currentDate

        if "ILOC Remnant Estimated Cost" in fieldsData:
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILOCRemnantEstimatedCost__c': fieldsData["ILOC Remnant Estimated Cost"]})
            opportunitiesDetails["ILOC Remnant Estimated Cost".upper(
            )] = fieldsData["ILOC Remnant Estimated Cost"]

        if "ILOC Remnant Insert Date" in fieldsData:
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILOCRemnantInsertDate__c': str(insDate)})
            opportunitiesDetails["ILOC Remnant Insert Date".upper()
                                 ] = currentDate

        if "ILOC Total Remnant Charge Detail" in fieldsData:
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILOCTotalRemnantChargeDetail__c': fieldsData["ILOC Total Remnant Charge Detail"]})
            opportunitiesDetails["ILOC Total Remnant Charge Detail".upper(
            )] = fieldsData["ILOC Total Remnant Charge Detail"]

        isAddRemnantGuidelines = True
        if "Add Remnant Guidelines" in fieldsData:
            if fieldsData["Add Remnant Guidelines"].lower() == "yes":
                isAddRemnantGuidelines = True
            else:
                isAddRemnantGuidelines = False
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'AddRemnantGuidelines__c': isAddRemnantGuidelines})
            opportunitiesDetails["Add Remnant Guidelines".upper(
            )] = fieldsData["Add Remnant Guidelines"]

        isBilledBasesOnActualExecution = True
        if "Billed based on Actual Execution" in fieldsData:
            if fieldsData["Billed based on Actual Execution"].lower() == "checked":
                isBilledBasesOnActualExecution = True
            else:
                isBilledBasesOnActualExecution = False
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'Billed_based_on_Actual_Execution__c': isBilledBasesOnActualExecution})
            opportunitiesDetails["Billed based on Actual Execution".upper(
            )] = fieldsData["Billed based on Actual Execution"]

        # if "ILoc Category" in fieldsData:
        #     Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {'ILocCategory__c': fieldsData["ILoc Category"]})
        #     opportunitiesDetails["ILoc Category".upper()] = fieldsData["ILoc Category"]

        if "ILoc Category" in fieldsData:
            randomString = str(uuid.uuid4()).upper().replace("-", ", ")
            if fieldsData["ILoc Category"] == "RANDOM":
                randomString = randomString[0:200]
            else:
                randomString = fieldsData["ILoc Category"]
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILocCategory__c': randomString})
            opportunitiesDetails["ILoc Category".upper()] = randomString

        if "Status" in fieldsData:
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'Status__c': fieldsData["Status"]})
            opportunitiesDetails["Status".upper()] = fieldsData["Status"]

        # IS FIELDS
        randomString = None
        if "ILoc Store Count" in fieldsData:
            if fieldsData["ILoc Store Count"] == "RANDOM":
                value = str(randint(1, 100))
            else:
                value = fieldsData["ILoc Store Count"]
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILocStoreCount__c': value})
            opportunitiesDetails["ILoc Store Count".upper()] = value

        if "ILoc Trade Class" in fieldsData:
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILocTradeClass__c': fieldsData["ILoc Trade Class"]})
            opportunitiesDetails["ILoc Trade Class".upper(
            )] = fieldsData["ILoc Trade Class"]

        if "ILoc Brand" in fieldsData:
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILocBrand__c': fieldsData["ILoc Brand"]})
            opportunitiesDetails["Status".upper()] = fieldsData["ILoc Brand"]

        if "ILoc Loc Type" in fieldsData:
            randomString = str(uuid.uuid4()).upper().replace("-", " ")
            if fieldsData["ILoc Loc Type"] == "RANDOM":
                randomString = randomString[0:15]
            else:
                randomString = fieldsData["ILoc Loc Type"]
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILocLocType__c': randomString})
            opportunitiesDetails["ILoc Brand".upper()] = randomString

        if "ILoc Program" in fieldsData:
            randomString = str(uuid.uuid4()).upper().replace("-", " ")
            if fieldsData["ILoc Program"] == "RANDOM":
                randomString = randomString[0:15]
            else:
                randomString = fieldsData["ILoc Program"]
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILocProgram__c': randomString})
            opportunitiesDetails["ILoc Program".upper()] = randomString

        if "ILoc Geography" in fieldsData:
            randomString = str(uuid.uuid4()).upper().replace("-", " ")
            if fieldsData["ILoc Geography"] == "RANDOM":
                randomString = randomString[0:15]
            else:
                randomString = fieldsData["ILoc Geography"]
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILocGeography__c': randomString})
            opportunitiesDetails["ILoc Geography".upper()] = randomString

        # CO51 Fields
        if "ILoc Asset Due Date" in fieldsData:
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILoc_Assest_Due_Date__c': str(insDate)})
            opportunitiesDetails["ILoc Asset Due Date".upper()] = currentDate

        # DG Fields
        if "Client List Due Date" in fieldsData:
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ClientListDueDate__c': str(insDate)})
            opportunitiesDetails["Client List Due Date".upper()] = currentDate

        if "Artwork Due Date" in fieldsData:
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'Artwork_Due_Date__c': str(insDate)})
            opportunitiesDetails["Artwork Due Date".upper()] = currentDate

#         pdb.set_trace()
        if "ILOC Client processing fee" in fieldsData:
            if fieldsData["ILOC Client processing fee"] == "RANDOM":
                #                 value = str(round(random.uniform(10.1, 10000.9),2))
                value = "1250.85"
            else:
                value = fieldsData["ILOC Client processing fee"]
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILOC_Client_s_processing_fee__c': "1250.85"})
            opportunitiesDetails["ILOC Client processing fee".upper()
                                 ] = "1250.85"

        if "ILOC Client clip rate" in fieldsData:
            if fieldsData["ILOC Client clip rate"] == "RANDOM":
                value = "50.55"
            else:
                value = fieldsData["ILOC Client clip rate"]
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILOC_Client_s_clip_rate__c': "50.55"})
            opportunitiesDetails["ILOC Client clip rate".upper()] = "50.55"

        # SSD FIELDS
        if "ILoc Format" in fieldsData:
            randomString = str(uuid.uuid4()).upper().replace("-", " ")
            randomString = randomString[0:20]

            if fieldsData["ILoc Format"] == "RANDOM":
                value = randomString
            else:
                value = fieldsData["ILoc Format"]
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILoc_Format__c': value})
            opportunitiesDetails["ILoc Format".upper()] = value

        if "ILoc Targeting" in fieldsData:
            randomString = str(uuid.uuid4()).upper().replace("-", " ")
            randomString = randomString[0:15]

            if fieldsData["ILoc Targeting"] == "RANDOM":
                value = randomString
            else:
                value = fieldsData["ILoc Targeting"]
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILoc_Targeting__c': value})
            opportunitiesDetails["ILoc Targeting".upper()] = value

        if "ILoc Sample Due Date" in fieldsData:
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILoc_Sample_Due_Date__c': str(insDate)})
            opportunitiesDetails["ILoc Sample Due Date".upper()] = currentDate

        if "ILoc Creative Due Date" in fieldsData:
            Drivers.sf.Opportunity.update(str(recDetails[0]["Id"]), {
                                          'ILoc_Creative_Due_Date__c': str(insDate)})
            opportunitiesDetails["ILoc Creative Due Date".upper()
                                 ] = currentDate

        Messages.write_message("Opportunity Updated Successfully: " + oppName)
        oppDetails.update(opportunitiesDetails)
        data_store.spec[productName] = oppDetails
        print("\n", "Updated Detail: ")
        Messages.write_message(oppDetails)
        Utils.setDataInXlsx(productName, oppDetails)

#         opportunitiesDetails["NAME"] = oppData["NAME"]
#         opportunityDictionary[productName + " ILOC FIELDS"] = opportunitiesDetails
#         Utils.set_Opportunities_Details(
#             'OpportunityData.yaml', opportunityDictionary)

    @step("Verify Saved <productName> Opportunity")
    def verify_saved_opportunity(self, productName):
        opportunitiesDetails = {}
        opportunitiesDetails = data_store.spec.get(productName)
        Messages.write_message(f"Opportunity Details: {opportunitiesDetails}")
        Drivers.driverWait.until(
            EC.visibility_of_element_located((By.ID, "topButtonRow")))
        opportunityName = Drivers.driver.find_element_by_xpath(
            "//td[span[text()='Opportunity Name']]/following-sibling::*[position()=1][name()='td']/div").text
        assert opportunityName == opportunitiesDetails["NAME"]
        Messages.write_message(opportunityName + " : " +
                               opportunitiesDetails["NAME"])

    @step("Verify <productName> Opportunity Details")
    def verify_saved_opportunity_details(self, productName):

        opportunitiesDetails = {}
        opportunitiesDetails = data_store.spec.get(productName)
        Messages.write_message(f"Opportunity Details: {opportunitiesDetails}")
        Drivers.driverWait.until(
            EC.visibility_of_element_located((By.ID, "topButtonRow")))

        parentPath = Path(__file__).parents[1]
        opportunityObjectRepositoryFileName = str(
            parentPath) + "\\ObjectRepository\\" + os.getenv("OPPORTUNITY_OBJECT_REPOSITORY_FILE")
        opportunityObjectRepositorySheet = os.getenv(
            "OPPORTUNITY_DETAILS_OBJECT_REPOSITORY_SHEET")
        opportunityObjectRepositoryJsonFileName = str(
            parentPath) + "\\ObjectRepository\\" + f"{opportunityObjectRepositorySheet}.json"

        if os.path.exists(opportunityObjectRepositoryFileName):
            Utils.set_json_from_object_repository(
                opportunityObjectRepositorySheet, opportunityObjectRepositoryFileName)

        if os.path.exists(opportunityObjectRepositoryJsonFileName):
            f = open(opportunityObjectRepositoryJsonFileName)
            data_store.spec[opportunityObjectRepositorySheet] = json.load(f)
            opportunityObjectDetails = data_store.spec[opportunityObjectRepositorySheet]

        opportunitiesDetailsPath = Path(__file__).parents[1]
        opportunitiesDetailsFileName = str(
            opportunitiesDetailsPath) + "\\Data\\" + os.getenv("OPPORTUNITY_DETAILS_FILE")

        wb = None
        ws = None
        oppToSearch = opportunitiesDetails["NAME"]
        isOpportunityFound = False
        rowNum = 0
        if os.path.exists(opportunitiesDetailsFileName):
            wb = load_workbook(filename=opportunitiesDetailsFileName)
            ws = wb[productName]
            df = pd.read_excel(opportunitiesDetailsFileName,
                               sheet_name=productName)
            columnList = df.columns.tolist()

        from xlrd import open_workbook
        book = open_workbook(opportunitiesDetailsFileName)
        sh = book.sheet_by_name(productName)
        # for sheet in book.sheets():
        for rowidx in range(sh.nrows):
            row = sh.row(rowidx)
            for colidx, cell in enumerate(row):
                if cell.value == oppToSearch:
                    print(
                        ws[f"{xlsxwriter.utility.xl_col_to_name(colidx)}{rowidx+1}"].value)
                    Messages.write_message(
                        f"Found the opportunity name: {oppToSearch}")
                    rowNum = rowidx + 1
                    isOpportunityFound = True
                    break
#         pdb.set_trace()
        for columnName in columnList:
            #             pdb.set_trace()
            if columnName in opportunityObjectDetails:
                oppFieldValue = ws[f"{xlsxwriter.utility.xl_col_to_name(columnList.index(columnName))}{rowNum}"].value
                print("Getting Element: ", columnName)
                oppField = Utils.get_element(
                     columnName, sheet_name=opportunityObjectRepositorySheet, file_path=opportunityObjectRepositoryFileName)
#                 pdb.set_trace()
                if columnName in ('BILLED BASED ON ACTUAL EXECUTION', 'FREEDOM'):
                    attributeValue = oppField.get_attribute("title")

                    Messages.write_message(
                        f"Verifying value of {columnName} --> {oppFieldValue} == {attributeValue}")
                    assert attributeValue == oppFieldValue

                    print(f"{columnName} --> {oppFieldValue} == {attributeValue}")
                else:
                    Messages.write_message(
                        f"Verifying value of {columnName} --> {oppFieldValue} == {oppField.text}")
                    assert oppField.text == oppFieldValue
                    print(f"{columnName} --> {oppFieldValue} == {oppField.text}")

        totalLineItems = ws[f"{xlsxwriter.utility.xl_col_to_name(columnList.index('TOTAL LINE ITEMS'))}{rowNum}"].value

        iframeElement = Drivers.driver.find_element_by_xpath(
            "//iframe[@title='inlineRevenueCalculation']")
        Drivers.driver.switch_to.frame(iframeElement)
        revenueTable = Drivers.driver.find_element_by_xpath(
            "//table[contains(@class,'inlineRevenue')]/tbody")
        tableRows = revenueTable.find_elements_by_tag_name("tr")
#         pdb.set_trace()
        revenueInformatinTableXPATH = "//table[@class='list inlineRevenue dataTable no-footer']/thead/tr/th"
        columnsList = Utils.get_column_index(revenueInformatinTableXPATH)
        print(columnsList)

        if productName not in ("SMARTSOURCE_DIRECT_US") and str(productName.split("_")[1]) == "US":
            oppCurrencyLabel = "Usd"
            xlsxCurrencyLabel = "USD"
        elif productName not in ("SMARTSOURCE_DIRECT_CA") and str(productName.split("_")[1]) == "CA":
            oppCurrencyLabel = "Cad"
            xlsxCurrencyLabel = "CAD"
        elif productName in ("SMARTSOURCE_DIRECT_US") and str(productName.split("_")[2]) == "US":
            oppCurrencyLabel = "Usd"
            xlsxCurrencyLabel = "USD"
        elif productName in ("SMARTSOURCE_DIRECT_CA") and str(productName.split("_")[2]) == "CA":
            oppCurrencyLabel = "Cad"
            xlsxCurrencyLabel = "CAD"

        salesPriceLabel = f"Sales Price ({oppCurrencyLabel})"
        subTotalLabel = f"Sub Total ({oppCurrencyLabel})"
        revenueData = {}
        revenueData["Space"] = Decimal("00.00")
        revenueData["Production"] = Decimal("00.00")
        revenueData["Other Commissionable"] = Decimal("00.00")
        revenueData["Other-Commissionable"] = Decimal("00.00")
        revenueData["Other-Non-Commissionable"] = Decimal("00.00")
        revenueData["Other Non-Commissionable"] = Decimal("00.00")
        revenueData["Other-Non Commissionable"] = Decimal("00.00")
        revenueData["Space Revenue"] = Decimal("00.00")
        revenueData["Production Revenue"] = Decimal("00.00")
        revenueData["Other Commissionable Revenue"] = Decimal("00.00")
        revenueData["Other Non Commissionable Revenue"] = Decimal("00.00")
        revenueData["Commissionable Revenue"] = Decimal("00.00")
        revenueData["Non Commissionable Revenue"] = Decimal("00.00")
        revenueData["Total Revenue"] = Decimal("00.00")

        totalRevenue = 0
        for cnt in range(1, int(totalLineItems) + 1):
            cntr = 1
            colProductName = f"PRODUCT NAME"
            colChargeName = f"CHARGE NAME-{cnt}"
            colChargeDescription = f"CHARGE DESCRIPTION-{cnt}"
            colchargeTypeCategory = f"CHARGE TYPE CATEGORY-{cnt}"
            colCommissionable = f"COMMISSIONABLE?-{cnt}"
            colSalesPrice = f"SALES PRICE ({xlsxCurrencyLabel})-{cnt}"
            colQuantity = f"QUANTITY-{cnt}"
            colSubTotal = f"SUB TOTAL ({xlsxCurrencyLabel})-{cnt}"
            colQuanityBy1000 = f"Quantity/1000-{cnt}"
            colIsQuantityBy1000 = f"QUANTITY/1000?-{cnt}"
            for rowData in tableRows:
                print("\n", rowData.text)
                colProductNameValue = ws[
                    f"{xlsxwriter.utility.xl_col_to_name(columnList.index(colProductName))}{rowNum}"].value
                productXpath = f"//table[contains(@class,'inlineRevenue')]/tbody/tr[{cntr}]/td[{columnsList['Product Name']}]"
                productValue = Drivers.driver.find_element_by_xpath(
                    productXpath).text

                colChargeNameValue = ws[
                    f"{xlsxwriter.utility.xl_col_to_name(columnList.index(colChargeName))}{rowNum}"].value
                chargeNameXpath = f"//table[contains(@class,'inlineRevenue')]/tbody/tr[{cntr}]/td[{columnsList['Charge Name']}]"
                chargeNameValue = Drivers.driver.find_element_by_xpath(
                    chargeNameXpath).text

                colChargeDescriptionValue = ws[
                    f"{xlsxwriter.utility.xl_col_to_name(columnList.index(colChargeDescription))}{rowNum}"].value
                chargeDescriptionXpath = f"//table[contains(@class,'inlineRevenue')]/tbody/tr[{cntr}]/td[{columnsList['Charge Description']}]"
                chargeDescriptionValue = Drivers.driver.find_element_by_xpath(
                    chargeDescriptionXpath).text

                colChargeTypeCategoryValue = ws[
                    f"{xlsxwriter.utility.xl_col_to_name(columnList.index(colchargeTypeCategory))}{rowNum}"].value
                chargeTypeCategoryXpath = f"//table[contains(@class,'inlineRevenue')]/tbody/tr[{cntr}]/td[{columnsList['Charge Type Category']}]"
                chargeTypeCategoryValue = Drivers.driver.find_element_by_xpath(
                    chargeTypeCategoryXpath).text

                colCommissionableValue = ws[
                    f"{xlsxwriter.utility.xl_col_to_name(columnList.index(colCommissionable))}{rowNum}"].value
                commissionableXpath = f"//table[contains(@class,'inlineRevenue')]/tbody/tr[{cntr}]/td[{columnsList['Commissionable?']}]//img"
                commissionableElement = Drivers.driver.find_element_by_xpath(
                    commissionableXpath)
                attributeValue = commissionableElement.get_attribute("title")
                if attributeValue == "Checked":
                    commissionableValue = True
                else:
                    commissionableValue = False

                colSalesPriceValue = ws[
                    f"{xlsxwriter.utility.xl_col_to_name(columnList.index(colSalesPrice))}{rowNum}"].value
                salesPriceXpath = f"//table[contains(@class,'inlineRevenue')]/tbody/tr[{cntr}]/td[{columnsList[salesPriceLabel]}]"
                salesPriceValue = Drivers.driver.find_element_by_xpath(
                    salesPriceXpath).text

                colQuantityValue = ws[
                    f"{xlsxwriter.utility.xl_col_to_name(columnList.index(colQuantity))}{rowNum}"].value
                quantityXpath = f"//table[contains(@class,'inlineRevenue')]/tbody/tr[{cntr}]/td[{columnsList['Quantity']}]"
                quantityValue = Drivers.driver.find_element_by_xpath(
                    quantityXpath).text

                colSubTotalValue = ws[
                    f"{xlsxwriter.utility.xl_col_to_name(columnList.index(colSubTotal))}{rowNum}"].value
                if str(colSubTotalValue).find(".") > 0 and len(str(colSubTotalValue).split(".")[1]) == 1:
                    colSubTotalValue = str(colSubTotalValue).split(
                        ".")[0] + "." + str(colSubTotalValue).split(".")[1] + "0"
                elif str(colSubTotalValue).find(".") == 0:
                    colSubTotalValue = str(colSubTotalValue) + "." "00"
                subTotalXpath = f"//table[contains(@class,'inlineRevenue')]/tbody/tr[{cntr}]/td[{columnsList[subTotalLabel]}]"
                subTotalValue = Drivers.driver.find_element_by_xpath(
                    subTotalXpath).text
                subTotalValue = str(subTotalValue).replace(",", "")

#                 isQuantityBy1000 = ws[f"{xlsxwriter.utility.xl_col_to_name(columnList.index(colIsQuantityBy1000))}{rowNum}"].value

                print("\n")
                print(
                    f"Product Name --> {colProductNameValue} == {productValue}")
                print(
                    f"Charge Description --> {colChargeDescriptionValue} == {chargeDescriptionValue}")
                print(
                    f"Charge Type Category --> {colChargeTypeCategoryValue} == {chargeTypeCategoryValue}")
                print(
                    f"Commissionable? --> {colCommissionableValue} == {commissionableValue}")
                print(
                    f"Sales Price --> {colSalesPriceValue} == {salesPriceValue}")
                print(f"Quantity --> {colQuantityValue} == {quantityValue}")
                print(f"Sub Total --> {colSubTotalValue} == {subTotalValue}")
                print("\n")

                if (colProductNameValue == productValue):
                    print(
                        f"Verified Product Name --> {colProductNameValue} == {productValue}")
                    if (colChargeDescriptionValue == chargeDescriptionValue):
                        print(
                            f"Verified Charge Description --> {colChargeDescriptionValue} == {chargeDescriptionValue}")
                        if (colChargeTypeCategoryValue == chargeTypeCategoryValue):
                            print(
                                f"Verified Charge Type Category --> {colChargeTypeCategoryValue} == {chargeTypeCategoryValue}")
                            if (colCommissionableValue == commissionableValue):
                                print(
                                    f"Verified Commissionable? --> {colCommissionableValue} == {commissionableValue}")
                                if (str(colSalesPriceValue) == str(salesPriceValue)):
                                    print(
                                        f"Verified Sales Price --> {colSalesPriceValue} == {salesPriceValue}")
                                    if (str(colQuantityValue) == str(quantityValue)):
                                        print(
                                            f"Verified Quantity --> {colQuantityValue} == {quantityValue}")
                                        if (str(colSubTotalValue) == str(subTotalValue)):
                                            print(
                                                f"Verified Sub Total --> {colSubTotalValue} == {subTotalValue}")
                                            print(
                                                f"Add revenue for {chargeTypeCategoryValue}")
                                            subTotal = 0
                                            subTotal = revenueData[chargeTypeCategoryValue] + round(
                                                Decimal(subTotalValue), 2)
                                            revenueData[chargeTypeCategoryValue] = subTotal
#                                             pdb.set_trace()
#                                             break
#                                             if isQuantityBy1000:
#                                                 print(f"isQuantityBy1000 --> {isQuantityBy1000}")
#                                                 colQuanityBy1000Value = ws[f"{xlsxwriter.utility.xl_col_to_name(columnList.index(colQuanityBy1000))}{rowNum}"].value
#                                                 if len(str(colQuanityBy1000Value).split(".")[1]) == 1:
#                                                     colQuanityBy1000Value = str(colQuanityBy1000Value).split(".")[0] + "." + str(colQuanityBy1000Value).split(".")[1] + "0"
#                                                 quantityBy1000Xpath = f"//table[contains(@class,'inlineRevenue')]/tbody/tr[{cntr}]/td[{columnsList['Quantity/1000']}]"
#                                                 quantityBy1000Value = Drivers.driver.find_element_by_xpath(quantityBy1000Xpath).text
#                                                 quantityBy1000Value = str(quantityBy1000Value).replace(",", "")
#
#                                                 if (str(colQuanityBy1000Value) == str(quantityBy1000Value)):
#                                                     print(f"Verified QuanityBy1000 --> {colQuanityBy1000Value} == {quantityBy1000Value}")
#                                                     break
#                                             else:
                cntr = cntr + 1
        revenueData["Space Revenue"] = revenueData["Space"]
        print("\n\n", "Space Revenue: ", revenueData["Space Revenue"])

        revenueData["Production Revenue"] = revenueData["Production"]
        print("\n\n", "Production Revenue: ",
              revenueData["Production Revenue"])

        revenueData["Other Commissionable Revenue"] = revenueData["Other Commissionable"] + \
            revenueData["Other-Commissionable"]
        print("\n\n", "Other Commissionable Revenue: ",
              revenueData["Other Commissionable Revenue"])

        revenueData["Other Non Commissionable Revenue"] = revenueData["Other-Non-Commissionable"] + \
            revenueData["Other Non-Commissionable"] + \
            revenueData["Other-Non Commissionable"]
        print("\n\n", "Other Non Commissionable Revenue: ",
              revenueData["Other Non Commissionable Revenue"])

        revenueData["Commissionable Revenue"] = revenueData["Space Revenue"] + \
            revenueData["Production Revenue"] + \
            revenueData["Other Commissionable Revenue"]
        print("\n\n", "Commissionable Revenue: ",
              revenueData["Commissionable Revenue"])

        revenueData["Non Commissionable Revenue"] = revenueData["Other Non Commissionable Revenue"]
        print("\n\n", "Non Commissionable Revenue: ",
              revenueData["Other Non Commissionable Revenue"])

        revenueData["Total Revenue"] = revenueData["Commissionable Revenue"] + \
            revenueData["Non Commissionable Revenue"]
        print("\n\n", "Total Revenue: ", revenueData["Total Revenue"])

        spaceRevenueField = Drivers.driver.find_element_by_xpath(
            "//th[descendant::label[text()='Space Revenue']]/following-sibling::td").text.replace(",", "")
        productionRevenueField = Drivers.driver.find_element_by_xpath(
            "//th[descendant::label[text()='Production Revenue']]/following-sibling::td").text.replace(",", "")
        otherCommissionableField = Drivers.driver.find_element_by_xpath(
            "//th[descendant::label[text()='Other-Commissionable']]/following-sibling::td").text.replace(",", "")
        otherNonCommissionableField = Drivers.driver.find_element_by_xpath(
            "//th[descendant::label[text()='Other-Non Commissionable']]/following-sibling::td").text.replace(",", "")
        commissionableRevenueField = Drivers.driver.find_element_by_xpath(
            "//th[descendant::label[text()='Commissionable Revenue']]/following-sibling::td/span").text.replace(",", "")
        nonCommissionableRevenueField = Drivers.driver.find_element_by_xpath(
            "//th[descendant::label[text()='Non Commissionable Revenue']]/following-sibling::td").text.replace(",", "")
        totalRevenueField = Drivers.driver.find_element_by_xpath(
            "//th[descendant::label[text()='Total Revenue']]/following-sibling::td").text.replace(",", "")

        print("\n", spaceRevenueField, "\t", productionRevenueField, "\t", otherCommissionableField, "\t",
              otherNonCommissionableField, "\t", commissionableRevenueField, "\t", nonCommissionableRevenueField, "\t", totalRevenueField)

        if str(revenueData["Space Revenue"]) in spaceRevenueField:
            Messages.write_message(
                f"Verified Space Revenue: {str(revenueData['Space Revenue'])}")
            print(
                f"Verified Space Revenue: {str(revenueData['Space Revenue'])}")

        if str(revenueData["Production Revenue"]) in productionRevenueField:
            Messages.write_message(
                f"Verified Production Revenue: {str(revenueData['Production Revenue'])}")
            print(
                f"Verified Production Revenue: {str(revenueData['Production Revenue'])}")

        if str(revenueData["Other Commissionable Revenue"]) in otherCommissionableField:
            Messages.write_message(
                f"Verified Other Commissionable Revenue: {str(revenueData['Other Commissionable Revenue'])}")
            print(
                f"Verified Other Commissionable Revenue: {str(revenueData['Other Commissionable Revenue'])}")

        if str(revenueData["Other Non Commissionable Revenue"]) in otherNonCommissionableField:
            Messages.write_message(
                f"Verified Other Non Commissionable Revenue: {str(revenueData['Other Non Commissionable Revenue'])}")
            print(
                f"Verified Other Non Commissionable Revenue: {str(revenueData['Other Non Commissionable Revenue'])}")

        if str(revenueData["Commissionable Revenue"]) in commissionableRevenueField:
            Messages.write_message(
                f"Verified Commissionable Revenue: {str(revenueData['Commissionable Revenue'])}")
            print(
                f"Verified Commissionable Revenue: {str(revenueData['Commissionable Revenue'])}")

        if str(revenueData["Non Commissionable Revenue"]) in nonCommissionableRevenueField:
            Messages.write_message(
                f"Verified Non Commissionable Revenue: {str(revenueData['Non Commissionable Revenue'])}")
            print(
                f"Verified Non Commissionable Revenue: {str(revenueData['Non Commissionable Revenue'])}")

        if str(revenueData["Total Revenue"]) in totalRevenueField:
            Messages.write_message(
                f"Verified Total Revenue: {str(revenueData['Total Revenue'])}")
            print(
                f"Verified Total Revenue: {str(revenueData['Total Revenue'])}")

        Drivers.driver.switch_to.default_content()

    @after_step
    def after_step_hook(self, context):
        if context.step.is_failing == True:
            Messages.write_message(context.step.text)
            # Messages.write_message(context.step.message)
            Screenshots.capture_screenshot()
            
    @after_spec("<OpportunityCreation>")
    def after_spec_hook(self):
        Drivers.driver.quit()
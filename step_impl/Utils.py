from tabulate import tabulate
from xlrd import open_workbook
import pdb
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib, ssl
import json
import requests
import base64
import pandas as pd
import mimetypes
from step_impl import Drivers
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
from email import encoders
from getgauge.python import step, before_scenario, Messages, before_suite, before_spec, after_spec, after_step
from getgauge.python import after_suite
from getgauge.python import ExecutionContext, Screenshots, before_suite
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import xlsxwriter
import csv
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl import Workbook
from xlrd import open_workbook
import yaml
from time import sleep
from datetime import date
import sys
import xml.dom.minidom
import os
from pathlib import Path
import shutil
from openpyxl.styles.alignment import Alignment
from selenium.common.exceptions import NoSuchElementException

# def get_Configuration_Data( mapName):
#     cwd = os.path.dirname(os.path.realpath(__file__))
#     with open(cwd + "\\" + "OpportunityData.yml") as f:
#         dataMap = yaml.safe_load(f)
#         return dataMap[mapName]

def moveFile(sourceFile, destFile):
        shutil.copy(sourceFile, destFile)


def read_xml_copy_to_xlxs(xmlFileName, xlxsFileName, csvFileName):
    currentDate = date.today()
    currentDate = currentDate.strftime("%d%m%Y")
    # cwd = os.path.dirname(os.path.realpath(__file__))
    # parentDir = Path().resolve().parent
    cwd = os.path.dirname(os.getcwd())
    # xmlFileName = cwd + "\\NAM-PHASE-2\\reports\\xml-report\\" + xmlFileName
    # xlxsFileName = cwd + "\\NAM-PHASE-2\\reports\\" + xlxsFile
    # csvFileName = cwd + "\\NAM-PHASE-2\\reports\\" + csvFile
    doc = xml.dom.minidom.parse(xmlFileName)
    ws1 = None
    
    if os.path.exists(xlxsFileName):                    
        wb = load_workbook(filename=xlxsFileName)
    else:
        wb = Workbook()
        wb.save(xlxsFileName)                
    if currentDate in wb.sheetnames:
        std = wb.get_sheet_by_name(currentDate)
        wb.remove_sheet(std)
    ws1 = wb.create_sheet(currentDate, 1)
    
    testSuites = doc.getElementsByTagName("testsuite")
    testCasePassFail = "Pass"
    colIndex = 0
    patternFill = PatternFill(start_color='BDB76B', end_color='BDB76B', fill_type='solid')
    cnt = 2
    ws1.merge_cells('A1:E1')  
    
    cell = ws1.cell(row=1, column=5)  
    testSuitesName = ''
    for testSuite in testSuites:
        testsuiteSerialNo = testSuite.getAttribute("id")
        ws1[f"A{cnt}"] = "Test Suite Sr. No."
        ws1[f"A{cnt}"].fill = patternFill
        ws1[f"B{cnt}"] = testsuiteSerialNo

        
        testSuiteName = testSuite.getAttribute("name")
        testSuitesName = testSuiteName + "\n" + testSuitesName
        cell.value = testSuitesName 
        cell.alignment = Alignment(horizontal='center', vertical='center')  

        ws1.cell(row = 1, column = 1).value = testSuitesName        
        ws1[f"A{cnt+1}"] = "TEST SUITE NAME"
        ws1[f"A{cnt+1}"].fill = patternFill
        ws1[f"B{cnt+1}"] = testSuiteName

        testSuiteExecutionTime = testSuite.getAttribute("time")
        ws1[f"A{cnt+2}"] = "TEST SUITE TOTAL EXECUTION TIME(sec)"
        ws1[f"A{cnt+2}"].fill = patternFill
        ws1[f"B{cnt+2}"] = testSuiteExecutionTime
        
        testSuiteTotalTests = testSuite.getAttribute("tests")
        ws1[f"A{cnt+3}"] = "TEST SUITE TOTAL TESTS"
        ws1[f"A{cnt+3}"].fill = patternFill
        ws1[f"B{cnt+3}"] = testSuiteTotalTests
        
        testSuiteTestsFailures = testSuite.getAttribute("failures")
        ws1[f"A{cnt+4}"] = "TEST SUITE FAILURE TESTS COUNT"
        ws1[f"A{cnt+4}"].fill = patternFill
        ws1[f"B{cnt+4}"] = testSuiteTestsFailures
        
        testSuiteTestsPass = int(testSuiteTotalTests) - int(testSuiteTestsFailures)
        ws1[f"A{cnt+5}"] = "TEST SUITE PASS TESTS COUNT"
        ws1[f"A{cnt+5}"].fill = patternFill
        ws1[f"B{cnt+5}"] = str(testSuiteTestsPass)
            
        testSuiteExecutionTime = testSuite.getAttribute("timestamp")
        ws1[f"A{cnt+6}"] = "TEST SUITE EXECUTION DATE_TIME"
        ws1[f"A{cnt+6}"].fill = patternFill
        ws1[f"B{cnt+6}"] = testSuiteExecutionTime

        print(testSuiteName)
        print(testSuiteExecutionTime)
        print(testSuiteTotalTests)
        print(testSuiteTestsFailures)
        print(testsuiteSerialNo)
        print(testSuiteExecutionTime)

        ws1[f"A{cnt+8}"] = "SCENARIO NAME"
        ws1[f"A{cnt+8}"].fill = patternFill
        ws1[f"B{cnt+8}"] = "SCENARIO STATUS"
        ws1[f"B{cnt+8}"].fill = patternFill
        ws1[f"C{cnt+8}"] = "SCENARIO EXECUTION TIME"
        ws1[f"C{cnt+8}"].fill = patternFill
        ws1[f"D{cnt+8}"] = "SCENARIO FAILED STEP"
        ws1[f"D{cnt+8}"].fill = patternFill
        ws1[f"E{cnt+8}"] = "ERROR MESSAGE"
        ws1[f"E{cnt+8}"].fill = patternFill
        
        cnt = cnt + 9
        
        testCases = testSuite.getElementsByTagName("testcase")
        for testCase in testCases:
            testCaseName = testCase.getAttribute("name")
            testCaseExecutionTime = testCase.getAttribute("time")
            testCasePassFail = "PASS"

            ws1[f"A{cnt}"] = testCaseName
            ws1[f"C{cnt}"] = testCaseExecutionTime
            
            print(testCaseExecutionTime)        
            
            if testCase.hasChildNodes():
                failureTestCases = testCase.getElementsByTagName("failure")
                for failureTestCase in failureTestCases:
                    failureStep = failureTestCase.getAttribute("type")
                    failureMessage = failureTestCase.getAttribute("message")
                    testCasePassFail = "FAIL"
                    ws1[f"B{cnt}"] = testCasePassFail
                    ws1[f"D{cnt}"] = str(failureStep).split('Step Execution Failure')[0]
                    ws1[f"E{cnt}"] = failureMessage
                    print(failureTestCase)
                    print(failureStep)
                    print(failureMessage)
            else:
                ws1[f"B{cnt}"] = testCasePassFail
                ws1[f"D{cnt}"] = "-"
                ws1[f"E{cnt}"] = "-"

            print(testCasePassFail)
            cnt = cnt + 1
        cnt = cnt + 1
    wb.save(filename=xlxsFileName)
    with open(csvFileName, 'w', newline="") as f:  # open('test.csv', 'w', newline="") for python 3
        c = csv.writer(f)
        for r in ws1.rows:
            c.writerow([cell.value for cell in r])


def sendExecutionStatusMail():
    try:
        cwd = os.path.dirname(os.getcwd())
        smtp_server = os.getenv("SMTP_SERVER")
        port = os.getenv("SMTP_PORT")
#             sender_email = "namautomationphase1@gmail.com"
#             password = "automation@nam"
        sender_email = os.getenv("SENDER_ID")
        password = os.getenv("SENDER_PASSWORD")
        toaddr = os.getenv("TO_RECIPIENT").split(",")
        context = ssl.create_default_context()
            
        text = """
        Hi All,
            
        Please find below the NAM-PHASE-2 REGRESSION TCs EXECUTION RESULT.
        Also Also please find the attached report "result.xlsx" having complete information about regression suite execution.
            
        {table}
            
        Regards,
            
        System Admin"""
            
        html = """
        <html>
        <head>
        <style> 
            table, th, td {{ border: 1px solid black; border-collapse: collapse; }}
            th, td {{ padding: 1px; }}
        </style>
        </head>
        <body><p>Hi All,</p>
        <p>Please find below the NAM-PHASE-2 REGRESSION TCs EXECUTION RESULT.</p>
        <p>Also Also please find the attached report "result.xlsx" having complete information about regression suite execution.</p>
        {table}
        <p>Regards,</p>
        <p>NAM Admin</p>
        </body></html>
        """
        csvFileName = os.getenv("CSV_RESULT_FILE_NAME")
        with open(csvFileName) as input_file:
            reader = csv.reader(input_file)
            data = list(reader)
            
        text = text.format(table=tabulate(data, tablefmt="grid", stralign="center"))
        html = html.format(table=tabulate(data, tablefmt="html", stralign="center"))
            
        message = MIMEMultipart(
            "alternative", None, [MIMEText(text), MIMEText(html, 'html')])
            
        message['Subject'] = "NAM-PHASE-2 REGRESSION TCs EXECUTION RESULT"
        message['From'] = sender_email
        message['To'] = ", ".join(toaddr)
        
        # filename = os.getenv("XML_REPORT_FILE_COPY")
        # fp = open(filename, 'rb')
        # xls = MIMEBase('application', 'vnd.ms-excel')
        # xls.set_payload(fp.read())
        # fp.close()
        # encoders.encode_base64(xls)
        # xls.add_header('Content-Disposition', 'attachment', filename=filename)
        # message.attach(xls)
            
        try:
            server = smtplib.SMTP(smtp_server, port)
            server.ehlo()  # Can be omitted
            server.starttls(context=context)  # Secure the connection
            server.ehlo()  # Can be omittedju8
            server.login(sender_email, password)
            server.sendmail(sender_email, toaddr, message.as_string())
                        # TODO: Send email here
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)            
        finally:
            server.quit()

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)

                
def get_Opportunities_Details(fileName, mapName):
    cwd = os.path.dirname(os.path.realpath(__file__))
    with open(cwd + "\\" + fileName) as f:
        dataMap = yaml.safe_load(f)
        return dataMap[mapName]

def get_Opportunities_Header_Details(fileName, mapName):
    with open(fileName) as f:
        dataMap = yaml.safe_load(f)
        return dataMap[mapName]

def get_yaml_file_Details(fileName):
    with open(fileName) as f:
        dataMap = yaml.safe_load(f)
        return dataMap

def set_Opportunities_Details(fileName, data):
    cwd = os.path.dirname(os.path.realpath(__file__))
    cur_yaml = None
    with open(cwd + "\\" + fileName, 'r') as yamlfile:
        cur_yaml = yaml.load(yamlfile)
        cur_yaml.update(data)

    with open(cwd + "\\" + fileName, 'w') as yamlfile:
        data = yaml.safe_dump(cur_yaml, yamlfile)
        Messages.write_message(data)

        
def set_Opportunities_Header_Details(fileName, data):
    with open(fileName, 'r') as yamlfile:
        cur_yaml = yaml.load(yamlfile)
        if cur_yaml != None:
            cur_yaml.update(data)            
        else:
            cur_yaml = data
        print(cur_yaml)

    with open(fileName, 'w') as yamlfile:
        data = yaml.dump(cur_yaml, yamlfile, default_flow_style=False)
        Messages.write_message(data)


def removeFiles(location):
    try:
        filelist = [ f for f in os.listdir(location) if f.endswith(".xml")]
        for f in filelist:
            os.remove(os.path.join(location, f))
    except Exception:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)


def getNAPIDate():
    try:
        soql = f"SELECT Name FROM NAPI_Insert_date__c WHERE NAPI_Date__c > Today ORDER BY NAPI_Date__c ASC NULLS LAST"
        queryResult = Drivers.sf.query_all(query=soql)
        recDetails = queryResult['records']
        Messages.write_message("NAPI Date:" + str(recDetails[0]["Name"]))
        print(recDetails[0]["Name"])
        return str(recDetails[0]["Name"])
    except Exception:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)

# def setExcelColumnHeaders( productType, colName, oppFileName, ws1, wb):
#     df = pd.read_excel(oppFileName, sheet_name=productType)
#     print(df.columns)
#     columnList = df.columns.tolist()
#     
#     if colName in columnList:
#         print(f"{colName} Index: ", xlsxwriter.utility.xl_col_to_name(columnList.index(colName)), columnList.index(colName))
#     else:
#         if len(columnList) > 0:
#             columnCount = int(ws1.max_column)
#         else:
#             columnCount = 0
#         ws1[f"{xlsxwriter.utility.xl_col_to_name(columnCount)}{1}"] = colName
#     wb.save(filename=oppFileName)
        
def setColumnDetails(productName, ws1, columnIndex, colHeaderMap, wb, colName):
    opportunitiesDetailsPath = Path(__file__).parents[1]
    opportunityHeaderFilePath = str(opportunitiesDetailsPath) + "\\Data\\" + os.getenv("OPPORTUNITY_COLUMN_HEADER_FILE")
    data = get_yaml_file_Details(opportunityHeaderFilePath)
    oppHeaderData = {}
    if data != None and productName in data:
        print(productName, " found in ", data)
        oppHeaderData = data[productName]
    else:            
        ws1[f"{xlsxwriter.utility.xl_col_to_name(columnIndex)}{1}"] = colName
        colVal = f"{xlsxwriter.utility.xl_col_to_name(columnIndex)}"
        colHeaderMap[colName] = colVal
        print(colName, "-> ", columnIndex, "-> ", xlsxwriter.utility.xl_col_to_name(columnIndex))
        columnIndex = columnIndex + 1
    colVal = ""

    if (oppHeaderData != None and len(oppHeaderData) != 0) and colName in oppHeaderData:
        colVal = oppHeaderData[colName]
        colHeaderMap[colName] = colVal
        ws1[f"{colVal}{1}"] = colName
    elif (oppHeaderData != None and len(oppHeaderData) != 0) and colName not in oppHeaderData:
        print(colName, " not found in YAML")
        colIndex = int(ws1.max_column)
        ws1[f"{xlsxwriter.utility.xl_col_to_name(colIndex)}{1}"] = colName
        print("Sheet data index for coulmn: ", colName, "--", xlsxwriter.utility.xl_col_to_name(colIndex))
        colVal = f"{xlsxwriter.utility.xl_col_to_name(colIndex)}"
        colHeaderMap[colName] = colVal
        
    opportunitiesDetailsPath = Path(__file__).parents[1]
    opportunitiesDetailsFileName = str(opportunitiesDetailsPath) + "\\Data\\" + os.getenv("OPPORTUNITY_DETAILS_FILE")    
    wb.save(filename=opportunitiesDetailsFileName)
    print("------>", colHeaderMap)
    return colHeaderMap, columnIndex

def setILOCcolumnDetails(productName, ws1, columnIndex, colHeaderMap, wb, colName):
    iLOCDetailsPath = Path(__file__).parents[1]
    iLOCHeaderFilePath = str(iLOCDetailsPath) + "\\Data\\" + os.getenv("ILOC_COLUMN_HEADER_YAML")
    data = get_yaml_file_Details(iLOCHeaderFilePath)
    iLOCHeaderData = {}
    
    if (data != None) and (productName in data):
        print(productName, " found in ", data)
        iLOCHeaderData = data[productName]
    else:            
        ws1[f"{xlsxwriter.utility.xl_col_to_name(columnIndex)}{1}"] = colName
        colVal = f"{xlsxwriter.utility.xl_col_to_name(columnIndex)}"
        colHeaderMap[colName] = colVal
        print(colName, "-> ", columnIndex, "-> ", xlsxwriter.utility.xl_col_to_name(columnIndex))
        columnIndex = columnIndex + 1
    colVal = ""

    if (iLOCHeaderData != None and len(iLOCHeaderData) != 0) and colName in iLOCHeaderData:
        colVal = iLOCHeaderData[colName]
        colHeaderMap[colName] = colVal
        ws1[f"{colVal}{1}"] = colName
    elif (iLOCHeaderData != None and len(iLOCHeaderData) != 0) and colName not in iLOCHeaderData:
        print(colName, " not found in YAML")
        colIndex = int(ws1.max_column)
        ws1[f"{xlsxwriter.utility.xl_col_to_name(colIndex)}{1}"] = colName
        print("Sheet data index for coulmn: ", colName, "--", xlsxwriter.utility.xl_col_to_name(colIndex))
        colVal = f"{xlsxwriter.utility.xl_col_to_name(colIndex)}"
        colHeaderMap[colName] = colVal
        
    iLOCDetailsPath = Path(__file__).parents[1]
    iLOCDetailsFileName = str(iLOCDetailsPath) + "\\Data\\" + os.getenv("ILOC_DETAILS_XLSX")    
    wb.save(filename=iLOCDetailsFileName)
    print("------>", colHeaderMap)
    return colHeaderMap, columnIndex
    
def setDataInXlsx( productType, oppData):
    columnCount = 0
    rowCount = 0
    
    opportunitiesDetailsPath = Path(__file__).parents[1]
    opportunitiesDetailsFileName = str(opportunitiesDetailsPath) + "\\Data\\" + os.getenv("OPPORTUNITY_DETAILS_FILE")
    wb = None
    ws1 = None
    
    if os.path.exists(opportunitiesDetailsFileName):                    
        wb = load_workbook(filename=opportunitiesDetailsFileName)
    else:
        wb = Workbook()
        wb.save(opportunitiesDetailsFileName)                
    if productType in wb.sheetnames:
        ws1 = wb[productType]
    else:
        ws1 = wb.create_sheet(productType, 1)
    
    wb.save(filename=opportunitiesDetailsFileName)   
         
    df = pd.read_excel(opportunitiesDetailsFileName, sheet_name=productType)
    Messages.write_message(df.columns)
    columnList = df.columns.tolist()    
    
    rowCount = int(ws1.max_row) + 1
    for key, value in oppData.items():
#         pdb.set_trace()
        if key in columnList:
            ws1[f"{xlsxwriter.utility.xl_col_to_name(columnList.index(key))}{rowCount}"] = value
        else:
            if len(columnList) == 0 and ws1["A1"].value == None:
                ws1[f"{xlsxwriter.utility.xl_col_to_name(columnCount)}{1}"] = key
                ws1[f"{xlsxwriter.utility.xl_col_to_name(columnCount)}{2}"] = value
            else:
                columnCount = int(ws1.max_column)
                ws1[f"{xlsxwriter.utility.xl_col_to_name(columnCount)}{1}"] = key
                ws1[f"{xlsxwriter.utility.xl_col_to_name(columnCount)}{rowCount}"] = value
        wb.save(filename=opportunitiesDetailsFileName)        
    wb.save(filename=opportunitiesDetailsFileName)

def setILOCDataInXlsx( ilocRegion, oppData):
    columnCount = 0
    rowCount = 0
    
    parentPath = Path(__file__).parents[1]
    iLOCDetailsFileName = str(parentPath) + "\\Data\\" + os.getenv("ILOC_DETAILS_XLSX")
    wb = None
    ws1 = None
    
    if os.path.exists(iLOCDetailsFileName):                    
        wb = load_workbook(filename=iLOCDetailsFileName)
    else:
        wb = Workbook()
        wb.save(iLOCDetailsFileName)                
    if ilocRegion in wb.sheetnames:
        ws1 = wb[ilocRegion]
    else:
        ws1 = wb.create_sheet(ilocRegion, 1)
    
    wb.save(filename=iLOCDetailsFileName)   
         
    df = pd.read_excel(iLOCDetailsFileName, sheet_name=ilocRegion)
    Messages.write_message(df.columns)
    columnList = df.columns.tolist()
    contractNumber = oppData['CONTRACT_NUMBER']
    rowCount = 0
    isOpportunityFound = False
    if 'CONTRACT_NUMBER' in columnList:
        from xlrd import open_workbook
        book = open_workbook(iLOCDetailsFileName)
        sh = book.sheet_by_name(ilocRegion)
        # for sheet in book.sheets():
        while isOpportunityFound == False:
            for rowidx in range(sh.nrows):
                row = sh.row(rowidx)
                for colidx, cell in enumerate(row):
                    print(cell.value)
                    if cell.value == contractNumber:
                        print(
                            ws1[f"{xlsxwriter.utility.xl_col_to_name(colidx)}{rowidx}"].value)
                        Messages.write_message(
                            f"Found the opportunity name: {contractNumber}")
                        isOpportunityFound = True
                        rowCount = rowidx + 1
                        
                        break
                if isOpportunityFound == True:
                    break
            if isOpportunityFound == False:    
                rowCount = int(ws1.max_row) + 1
                break
    else:
        rowCount = int(ws1.max_row) + 1
    for key, value in oppData.items():
        if key in columnList:
            ws1[f"{xlsxwriter.utility.xl_col_to_name(columnList.index(key))}{rowCount}"] = value
        else:
            if len(columnList) == 0 and ws1["A1"].value == None:
                ws1[f"{xlsxwriter.utility.xl_col_to_name(columnCount)}{1}"] = key
                ws1[f"{xlsxwriter.utility.xl_col_to_name(columnCount)}{2}"] = value
            else:
                columnCount = int(ws1.max_column)
                ws1[f"{xlsxwriter.utility.xl_col_to_name(columnCount)}{1}"] = key
                ws1[f"{xlsxwriter.utility.xl_col_to_name(columnCount)}{rowCount}"] = value
        wb.save(filename=iLOCDetailsFileName)        
    wb.save(filename=iLOCDetailsFileName)

def get_element( fieldName, *args, **kwargs):
    #Optional Arguments
    sheetName = kwargs.get('sheet_name', None)
    filePath = kwargs.get('file_path', None)
    jsonData = kwargs.get('json_data', None)
       
    ilocElement = None
    byValue = None
    rowid = 0 
    isFound = False   
    # Messages.write_message(jsonData)
    if jsonData == None:
        df = pd.read_excel(filePath, sheet_name=sheetName)
        # print(df.columns)
        columnList = df.columns.tolist()
        print(columnList)
        
        book = open_workbook(filePath)
        sheet = book.sheet_by_name(sheetName)
        
        for rowidx in range(sheet.nrows):
            row = sheet.row(rowidx)
            for colidx, cell in enumerate(row):
                if cell.value == fieldName:                   
                    rowid = rowidx
                    print("Found at Row: ",rowid)
                    isFound = True
                    break
            if isFound == True:
                break
        print(rowid)  
        wb = load_workbook(filename=filePath)
        ws1 = wb[sheetName]
        print("\n", f"{xlsxwriter.utility.xl_col_to_name(columnList.index('Field'))}{rowid}", "\n")
        print(ws1[f"{xlsxwriter.utility.xl_col_to_name(columnList.index('Field'))}{rowid+1}"].value)
        
        locatorType = ws1[f"{xlsxwriter.utility.xl_col_to_name(columnList.index('Locator'))}{rowid+1}"].value
        print(ws1[f"{xlsxwriter.utility.xl_col_to_name(columnList.index('Locator'))}{rowid+1}"].value)
        
        locatorValue = ws1[f"{xlsxwriter.utility.xl_col_to_name(columnList.index('Value'))}{rowid+1}"].value
        print(ws1[f"{xlsxwriter.utility.xl_col_to_name(columnList.index('Value'))}{rowid+1}"].value)
        
        locatorExpectedCondition = ws1[f"{xlsxwriter.utility.xl_col_to_name(columnList.index('Expected_Conditions'))}{rowid+1}"].value
        print(ws1[f"{xlsxwriter.utility.xl_col_to_name(columnList.index('Value'))}{rowid+1}"].value)
    else:
        locatorType = jsonData[fieldName]['Locator']
        locatorValue = jsonData[fieldName]['Value']
        locatorExpectedCondition = jsonData[fieldName]['Expected_Conditions']
        locatorLabel = jsonData[fieldName]['Label']
        
    if locatorType == "ID":
        byValue = By.ID
        expected_condition_for_waiting( locatorExpectedCondition, locator_value = locatorValue, by_value = byValue)
        ilocElement = Drivers.driver.find_element_by_id(f"{locatorValue}")        
    elif locatorType == "XPATH":
        byValue = By.XPATH
        expected_condition_for_waiting( locatorExpectedCondition, locator_value = locatorValue, by_value = byValue) 
        ilocElement = Drivers.driver.find_element_by_xpath(f"{locatorValue}")        
    elif locatorType == "LINK_TEXT":
        byValue = By.LINK_TEXT
        expected_condition_for_waiting( locatorExpectedCondition, locator_value = locatorValue, by_value = byValue)
        ilocElement = Drivers.driver.find_element_by_link_text(f"{locatorValue}")
    elif locatorType == "PARTIAL_LINK_TEXT":
        byValue = By.PARTIAL_LINK_TEXT
        expected_condition_for_waiting( locatorExpectedCondition, locator_value = locatorValue, by_value = byValue)
        ilocElement = Drivers.driver.find_element_by_partial_link_text(f"{locatorValue}")        
    elif locatorType == "CLASS_NAME":
        byValue = By.CLASS_NAME
        expected_condition_for_waiting( locatorExpectedCondition, locator_value = locatorValue, by_value = byValue)
        ilocElement = Drivers.driver.find_element_by_class_name(f"{locatorValue}")        
    elif locatorType == "CSS_SELECTOR":
        byValue = By.CSS_SELECTOR
        expected_condition_for_waiting( locatorExpectedCondition, locator_value = locatorValue, by_value = byValue)
        ilocElement = Drivers.driver.find_element_by_css_selector(f"{locatorValue}")        
    elif locatorType == "NAME":
        byValue = By.NAME
        expected_condition_for_waiting( locatorExpectedCondition, locator_value = locatorValue, by_value = byValue)
        ilocElement = Drivers.driver.find_element_by_name(f"{locatorValue}")        
    elif locatorType == "TAG_NAME":
        byValue = By.TAG_NAME
        expected_condition_for_waiting( locatorExpectedCondition, locator_value = locatorValue, by_value = byValue)
        ilocElement = Drivers.driver.find_element_by_tag_name(f"{locatorValue}")    
    return ilocElement

def expected_condition_for_waiting( expectedConditionValue, *args, **kwargs): 
    #Optional Arguments
    locatorValue = kwargs.get('locator_value', None)
    # Messages.write_message(locatorValue)
    byValue = kwargs.get('by_value', None)
    # Messages.write_message(byValue)
    numberOfWindows = kwargs.get('number_of_windows', None)
    fieldText = kwargs.get('field_text', None)
    titleText = kwargs.get('title_text', None)
    windowHandle = kwargs.get('window_handle', None)
    
    if expectedConditionValue == 'element_to_be_clickable':
        Drivers.driverWait.until(EC.element_to_be_clickable(
                (byValue, f"{locatorValue}")))
        print()
    elif expectedConditionValue == 'element_to_be_selected':
        Drivers.driverWait.until(EC.element_to_be_selected(
                (byValue, f"{locatorValue}")))
        print()
    elif expectedConditionValue == 'frame_to_be_available_and_switch_to_it':
        Drivers.driverWait.until(EC.frame_to_be_available_and_switch_to_it(
                (byValue, f"{locatorValue}")))
        print()
    elif expectedConditionValue == 'invisibility_of_element':
        Drivers.driverWait.until(EC.invisibility_of_element(
                (byValue, f"{locatorValue}")))
        print()
    elif expectedConditionValue == 'invisibility_of_element_located':
        Drivers.driverWait.until(EC.invisibility_of_element_located(
                (byValue, f"{locatorValue}")))
        print()
    elif expectedConditionValue == 'new_window_is_opened':
        Drivers.driverWait.until(EC.new_window_is_opened(windowHandle))
        print()
    elif expectedConditionValue == 'number_of_windows_to_be':
        Drivers.driverWait.until(EC.number_of_windows_to_be(numberOfWindows))
        Messages.write_message("Number of Windows Found: " + str(numberOfWindows))
    elif expectedConditionValue == 'presence_of_all_elements_located':
        Drivers.driverWait.until(EC.presence_of_all_elements_located(
                (byValue, f"{locatorValue}")))
        print()
    elif expectedConditionValue == 'presence_of_element_located':
        Drivers.driverWait.until(EC.presence_of_element_located(
                (byValue, f"{locatorValue}")))
        print()
    elif expectedConditionValue == 'text_to_be_present_in_element':
        Drivers.driverWait.until(EC.text_to_be_present_in_element(
                (byValue, f"{locatorValue}")))
        print()
    elif expectedConditionValue == 'text_to_be_present_in_element_value':
        Drivers.driverWait.until(EC.text_to_be_present_in_element_value(
                (byValue, f"{locatorValue}")))
        print()
    elif expectedConditionValue == 'title_contains':
        Drivers.driverWait.until(EC.title_contains(
                (byValue, f"{locatorValue}")))
        print()
    elif expectedConditionValue == 'title_is':
        Drivers.driverWait.until(EC.title_is(
                (byValue, f"{locatorValue}")))
        print()
    elif expectedConditionValue == 'url_contains':
        Drivers.driverWait.until(EC.url_contains(
                (byValue, f"{locatorValue}")))
        print()
    elif expectedConditionValue == 'url_matches':
        Drivers.driverWait.until(EC.url_matches(
                (byValue, f"{locatorValue}")))
        print()
    elif expectedConditionValue == 'url_to_be':
        Drivers.driverWait.until(EC.url_to_be(
                (byValue, f"{locatorValue}")))
        print()
    elif expectedConditionValue == 'visibility_of_all_elements_located':
        Drivers.driverWait.until(EC.visibility_of_all_elements_located(
                (byValue, f"{locatorValue}")))
        print()
    elif expectedConditionValue == 'visibility_of_element_located':
        Drivers.driverWait.until(EC.visibility_of_element_located(
                (byValue, f"{locatorValue}")))
        print()    
    elif expectedConditionValue == 'alert_is_present':
        alert = Drivers.driverWait.until(EC.alert_is_present())
        alert.accept()
        print()                                    
            
def set_json_from_object_repository( sheetName, filePath):
    parentPath = Path(__file__).parents[1]
    ilocObjectRepositoryJsonFileName = str(parentPath) + "\\ObjectRepository\\" + f"{sheetName}.json"
    book = open_workbook(filePath)
    sh = book.sheet_by_name(sheetName)
    data_list = {}
    dataField = {}
    # data_list[dataField] = {}
    dataValue = {}
    for rownum in range(1, sh.nrows):
        row_values = sh.row_values(rownum)
#         pdb.set_trace()
        data_list[row_values[0]] = {'Locator':row_values[1],'Value'  : row_values[2], 'Expected_Conditions' : row_values[3], 'Label' : row_values[4]}
    jsonData = json.dumps(data_list)
    with open(ilocObjectRepositoryJsonFileName, 'w') as f:
        f.write(jsonData)
    print("completed")

def get_column_index( tableHeaderPath):
        labelsMap = {}
        try:            
            tableHeader = Drivers.driver.find_elements_by_xpath(tableHeaderPath)
                
            columnIndex = 1
            for theader in tableHeader:
                labelXpath = "//table[@id='SelectedProductType']/thead/tr/th[" + \
                    str(columnIndex) + "]/div"
                # labelCount = "//table[@id='SelectedProductType']/thead/tr/th[count(*)=1][" + str(columnIndex) + "]"
                # labelName = Drivers.driver.find_element_by_xpath(labelXpath)
                # print("Label Count: ", labelCount)
                # columnName = theader.find_element_by_css_selector(".slds-truncate").get_attribute("title")
                columnName = theader.text.title()
                labelsMap[columnName] = columnIndex
                Messages.write_message(
                    columnName + " : " + str(columnIndex))
                columnIndex = columnIndex + 1
            return labelsMap
        except NoSuchElementException:
            print("element not found")
def get_row_col_num_from_xlsx( sheetName, fileName, valueToSearch, *args, **kwargs):
    oppCurrency = kwargs.get('opp_currency', None)
    
    from xlrd import open_workbook
    book = open_workbook(fileName)
    sh = book.sheet_by_name(sheetName)
    # for sheet in book.sheets():
    rowId = 0
    for rowidx in range(sh.nrows):
        row = sh.row(rowidx)
        for colidx, cell in enumerate(row):
            if cell.value == valueToSearch:
#                 if ws[f"{xlsxwriter.utility.xl_col_to_name(columnList.index('Currency'))}{rowidx+1}"].value == oppCurrency:
                print(rowidx , colidx)
                rowId = rowidx + 1
                break
    return rowId
#     print(ws[f"{xlsxwriter.utility.xl_col_to_name(columnList.index('Product'))}{rowid}"].value)            
#     print(ws[f"{xlsxwriter.utility.xl_col_to_name(columnList.index('Charge'))}{rowid}"].value)            
#     print(ws[f"{xlsxwriter.utility.xl_col_to_name(columnList.index('Effective_Date'))}{rowid}"].value)
@after_step
def after_step_hook(context):
    if context.step.is_failing == True:
        Messages.write_message(context.step.text)
        # Messages.write_message(context.step.message)
        Screenshots.capture_screenshot()        
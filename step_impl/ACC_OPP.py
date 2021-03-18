from getgauge.python import Messages, Screenshots, after_step, data_store, step
import re
from simple_salesforce import Salesforce, SalesforceLogin
from datetime import date
import uuid
import xlsxwriter
from openpyxl import Workbook, load_workbook
import csv
import os
import pytz
from dateutil.parser import parse
from dateutil import tz
from datetime import timezone, tzinfo
import datetime
from random import randint
import pdb
from time import sleep
# from jproperties import Properties, codecs
from pathlib import Path
import random
import json
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl import Workbook
from simple_salesforce.format import format_soql
from step_impl import Drivers
from step_impl import Common_Steps
from step_impl import Utils


class AccountOpportunityCreation:
    session_id, instance = SalesforceLogin(
        username=os.getenv("USER_ID"), password=os.getenv("USER_PASSWORD"), security_token=os.getenv("USER_SECURITY_TOKEN"), domain='test')
    print(session_id, "\n", instance)
    Messages.write_message(str(session_id) + " : " + str(instance))
    sf = Salesforce(instance=instance, session_id=session_id)

    def verifyATAB(self, createdByUserID, formattedISODate):
        batchATAB = ["AccountTerritoryAssociationBatch", "UserHierarchyQueueDataSetupBatch", "UserHierarchyQueueDataSetupBatchUUHList", "UserHierarchyQueueDataSetupBatchO2AList", "UserHierarchyQueueAccountBatch", "UserHierarchyQueueOpportunityBatch",
                     "UserHierarchyQueueGoalBatch", "SupportUserHierarchySharingBatch", "userSharingBatch", "UserSharingBatchForFormHeaders", "UserRevokeSharingDataSetUpBatch", "UserSharingRevokeBatch", "UserSharingRevokeForFormHeaders"]

        batchATABStatus = {}
        apexClassName = None
        apexClassStatus = None
        for batchValue in batchATAB:
            print("Searching Current Batch:", batchValue)
            isCompleted = False
            while isCompleted == False:
                print("in while block")
                batchATABSOQL = f"SELECT CreatedDate, ApexClass.Name, MethodName, Status, CompletedDate FROM AsyncApexJob where ApexClass.Name = '{batchValue}' and CreatedById = '{createdByUserID}' and CreatedDate > {formattedISODate}"
                apexJobsResult = AccountOpportunityCreation.sf.query_all(
                    query=batchATABSOQL)
                apexJobsData = apexJobsResult['records']
                print("\n", apexJobsData)
                if len(apexJobsData) > 0:
                    if apexJobsData[0]['ApexClass'] != None:
                        print(
                            f"Verifying {apexJobsData[0]['ApexClass']['Name']} -- {apexJobsData[0]['Status']}")
                        if apexJobsData[0]['Status'] == 'Completed':
                            print(
                                f"Verified {apexJobsData[0]['ApexClass']['Name']} -- {apexJobsData[0]['Status']}")
                            apexClassName = apexJobsData[0]['ApexClass']['Name']
                            apexClassStatus = apexJobsData[0]['Status']
                            batchATABStatus[apexClassName] = apexClassStatus
                            isCompleted = True
                            Messages.write_message(
                                f"Verified {apexJobsData[0]['ApexClass']['Name']} -- {apexJobsData[0]['Status']}")
                        elif apexJobsData[0]['Status'] == 'Failed':
                            print(
                                f"Verified {apexJobsData[0]['ApexClass']['Name']} -- {apexJobsData[0]['Status']}")
                            Messages.write_message(
                                f"Verified {apexJobsData[0]['ApexClass']['Name']} -- {apexJobsData[0]['Status']}")
                            apexClassName = apexJobsData[0]['ApexClass']['Name']
                            apexClassStatus = apexJobsData[0]['Status']
                            batchATABStatus[apexJobsData[0]
                                            ['ApexClass']['Name']] = apexClassStatus
                            isCompleted = True
                        else:
                            sleep(15)
                else:
                    sleep(15)
                    isCompleted = False
        return batchATABStatus

    def getDateTime(self):
        currentDateTime = datetime.datetime.now(
            pytz.timezone("America/Los_Angeles"))
        twoMinutes = datetime.timedelta(minutes=2)
        end = currentDateTime - twoMinutes

        current_month = end.strftime('%m')
        # current_month_text = end.strftime('%h')
        # current_month_text = end.strftime('%B')

        current_day = end.strftime('%d')
        # current_day_text = end.strftime('%a')
        # current_day_full_text = end.strftime('%A')

        # current_weekday_day_of_today = end.strftime('%w')

        current_year_full = end.strftime('%Y')
        # current_year_short = end.strftime('%y')

        current_second = end.strftime('%S')
        current_minute = end.strftime('%M')
        current_hour = end.strftime('%H')
        # current_hour_word = end.strftime('%I')

        # d = datetime.datetime.today().replace(microsecond=0)
        # print(d.isoformat())

        finalDateTime = datetime.datetime(int(current_year_full), int(current_month), int(current_day), int(
            current_hour), int(current_minute), int(current_second), tzinfo=tz.gettz("America/Los_Angeles")).isoformat()
        print(datetime.datetime(int(current_year_full), int(current_month), int(current_day), int(current_hour), int(
            current_minute), int(current_second), tzinfo=tz.gettz("America/Los_Angeles")).isoformat())

        Messages.write_message(f"ATAB RUN TIME: {finalDateTime}")
        return finalDateTime

    def getNAPIDate(self, sf, isPastOrFuture):
        if isPastOrFuture == "Future":
            soql = f"SELECT Is_Active__c,Id, NAPI_Date__c, Name FROM NAPI_Insert_date__c WHERE NAPI_Date__c > Today and Is_Active__c = true ORDER BY NAPI_Date__c ASC NULLS LAST"
        elif isPastOrFuture == "Past":
            soql = f"SELECT Id, NAPI_Date__c, Name FROM NAPI_Insert_date__c WHERE NAPI_Date__c < LAST_N_MONTHS:12 and Is_Active__c = true ORDER BY NAPI_Date__c DESC NULLS LAST"
        queryResult = sf.query_all(query=soql)
        recDetails = queryResult['records']
    #     Messages.write_message("NAPI Date:" + str(recDetails[0]["Name"]))
        print("NAPI Date: ", recDetails[0]["NAPI_Date__c"])

        Messages.write_message(f"NAPI DATE: {recDetails[0]['NAPI_Date__c']}")
        return str(recDetails[0]["Id"]), str(recDetails[0]["NAPI_Date__c"])

    def getInstoreCycle(self, sf, cycleDate, current_month, current_year_full):
        cycleSOQL = f"SELECT Id,Begin_Date__c,End_Date__c,Name,Week__c FROM Cycle__c where Name = '{cycleDate}' and Week__c = 1.0"

        cycleResult = sf.query_all(query=cycleSOQL)
        cycleData = cycleResult['records']

        for cycle in cycleData:
            if cycle['Name'] == f"{current_year_full}{current_month}" and cycle['Week__c'] == 1.0:
                print(cycle['Id'], "\t", cycle['Name'], "\t", cycle['Week__c'])
        Messages.write_message(f"INSTORE CYCLE: {cycle['Name']}")
        return cycle['Id'], cycle['Name']

    @step("Create <region> Account")
    def create_account_and_store_to(self, region):
        data_store.spec.clear()
        if os.getenv('OPPORUNITIES_TO_CREATE_FOR_ACCOUNT') == 'NEW':
            accountType = region
            data_store.spec["REGION"] = region

            # accountData = p.properties

            if accountType == "USD":
                currencyType = "USD"
                contactLastName = "US"
                assignedTerritories = os.getenv('US_TERRITORIES')
                # assignedTerritories = 'LA Sales 2-6,Ssd Sales 2-3,NTL - COLGATE,Merchandising Region,Digital Division,C51 West Sales 4 - US'
            elif accountType == "CAD":
                currencyType = "CAD"
                contactLastName = "CA"
                assignedTerritories = os.getenv('CA_TERRITORIES')
                # assignedTerritories = 'Montreal Sales 2,Merchandising Canada Team,SSD Canada Sales 1,Digital Canada Sales 5,C51 Division - US'

            pattern = ","
            assignedTerritoriesList = re.split(pattern, assignedTerritories)

            # assignedTerritoriesList = assignedTerritories.split(",")
            Messages.write_message(
                "ASSIGNED TERRITORIES: {assignedTerritoriesList}")

            assignedTerritoriesMap = {}
            for assignedTerritory in assignedTerritoriesList:
                territoryNameSOQL = f"SELECT Id,Name,ParentTerritory2Id FROM Territory2 where Name = '{assignedTerritory}' and isActive__c = true"
                territoryNameResult = AccountOpportunityCreation.sf.query_all(
                    query=territoryNameSOQL)
                territoryNameData = territoryNameResult['records']
                if len(territoryNameData) > 0:
                    territoryID = territoryNameData[0]['Id']
                    assignedTerritoriesMap[assignedTerritory] = territoryNameData[0]['Id']
            Messages.write_message(
                f"ASSIGNED TERRITORIES MAP: {assignedTerritoriesMap}")
            accountName = f"AT-{str(randint(1, 99999))}"
            isAccountCreated = AccountOpportunityCreation.sf.Account.create(
                {'RecordTypeId': '012f1000000n6QJAAY', 'Name': f'{accountName}', 'CurrencyIsoCode': f'{currencyType}'})

            if isAccountCreated["success"] == True:
                print(
                    f"Account {accountName} is created {isAccountCreated['id']}")
                Messages.write_message(
                    f"Account {accountName} is created {isAccountCreated['id']}")
                data_store.spec[f'{region}_ACCOUNT_NAME'] = accountName
                data_store.spec[f'{region}_ACCOUNT_ID'] = isAccountCreated['id']

                Common_Steps.CommonSteps.write_data_to_table_column(
                    self, "Account", "Id", "id", isAccountCreated['id'], isAccountCreated['id'])
                Common_Steps.CommonSteps.write_data_to_table_column(
                    self, "Account", "Name", "string", accountName, isAccountCreated['id'])
                Common_Steps.CommonSteps.write_data_to_table_column(
                    self, "Account", "CurrencyIsoCode", "reference", currencyType, isAccountCreated['id'])
                # p['REGION'] = region
                # p['ACCOUNT_NAME'] = accountName
                # p['ACCOUNT_ID'] = isAccountCreated['id']

                for assignedTerritory in assignedTerritoriesList:
                    isObjectTerritory2Associated = AccountOpportunityCreation.sf.ObjectTerritory2Association.create(
                        {'AssociationCause': 'Territory2Manual', 'ObjectId': isAccountCreated['id'], 'Territory2Id': f'{assignedTerritoriesMap[assignedTerritory]}'})
                    print(isObjectTerritory2Associated)

            # contactId = None
            # contactData = sf.quick_search('fagoon4u@gmail.com')
            # contactId = contactData['searchRecords'][0]['Id']
            # print("Found Contact: ",contactData['searchRecords'][0]['Id'])

            firstName = os.getenv('ILOC_CUTOMER_NAME')
            lastName = contactLastName
            isContaactCreated = AccountOpportunityCreation.sf.Contact.create(
                {'AccountId': isAccountCreated['id'], 'Email': 'nam.customer@outlook.com', 'IsPrimaryContact__c': True, 'FirstName': firstName, 'LastName': lastName})
            if isContaactCreated["success"] == True:
                # p['CONTACT_NAME'] = f"{firstName} {lastName}"
                # p['CONTACT_ID'] = isContaactCreated['id']
                print(f"Contact {isContaactCreated['id']} is created...")
                Messages.write_message(
                    f"Contact {firstName} {lastName} is created {isContaactCreated['id']}")
                data_store.spec[f'{region}_CONTACT_NAME'] = f"{firstName} {lastName}"
                data_store.spec[f'{region}_CONTACT_ID'] = isContaactCreated['id']
                Common_Steps.CommonSteps.write_data_to_table_column(
                    self, "Contact", "Id", "id", isContaactCreated['id'], isContaactCreated['id'])
                Common_Steps.CommonSteps.write_data_to_table_column(
                    self, "Contact", "AccountId", "reference", isAccountCreated['id'], isContaactCreated['id'])
                Common_Steps.CommonSteps.write_data_to_table_column(
                    self, "Contact", "IsPrimaryContact__c", "reference", "True", isContaactCreated['id'])
                Common_Steps.CommonSteps.write_data_to_table_column(
                    self, "Contact", "FirstName", "string", firstName, isContaactCreated['id'])
                Common_Steps.CommonSteps.write_data_to_table_column(
                    self, "Contact", "LastName", "string", lastName, isContaactCreated['id'])
                Common_Steps.CommonSteps.write_data_to_table_column(
                    self, "Contact", "Email", "string", 'nam.customer@outlook.com', isContaactCreated['id'])
        # else:
        #     if region == 'USD':
        #         data_store.spec[f'{region}_ACCOUNT_ID'] = os.getenv(
        #         'OPPORUNITIES_TO_CREATE_FOR_USD_ACCOUNT_ID')
        #     elif region == 'CAD':
        #         data_store.spec[f'{region}_ACCOUNT_ID'] = os.getenv(
        #         'OPPORUNITIES_TO_CREATE_FOR_CAD_ACCOUNT_ID')

    @step("Run ATAB Batch")
    def run_atab_batch(self):
        if os.getenv('OPPORUNITIES_TO_CREATE_FOR_ACCOUNT') == 'NEW':
            userData = AccountOpportunityCreation.sf.quick_search(
                f"{os.getenv('USER_ID')}")
            userId = userData['searchRecords'][0]['Id']
            print(userData['searchRecords'][0]['Id'])

            result = AccountOpportunityCreation.sf.restful('tooling/executeAnonymous',
                                                           {'anonymousBody': 'AccountTerritoryAssociationBatch a = new AccountTerritoryAssociationBatch(); Database.executeBatch(a, 25);'})
            print(result)
            Messages.write_message(f"{result}")
            sleep(10)

            formattedDateTime = self.getDateTime()
            data_store.spec["ATAB_RUNNING_TIME"] = formattedDateTime
            createdByUserID = userId
            batchATABStatus = self.verifyATAB(
                createdByUserID, formattedDateTime)
            print(batchATABStatus)
        else:
            Messages.write_message(f"Acccount already present")

    @step("Verify <region> Account Teams")
    def verify_account_teams(self, region):
        if os.getenv('OPPORUNITIES_TO_CREATE_FOR_ACCOUNT') == 'NEW':
            accountType = region
            # p = Properties()
            # parentPath = Path(__file__).parents[1]
            # accountPropertiesFileName = str(
            #     parentPath) + "\\Data\\" + "Accounts.properties"

            # with open(accountPropertiesFileName, "r+b") as f:
            #     p.load(f, encoding='latin-1')

            #     accountData = p.properties
            accountID = data_store.spec.get(f'{region}_ACCOUNT_ID')

            if accountType == "USD":
                assignedTerritories = os.getenv('US_TERRITORIES')
            elif accountType == "CAD":
                assignedTerritories = os.getenv('CA_TERRITORIES')

            pattern = ","
            assignedTerritoriesList = re.split(pattern, assignedTerritories)
            # assignedTerritoriesList = assignedTerritories.split(",")

            assignedTerritoriesMap = {}
            for assignedTerritory in assignedTerritoriesList:
                territoryNameSOQL = f"SELECT Id,Name,ParentTerritory2Id FROM Territory2 where Name = '{assignedTerritory}' and isActive__c = true"
                territoryNameResult = AccountOpportunityCreation.sf.query_all(
                    query=territoryNameSOQL)
                territoryNameData = territoryNameResult['records']
                if len(territoryNameData) > 0:
                    territoryID = territoryNameData[0]['Id']
                    # p[assignedTerritory] = territoryNameData[0]['Id']
                    assignedTerritoriesMap[assignedTerritory] = territoryNameData[0]['Id']

            usersVerified = []
            usersNotVerified = []

            for assignedTerritory in assignedTerritoriesList:
                verifyMember = False
                territoryMembersData = None
                territoryIdToVerify = assignedTerritoriesMap[assignedTerritory]
                print(
                    f"----------------------------------------------------------------")
                print(
                    f"Territory to Verify: {territoryIdToVerify} : {assignedTerritory}")
                print(
                    f"----------------------------------------------------------------")
                usersVerified = []
                usersNotVerified = []
                while verifyMember == False:
                    territoryMembersSOQL = f"SELECT RoleInTerritory2,Territory2Id,UserId,User.Name FROM UserTerritory2Association where Territory2Id='{territoryIdToVerify}'"
                    territoryMembersResult = AccountOpportunityCreation.sf.query_all(
                        query=territoryMembersSOQL)
                    territoryMembersData = territoryMembersResult['records']
                    print(
                        f"Total Members in Territory: {len(territoryMembersData)}")
                    if len(territoryMembersData) > 0:
                        for territoryMembers in territoryMembersData:
                            if territoryMembers['RoleInTerritory2'] in ('Primary', 'Primary Split', 'Later Primary'):
                                print(
                                    f"{territoryMembers['User']['Name']} == {territoryMembers['RoleInTerritory2']}")
                                Messages.write_message(
                                    f"{territoryMembers['User']['Name']} == {territoryMembers['RoleInTerritory2']}")
                                verifyMember = True

                    if len(territoryMembersData) == 0 or verifyMember == False:
                        territoryNameSOQL = f"SELECT Id,Name,ParentTerritory2Id,ParentTerritory2.Name FROM Territory2 where Id = '{territoryIdToVerify}'"
                        territoryNameResult = AccountOpportunityCreation.sf.query_all(
                            query=territoryNameSOQL)
                        territoryNameData = territoryNameResult['records']
                        print(
                            f"Record in Parent Territory: {len(territoryMembersData)}")
                        if len(territoryNameData) > 0:
                            territoryIdToVerify = territoryNameData[0]['ParentTerritory2Id']
                            print(
                                f"Parent Territory: {territoryNameData[0]['ParentTerritory2']['Name']}")
                            Messages.write_message(
                                f"Parent Territory: {territoryNameData[0]['ParentTerritory2']['Name']}")

                print(f"Verify Member: {verifyMember}")

                for territoryMembers in territoryMembersData:
                    # '{accID}'"
                    accountMembersSOQL = f"SELECT Role_In_Territory__c ,TeamMemberRole__c,TerritoryId__c,Territory_Category__c,User__c,User__r.Name FROM Account_Team__c where Account__c = '{accountID}'"
                    accountMembersResult = AccountOpportunityCreation.sf.query_all(
                        query=accountMembersSOQL)
                    accountMembersData = accountMembersResult['records']
            #         pdb.set_trace()
                    isMemberVerified = False
                    if len(accountMembersData) > 0:
                        for accountMembers in accountMembersData:
                            if not isMemberVerified:
                                print(
                                    "\nVerifying Users", territoryMembers['User']['Name'], "\t", accountMembers['User__r']['Name'])
                #                 if (accountMembers['Role_In_Territory__c'] == territoryMembers['RoleInTerritory2']) and (accountMembers['TerritoryId__c'] == territoryMembers['Territory2Id']) and (accountMembers['User__c'] == territoryMembers['UserId']):
                                if (accountMembers['Role_In_Territory__c'] in territoryMembers['RoleInTerritory2']) and (accountMembers['User__c'] == territoryMembers['UserId']):

                                    createTable = f"CREATE TABLE IF NOT EXISTS ACCOUNT_TEAM(ACCOUNT_ID STRING(20), TERRITORY_ID STRING(20), TERRITORY_CATEGORY STRING(20), ROLE_IN_TERRITORY STRING(50), USER_ID STRING(20), USER_NAME STRING(50))"
                                    tableData = Drivers.dbCursor.execute(
                                        createTable)
                                    print("Table Created", tableData)
                                    Messages.write_message(
                                        f"Table Created {tableData}")

                                    tableRecords = Drivers.dbCursor.execute(
                                        f"SELECT EXISTS(SELECT 1 FROM ACCOUNT_TEAM WHERE ACCOUNT_ID = '{accountID}' AND TERRITORY_ID = '{accountMembers['TerritoryId__c']}' AND USER_NAME = '{accountMembers['User__r']['Name']}')")
                                    recordCount = tableRecords.fetchone()[0]
                                    print("Record(s) Found: ", recordCount)
                                    if recordCount == 0:
                                        insertValue = Drivers.dbCursor.execute(
                                            f"INSERT INTO ACCOUNT_TEAM (ACCOUNT_ID,TERRITORY_ID,TERRITORY_CATEGORY,ROLE_IN_TERRITORY,USER_ID,USER_NAME) VALUES ('{accountID}', '{accountMembers['TerritoryId__c']}',  '{accountMembers['Territory_Category__c']}', '{accountMembers['Role_In_Territory__c']}', '{accountMembers['User__c']}','{accountMembers['User__r']['Name']}')")
                                        Drivers.dbConn.commit()

                                    print(
                                        "\nMember Verified: ", accountMembers['User__r']['Name'], "\t", accountMembers['Role_In_Territory__c'])
                                    Messages.write_message(
                                        f"Member Verified: {accountMembers['User__r']['Name']} \t {accountMembers['Role_In_Territory__c']}")
                                    if territoryMembers['User']['Name'] not in usersVerified:
                                        usersVerified.append(
                                            territoryMembers['User']['Name'])

                                    if territoryMembers['User']['Name'] in usersNotVerified:
                                        usersNotVerified.remove(
                                            territoryMembers['User']['Name'])
                                    isMemberVerified = True
                                else:
                                    #                 print("Member Not Verified: ", accountMembers['User__r']['Name'], "\t", accountMembers['Role_In_Territory__c'])
                                    if territoryMembers['User']['Name'] not in usersNotVerified:
                                        usersNotVerified.append(
                                            territoryMembers['User']['Name'])

                                    if territoryMembers['User']['Name'] in usersVerified:
                                        usersVerified.remove(
                                            territoryMembers['User']['Name'])
                                    isMemberVerified = False
                    else:
                        Messages.write_message("No Team Members Found...")
                        print("No Team Members Found...")
                    print("\nMembers Verified: ", usersVerified)
                    print("\nMembers not verified: ", usersNotVerified)
                # f.truncate(0)
                # p.store(f, encoding='latin-1')
        else:
            Messages.write_message(f"Acccount already present")

    @step("Create Opportunities for <region> Account")
    def create_opportunities_for_account(self, region):
        currencyCode = ''
        opportunitiesDetailsPath = Path(__file__).parents[1]
        opportunitiesDetailsFileName = str(
            opportunitiesDetailsPath) + "\\Data\\" + os.getenv("ILOC_OPPORTUNITY_DETAILS_FILE")
        wb = None
        ws = None

        if os.path.exists(opportunitiesDetailsFileName):
            wb = load_workbook(filename=opportunitiesDetailsFileName)
        else:
            wb = Workbook()
            wb.save(opportunitiesDetailsFileName)

        end = datetime.datetime.now()
        current_month = end.strftime('%m')
        current_day = end.strftime('%d')
        current_year_full = end.strftime('%Y')
        # current_second = end.strftime('%S')
        # current_minute = end.strftime('%M')
        # current_hour = end.strftime('%H')

        futureDate = datetime.datetime(int(current_year_full) + 1,
                                       int(current_month), int(current_day)).date()
        pastDate = datetime.datetime(int(current_year_full) - 1,
                                     int(current_month), int(current_day)).date()
        
        print("Future Date: ", futureDate, "Past Date: ", pastDate)
        userData = AccountOpportunityCreation.sf.quick_search(
            f"{os.getenv('USER_ID')}")
        userId = userData['searchRecords'][0]['Id']
        print(userData['searchRecords'][0]['Id'])
        data_store.spec["FUTURE_DATE"] = futureDate
        data_store.spec["PAST_DATE"] = pastDate
        data_store.spec["USER_ID"] = userId

        napiDateId, napiDate = AccountOpportunityCreation.getNAPIDate(
            self, AccountOpportunityCreation.sf, "Future")
        data_store.spec["FUTURE_NAPI_DATE"] = napiDate

        pastNapiDateId, pastNAPIDate = AccountOpportunityCreation.getNAPIDate(
            self, AccountOpportunityCreation.sf, "Past")
        data_store.spec["PAST_NAPI_DATE"] = pastNAPIDate

        instoreCycle, instoreCycleName = AccountOpportunityCreation.getInstoreCycle(
            self, AccountOpportunityCreation.sf, f"{str(int(current_year_full)+1)}{current_month}", current_month, current_year_full)
        pastInStoreCycle, pastInStoreCycleName = AccountOpportunityCreation.getInstoreCycle(
            self, AccountOpportunityCreation.sf, f"{str(int(current_year_full)-1)}{current_month}", current_month, current_year_full)
        data_store.spec["FUTURE_INSTORE_CYCLE"] = instoreCycle

        data_store.spec["PAST_INSTORE_CYCLE"] = pastInStoreCycle

        productLineName = os.getenv('OPPORTUNITY_FOR_PRODUCT_LINE')
        data_store.spec["PRODUCT_LINES"] = productLineName

        usProductLineMap = {}
        caProductLineMap = {}
        productLineList = None

        if os.getenv('OPPORUNITIES_TO_CREATE_FOR_ACCOUNT') == 'EXISTING':
            accountID = os.getenv(
                f'OPPORUNITIES_TO_CREATE_FOR_{region}_ACCOUNT_ID')
        elif os.getenv('OPPORUNITIES_TO_CREATE_FOR_ACCOUNT') == 'NEW':
            accountID = data_store.spec.get(f'{region}_ACCOUNT_ID')

        if region == 'USD':
            currencyCode = 'USD'
            # if productLineName == "ALL":
            productLine = os.getenv('US_PRODUCT_LINES')
            # else:
            #     productLine = productLineName

        elif region == 'CAD':
            currencyCode = 'CAD'            
            # if productLineName == "ALL":
            productLine = os.getenv('CA_PRODUCT_LINES')
            # else:
            #     productLine = productLineName

        # pattern = "^ ([A-Z0-9- a-z]+),"
        pattern = ","
        productLineList = re.split(pattern, productLine)
        print(*productLineList, sep="\n")
        # pdb.set_trace()
        # if "," in productLine:
        #     productLineList = productLine.split(",")
        # else:
        #     productLineList = [productLine]

        # Get all opportunity record types :
        opprtunityPLMap = {}
        opprtunityPLSOQL = f"select Id,Name from RecordType where sObjectType='Opportunity'"
        opprtunityPLResult = AccountOpportunityCreation.sf.query_all(
            query=opprtunityPLSOQL)
        opprtunityPLData = opprtunityPLResult['records']

        for opprtunityPL in opprtunityPLData:
            opprtunityPLMap[opprtunityPL['Name']] = opprtunityPL['Id']

        [print(key, value) for key, value in opprtunityPLMap.items()]

        for productLine in productLineList:
            childOpps = []
            oppDetails = {}
            oppDetails.clear()
            if f"{productLine}_{currencyCode}" in wb.sheetnames:
                ws = wb[f"{productLine}_{currencyCode}"]
            else:
                ws = wb.create_sheet(f"{productLine}_{currencyCode}", 1)

            wb.save(filename=opportunitiesDetailsFileName)

            df = pd.read_excel(opportunitiesDetailsFileName,
                               sheet_name=f"{productLine}_{currencyCode}")
            Messages.write_message(df.columns)
            columnList = df.columns.tolist()

            rowCount = int(df.shape[0])
            colCount = int(df.shape[1])
            # pdb.set_trace()
            print(
                f"\n{productLine} Started---------------------------------------------------------------------------\n")
            usProductLineMap[productLine] = {
                'PRODUCT_LINE_ID': opprtunityPLMap[productLine], 'PRODUCT_LINE_NAME': productLine, 'PARENT_OPP': ''}

            parentOpportunityName = f"{productLine}#P#{rowCount + 1}"
            if productLine in ('InStore', 'InStore-Canada'):
                parentOppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine], 'Opportunity_Category__c': 'Parent', 'Name': parentOpportunityName,
                                                                                     'StageName': 'Contract', 'InStore_Cycle__c': str(instoreCycle), 'CloseDate': str(futureDate), 'AccountId': accountID, 'ILocArtworkDueDate__c': str(futureDate)})
            elif productLine in ('FSI'):
                parentOppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine], 'Opportunity_Category__c': 'Parent', 'Name': parentOpportunityName,
                                                                                     'StageName': 'Contract', 'NAPI_Insert_date__c': str(napiDateId), 'CloseDate': str(napiDate), 'AccountId': accountID})
        #     elif productLine in ('Digital','Digital- Canada'):
        #         parentOppDetails = sf.Opportunity.create({'RecordTypeId' : opprtunityPLMap[productLine],'Opportunity_Category__c' : 'Parent', 'Name' : parentOpportunityName,'StageName' : 'Contract','Insert_Date__c' : str(futureDate),'CloseDate' : str(futureDate), 'End_Date__c' : str(futureDate), 'AccountId': accountID})
            else:
                # pdb.set_trace()
                parentOppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine], 'Opportunity_Category__c': 'Parent', 'Name': parentOpportunityName, 'StageName': 'Contract', 'Insert_Date__c': str(futureDate), 'CloseDate': str(futureDate), 'End_Date__c': str(futureDate), 'AccountId': accountID})

            print(parentOppDetails)
            if parentOppDetails["success"] == True:
                print(f"Parent Opp ID of {productLine}: ",
                      parentOppDetails['id'])
                Messages.write_message(
                    f"Parent Opp ID of {productLine}: {parentOppDetails['id']}")

                data_store.spec[f"1_PO_ID_{productLine}"] = parentOppDetails['id']
                data_store.spec[f"1_PO_NAME_{productLine}"] = parentOpportunityName
                oppDetails['P_ID'] = parentOppDetails['id']
                oppDetails['P_NAME'] = parentOpportunityName

                usProductLineMap[productLine]['PARENT_OPP_1'] = {
                    'ID': parentOppDetails['id'], 'CHILD1_OPP_ID': '', 'CHILD2_OPP_ID': ''}
                usProductLineMap[productLine]['PARENT_OPP_1']['CHILD1_OPP_ID'] = {
                    'ID': '', 'ORDER#': '', 'PARENT_ORDER#': ''}
                usProductLineMap[productLine]['PARENT_OPP_1']['CHILD2_OPP_ID'] = {
                    'ID': '', 'ORDER#': '', 'PARENT_ORDER#': ''}

            for i in range(1, 2):
                df = pd.read_excel(opportunitiesDetailsFileName,
                                   sheet_name=f"{productLine}_{currencyCode}")
                rowCount = int(df.shape[0])
                colCount = int(df.shape[1])
                today = date.today()
                oppDateToday = today.strftime("%Y%m%d")
                child1OpportunityName = f"{productLine}{oppDateToday}"
                p1OrderNumber = str(uuid.uuid4()).upper()[:10]

                if productLine in ('InStore', 'InStore-Canada'):
                    if productLine in ('InStore'):
                        projectCode = 300000
                    elif productLine in ('InStore-Canada'):
                        projectCode = 399999

                    child1OppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine],
                                                                                         'Opportunity_Category__c': 'Child',
                                                                                         'Name': child1OpportunityName,
                                                                                         'StageName': 'Contract',
                                                                                         'InStore_Cycle__c': str(instoreCycle),
                                                                                         'CloseDate': str(futureDate),
                                                                                         'End_Date__c': str(futureDate),
                                                                                         'Artwork_Due_Date__c': str(futureDate),
                                                                                         'AccountId': accountID,
                                                                                         'Parent_Opportunity__c': parentOppDetails['id'],
                                                                                         'Probability__c': '75',
                                                                                         'Estimated_Average_CPS__c': 1,
                                                                                         'Estimated_Store_Count__c': 1,
                                                                                         'Business_Type__c': 'New',
                                                                                         'Status__c': 'Reserved-RS1',
                                                                                         'Type': 'Tactical',
                                                                                         'Order__c': p1OrderNumber,
                                                                                         'Parent_Order__c': '',
                                                                                         'ILocArtworkDueDate__c': str(futureDate),
                                                                                         'ILocStoreCount__c': 10,
                                                                                         'ILocTradeClass__c': 'TRADE CLASS',
                                                                                         'ILocCategory__c': 'CATEGORY1, CATEGORY2, CATEGORY3',
                                                                                         'ILocBrand__c': 'BRAND',
                                                                                         'ILocLocType__c': 'TYPE',
                                                                                         'ILocProgram__c': 'PROGRAM',
                                                                                         'ILocGeography__c': '20',
                                                                                         'ILOCProjectCode__c': projectCode,
                                                                                         'Billed_based_on_Actual_Execution__c': True})
                    oppDetails['C_CYCLE'] = str(instoreCycleName)
                    oppDetails['C_STATUS'] = 'Reserved-RS1'
                    oppDetails['C_ILocStoreCount'] = 10
                    oppDetails['C_ILocTradeClass'] = 'TRADE CLASS'
                    oppDetails['C_ILocCategory'] = 'CATEGORY1, CATEGORY2, CATEGORY3'
                    oppDetails['C_ILocBrand'] = 'BRAND'
                    oppDetails['C_ILocLocType'] = 'TYPE'
                    oppDetails['C_ILocProgram'] = 'PROGRAM'
                    oppDetails['C_ILocGeography'] = '20'
                    oppDetails['C_ILOCProjectCode'] = projectCode
                    oppDetails['C_Billed_based_on_Actual_Execution'] = True
                    
                    if child1OppDetails["success"] == True:
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Id", 'id', child1OppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CurrencyIsoCode", 'text', currencyCode, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeId", 'id', str(opprtunityPLMap[productLine]), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeName", 'string', str(productLine), child1OppDetails['id'])                        
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Name", 'string', str(child1OpportunityName), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "PARENT_OPP_ID", 'id', parentOppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "StageName", 'text', 'Contract', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "InStore_Cycle__c", 'text', str(instoreCycle), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CloseDate", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "End_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Artwork_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "AccountId", 'reference', str(accountID), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Probability__c", 'text', '75', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Estimated_Average_CPS__c", 'currency', 1, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Estimated_Store_Count__c", 'currency', 1, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Business_Type__c", 'text', 'New', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Status__c", 'text', 'Reserved-RS1', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Type", 'text', 'Tactical', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Order__c", 'text', str(p1OrderNumber), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Parent_Order__c", 'text', '', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocArtworkDueDate__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocStoreCount__c", 'text', '1', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocTradeClass__c", 'text', 'TRADE CLASS', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocCategory__c", 'text', 'CATEGORY1, CATEGORY2, CATEGORY3', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocBrand__c", 'text', 'BRAND', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocLocType__c", 'text', 'TYPE', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocProgram__c", 'text', 'PROGRAM', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocGeography__c", 'text', '20', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOCProjectCode__c", 'text', str(projectCode), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])
                elif productLine in ('FSI'):
                    child1OppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine],
                                                                                         'Opportunity_Category__c': 'Child',
                                                                                         'Name': child1OpportunityName,
                                                                                         'StageName': 'Contract',
                                                                                         'NAPI_Insert_date__c': str(napiDateId),
                                                                                         'CloseDate': str(napiDate),
                                                                                         'End_Date__c': str(napiDate),
                                                                                         'AccountId': accountID,
                                                                                         'Parent_Opportunity__c': parentOppDetails['id'],
                                                                                         'Probability__c': '75',
                                                                                         'Estimated_Average_CPM__c': 1,
                                                                                         'Expected_Circulation__c': 1,
                                                                                         'Business_Type__c': 'New',
                                                                                         'Status__c': 'Reserved-RS1',
                                                                                         'Type': 'Standard',
                                                                                         'ILocCirculationCharges__c': '[{"ChargeType":"C","Charges":1000,"Description":"CIRCULATION CHARGE","Amount":1,"CirculationQty":1000}]',
                                                                                         'ILOCProductionCharges__c': '[{"ChargeType":"P","Charges":2000,"Description":"DISK HANDLING","Amount":1000,"CirculationQty":2}]',
                                                                                         'ILocOtherCharges__c': '[{"ChargeType":"o","Charges":5000,"Description":"OTHER CHARGE","Amount":50, "CirculationQty": 100}]',
                                                                                         'ILOCTotalProgramFee__c': 8000,
                                                                                         'Order__c': p1OrderNumber,
                                                                                         'Parent_Order__c': '',
                                                                                         'ILOCAdDescription__c': 'Description',
                                                                                         'ILocCategory__c': 'Category',
                                                                                         'ILocMarketListDueDate__c': str(napiDate),
                                                                                         'ILocArtworkDueDate__c': str(napiDate),
                                                                                         'Billed_based_on_Actual_Execution__c': True})
                    
                    if child1OppDetails["success"] == True: 
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Id", 'id', child1OppDetails['id'], child1OppDetails['id'])     
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CurrencyIsoCode", 'text', currencyCode, child1OppDetails['id'])    
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeId", 'id', str(opprtunityPLMap[productLine]), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeName", 'string', str(productLine), child1OppDetails['id'])                        
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Name", 'string', str(child1OpportunityName), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "PARENT_OPP_ID", 'id', parentOppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "StageName", 'text', 'Contract', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CloseDate", 'text', napiDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "End_Date__c", 'text', napiDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "NAPI_Insert_date__c", 'text', napiDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "AccountId", 'reference', str(accountID), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Probability__c", 'text', '75', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Estimated_Average_CPM__c", 'currency', 1, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Expected_Circulation__c", 'currency', 1, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Business_Type__c", 'text', 'New', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Status__c", 'text', 'Reserved-RS1', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Type", 'text', 'Standard', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocCirculationCharges__c", 'textarea', r'[{"ChargeType":"C","Charges":1000,"Description":"CIRCULATION CHARGE","Amount":1,"CirculationQty":1000}]', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOCProductionCharges__c", 'textarea', r'[{"ChargeType":"P","Charges":2000,"Description":"DISK HANDLING","Amount":1000,"CirculationQty":2}]', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocOtherCharges__c", 'textarea', r'[{"ChargeType":"o","Charges":5000,"Description":"OTHER CHARGE","Amount":50, "CirculationQty": 100}]', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOCTotalProgramFee__c", 'text', '8000', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Order__c", 'text', str(p1OrderNumber), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Parent_Order__c", 'text', '', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOCAdDescription__c", 'text', 'Description', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocCategory__c", 'text', 'Category', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocMarketListDueDate__c", 'text', napiDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocArtworkDueDate__c", 'text', napiDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])
                    oppDetails['C_NAPI_DATE'] = str(napiDateId)
                    oppDetails['C_STATUS'] = 'Reserved-RS1'
                    oppDetails['C_TOTAL_PROGRAM_FEE'] = 8000
                    oppDetails['C_ILOCAdDescription'] = 'Description'
                    oppDetails['C_ILocMarketListDueDate'] = str(
                        napiDateId)
                    oppDetails['C_ILocArtworkDueDate'] = str(napiDateId)
                    oppDetails['C_ILocCategory'] = 'CATEGORY1, CATEGORY2, CATEGORY3'
                    oppDetails['C_Billed_based_on_Actual_Execution'] = True
                elif productLine in ('SSMG'):
                    child1OppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine],
                                                                                         'Opportunity_Category__c': 'Child',
                                                                                         'Name': child1OpportunityName,
                                                                                         'StageName': 'Contract',
                                                                                         'Insert_Date__c': str(futureDate),
                                                                                         'CloseDate': str(futureDate),
                                                                                         'End_Date__c': str(futureDate),
                                                                                         'AccountId': accountID,
                                                                                         'Parent_Opportunity__c': parentOppDetails['id'],
                                                                                         'Probability__c': '75',
                                                                                         'Estimated_Average_CPM__c': 1,
                                                                                         'Expected_Circulation__c': 1,
                                                                                         'Business_Type__c': 'New',
                                                                                         'Type': 'Standard',
                                                                                         'Order__c': p1OrderNumber,
                                                                                         'Parent_Order__c': '',
                                                                                         'ILocMarketListDueDate__c': str(futureDate),
                                                                                         'ILoc_Material_Due_Date__c': str(futureDate),
                                                                                         'Billed_based_on_Actual_Execution__c': True})
                    if child1OppDetails["success"] == True:  
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Id", 'id', child1OppDetails['id'], child1OppDetails['id'])     
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CurrencyIsoCode", 'text', currencyCode, child1OppDetails['id'])    
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeId", 'id', str(opprtunityPLMap[productLine]), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeName", 'string', str(productLine), child1OppDetails['id'])                        
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Name", 'string', str(child1OpportunityName), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "PARENT_OPP_ID", 'id', parentOppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "StageName", 'text', 'Contract', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Insert_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CloseDate", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "End_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "AccountId", 'reference', str(accountID), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Probability__c", 'text', '75', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Estimated_Average_CPM__c", 'currency', 1, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Expected_Circulation__c", 'currency', 1, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Business_Type__c", 'text', 'New', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Status__c", 'text', 'Reserved-RS1', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Type", 'text', 'Standard', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Order__c", 'text', str(p1OrderNumber), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Parent_Order__c", 'text', '', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocMarketListDueDate__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Material_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])

                    oppDetails['C_ILocMarketListDueDate'] = str(
                        futureDate)
                    oppDetails['C_ILoc_Material_Due_Date'] = str(
                        futureDate)
                    oppDetails['C_Billed_based_on_Actual_Execution'] = True
                elif productLine in ('Checkout 51'):
                    child1OppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine],
                                                                                         'Opportunity_Category__c': 'Child',
                                                                                         'Name': child1OpportunityName,
                                                                                         'StageName': 'Contract',
                                                                                         'Insert_Date__c': str(futureDate),
                                                                                         'CloseDate': str(futureDate),
                                                                                         'End_Date__c': str(futureDate),
                                                                                         'AccountId': accountID,
                                                                                         'Parent_Opportunity__c': parentOppDetails['id'],
                                                                                         'Probability__c': '75',
                                                                                         'Business_Type__c': 'New',
                                                                                         'Order__c': p1OrderNumber,
                                                                                         'Parent_Order__c': '',
                                                                                         'ILoc_Assest_Due_Date__c': str(futureDate),
                                                                                         'Billed_based_on_Actual_Execution__c': True})
                    if child1OppDetails["success"] == True:   
                        # pdb.set_trace()
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Id", 'id', child1OppDetails['id'], child1OppDetails['id']) 
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CurrencyIsoCode", 'text', str(currencyCode), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeId", 'id', str(opprtunityPLMap[productLine]), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeName", 'string', str(productLine), child1OppDetails['id'])                        
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Name", 'string', str(child1OpportunityName), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "PARENT_OPP_ID", 'id', parentOppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "StageName", 'text', 'Contract', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Insert_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CloseDate", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "End_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "AccountId", 'reference', str(accountID), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Probability__c", 'text', '75', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Business_Type__c", 'text', 'New', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Order__c", 'text', str(p1OrderNumber), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Parent_Order__c", 'text', '', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Assest_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])
                    
                    oppDetails['C_ILoc_Assest_Due_Date'] = str(futureDate)
                    oppDetails['C_Billed_based_on_Actual_Execution'] = True
                elif productLine in ('Merchandising'):
                    projectCode = 600093
                    child1OppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine],
                                                                                         'Opportunity_Category__c': 'Child',
                                                                                         'Name': child1OpportunityName,
                                                                                         'StageName': 'Contract',
                                                                                         'Insert_Date__c': str(futureDate),
                                                                                         'CloseDate': str(futureDate),
                                                                                         'End_Date__c': str(futureDate),
                                                                                         'AccountId': accountID,
                                                                                         'Parent_Opportunity__c': parentOppDetails['id'],
                                                                                         'Probability__c': '75',
                                                                                         'Business_Type__c': 'New',
                                                                                         'Status__c': 'Reserved-RS1',
                                                                                         'Type': 'Subscription',
                                                                                         'Order__c': p1OrderNumber,
                                                                                         'Parent_Order__c': '',
                                                                                         'ILOCProjectCode__c': projectCode,
                                                                                         'ILOC_Store_Type__c': 'Store Type',
                                                                                         'ILOC_Tentative_geography__c': True,
                                                                                         'ILocStoreCount__c': 10,
                                                                                         'ILocTradeClass__c': 'Trade Class',
                                                                                         'ILocRemoveWave__c': True,
                                                                                         'ILocCategory__c': 'AIRTEL, IDEA, JIO, VODAFONE, TATA DOCOMO',
                                                                                         'Job_Description_Comments__c': 'INSTALLATION, MAINTENANACE, OFFERS, FITTING',
                                                                                         'Billed_based_on_Actual_Execution__c': True,
                                                                                         'ILocSignedLOCORStoreListDueDt__c': str(futureDate),
                                                                                         'POS_Materials_Produced_If_yes_above__c': 'NMCIS',
                                                                                         'ILocProgramDocumentDueDate__c': str(futureDate),
                                                                                         'POS_Material_Disposition__c': r"Return to Client at Client's expense",
                                                                                         'ILocArtworkDueDate__c': str(futureDate),
                                                                                         'Network_Retailer_List__c': 'AIRTEL, IDEA, JIO, VODAFONE, TATA DOCOMO',
                                                                                         'ILoc_Material_Due_Date__c': str(futureDate),
                                                                                         'Non_Network_Retailer_List__c': 'AIRTEL, IDEA, JIO, VODAFONE, TATA DOCOMO',
                                                                                         'Client_Supplied_Materials_Due_To__c': 'Other Vendor',
                                                                                         'Image_Request_Details__c': 'Subset List',
                                                                                         'Subset_List_Count__c': 46})
                    if child1OppDetails["success"] == True:   
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Id", 'id', child1OppDetails['id'], child1OppDetails['id'])     
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CurrencyIsoCode", 'text', currencyCode, child1OppDetails['id'])    
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeId", 'id', str(opprtunityPLMap[productLine]), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeName", 'string', str(productLine), child1OppDetails['id'])                        
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Name", 'string', str(child1OpportunityName), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "PARENT_OPP_ID", 'id', parentOppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "StageName", 'text', 'Contract', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Insert_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CloseDate", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "End_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "AccountId", 'reference', str(accountID), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Probability__c", 'text', '75', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Business_Type__c", 'text', 'New', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Status__c", 'text', 'Reserved-RS1', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Type", 'text', 'Subscription', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Order__c", 'text', str(p1OrderNumber), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Parent_Order__c", 'text', '', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Assest_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOC_Store_Type__c", 'text', 'Store Type', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOC_Tentative_geography__c", 'text', 'True', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocStoreCount__c", 'text', '10', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocTradeClass__c", 'text', 'Trade Class', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocRemoveWave__c", 'text', 'True', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocCategory__c", 'string', 'AIRTEL, IDEA, JIO, VODAFONE, TATA DOCOMO', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Job_Description_Comments__c", 'string', 'INSTALLATION, MAINTENANACE, OFFERS, FITTING', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocSignedLOCORStoreListDueDt__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "POS_Materials_Produced_If_yes_above__c", 'text', 'NMCIS', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocProgramDocumentDueDate__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "POS_Material_Disposition__c", 'string', r"Return to Client at Client expense", child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocArtworkDueDate__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Network_Retailer_List__c", 'string', 'AIRTEL, IDEA, JIO, VODAFONE, TATA DOCOMO', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Material_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Non_Network_Retailer_List__c", 'string', 'AIRTEL, IDEA, JIO, VODAFONE, TATA DOCOMO', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Client_Supplied_Materials_Due_To__c", 'text', 'Other Vendor', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Image_Request_Details__c", 'text', 'Image_Request_Details__c', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Subset_List_Count__c", 'text', '46', child1OppDetails['id'])
                    oppDetails['C_STATUS'] = 'Reserved-RS1'
                    oppDetails['C_ILOCProjectCode'] = projectCode
                    oppDetails['C_ILOC_Store_Type'] = 'Store Type'
                    oppDetails['C_ILOC_Tentative_geography'] = True
                    oppDetails['C_ILocStoreCount'] = 10
                    oppDetails['C_ILocTradeClass'] = 'Trade Class'
                    oppDetails['C_ILocRemoveWave'] = True
                    oppDetails['C_ILocCategory'] = 'AIRTEL IDEA JIO VODAFONE TATA DOCOMO'
                    oppDetails['C_Job_Description_Comments'] = 'INSTALLATION MAINTENANACE OFFERS FITTING'
                    oppDetails['C_Billed_based_on_Actual_Execution'] = True
                    oppDetails['C_ILocSignedLOCORStoreListDueDt'] = str(
                        futureDate)
                    oppDetails['C_POS_Materials_Produced_If_yes_above'] = 'NMCIS'
                    oppDetails['C_ILocProgramDocumentDueDate'] = str(
                        futureDate)
                    oppDetails['C_POS_Material_Disposition'] = r"Return to Client at Client's expense"
                    oppDetails['C_ILocArtworkDueDate'] = str(futureDate)
                    oppDetails['C_Network_Retailer_List'] = 'AIRTEL IDEA JIO VODAFONE TATA DOCOMO'
                    oppDetails['C_ILoc_Material_Due_Date'] = str(
                        futureDate)
                    oppDetails['C_Non_Network_Retailer_List'] = 'AIRTEL IDEA JIO VODAFONE TATA DOCOMO'
                    oppDetails['C_Client_Supplied_Materials_Due_To'] = 'Other Vendor'
                    oppDetails['C_Image_Request_Details'] = 'Subset List'
                    oppDetails['C_Subset_List_Count'] = 46
                elif productLine in ('SmartSource Direct', 'SmartSource Direct- Canada'):
                    if productLine in ('SmartSource Direct'):
                        projectCode = 549999
                    elif productLine in ('SmartSource Direct- Canada'):
                        projectCode = 599999
                    child1OppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine],
                                                                                         'Opportunity_Category__c': 'Child',
                                                                                         'Name': child1OpportunityName,
                                                                                         'StageName': 'Contract',
                                                                                         'Insert_Date__c': str(futureDate),
                                                                                         'CloseDate': str(futureDate),
                                                                                         'End_Date__c': str(futureDate),
                                                                                         'AccountId': accountID,
                                                                                         'Parent_Opportunity__c': parentOppDetails['id'],
                                                                                         'Probability__c': '75',
                                                                                         'Business_Type__c': 'New',
                                                                                         'Status__c': 'Reserved-RS1',
                                                                                         'Type': 'Standard',
                                                                                         'Order__c': p1OrderNumber,
                                                                                         'Parent_Order__c': '',
                                                                                         'ILOCProjectCode__c': projectCode,
                                                                                         'ILoc_Format__c': '4BF0D9BD 35F7 4FE6 B',
                                                                                         'ILoc_Targeting__c': 'C4C9F437 0703 4',
                                                                                         'ILocArtworkDueDate__c': str(futureDate),
                                                                                         'ILoc_Sample_Due_Date__c': str(futureDate),
                                                                                         'ILoc_Creative_Due_Date__c': str(futureDate),
                                                                                         'Billed_based_on_Actual_Execution__c': True,
                                                                                         'ILocCategory__c': '70888577, 4615, 4BDF, 8D4C, 350F21B2192D'})
                    if child1OppDetails["success"] == True: 
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Id", 'id', child1OppDetails['id'], child1OppDetails['id'])     
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CurrencyIsoCode", 'text', currencyCode, child1OppDetails['id'])    
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeId", 'id', str(opprtunityPLMap[productLine]), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeName", 'string', str(productLine), child1OppDetails['id'])                        
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Name", 'string', str(child1OpportunityName), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "PARENT_OPP_ID", 'id', parentOppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "StageName", 'text', 'Contract', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CloseDate", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "End_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Insert_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "AccountId", 'reference', str(accountID), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Probability__c", 'text', '75', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Business_Type__c", 'text', 'New', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Status__c", 'text', 'Reserved-RS1', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Type", 'text', 'Standard', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Order__c", 'text', str(p1OrderNumber), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Parent_Order__c", 'text', '', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Format__c", 'text',  '4BF0D9BD 35F7 4FE6 B', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOCProjectCode__c", 'text', str(projectCode), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Targeting__c", 'text', 'C4C9F437 0703 4', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocArtworkDueDate__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Sample_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Creative_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])                    
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocCategory__c", 'string', '70888577, 4615, 4BDF, 8D4C, 350F21B2192D', child1OppDetails['id'])

                    oppDetails['C_STATUS'] = 'Reserved-RS1'
                    oppDetails['C_ILOCProjectCode'] = projectCode
                    oppDetails['C_ILoc_Format'] = '4BF0D9BD 35F7 4FE6 B'
                    oppDetails['C_ILoc_Targeting'] = 'C4C9F437 0703 4'
                    oppDetails['C_ILocArtworkDueDate'] = str(
                        futureDate)
                    oppDetails['C_ILoc_Sample_Due_Date'] = str(
                        futureDate)
                    oppDetails['C_ILoc_Creative_Due_Date'] = str(
                        futureDate)
                    oppDetails['C_Billed_based_on_Actual_Execution'] = True
                    oppDetails['C_ILocCategory'] = '70888577 4615 4BDF 8D4C 350F21B2192D'
                elif productLine in ('Digital', 'Digital- Canada'):
                    child1OppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine],
                                                                                         'Opportunity_Category__c': 'Child',
                                                                                         'Name': child1OpportunityName,
                                                                                         'StageName': 'Contract',
                                                                                         'Insert_Date__c': str(futureDate),
                                                                                         'CloseDate': str(futureDate),
                                                                                         'End_Date__c': str(futureDate),
                                                                                         'AccountId': accountID,
                                                                                         'Parent_Opportunity__c': parentOppDetails['id'],
                                                                                         'Probability__c': '75',
                                                                                         'Business_Type__c': 'New',
                                                                                         'Status__c': 'Reserved-RS1',
                                                                                         'Type': 'Standard',
                                                                                         'Order__c': p1OrderNumber,
                                                                                         'Parent_Order__c': '',
                                                                                         'Artwork_Due_Date__c': str(futureDate),
                                                                                         'ClientListDueDate__c': str(futureDate),
                                                                                         'Billed_based_on_Actual_Execution__c': True,
                                                                                         'ILocCategory__c': 'AIRTEL IDEA JIO VODAFONE TATA DOCOMO',
                                                                                         'ILOC_Client_s_clip_rate__c': 10,
                                                                                         'ILOCProjectCode__c': 7755001})

                    if child1OppDetails["success"] == True: 
                        # pdb.set_trace()
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Id", 'id', child1OppDetails['id'], child1OppDetails['id'])     
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CurrencyIsoCode", 'text', currencyCode, child1OppDetails['id'])    
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeId", 'id', str(opprtunityPLMap[productLine]), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeName", 'string', str(productLine), child1OppDetails['id'])                        
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Name", 'string', str(child1OpportunityName), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "PARENT_OPP_ID", 'id', parentOppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "StageName", 'text', 'Contract', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CloseDate", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "End_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Insert_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "AccountId", 'reference', str(accountID), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Probability__c", 'text', '75', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Business_Type__c", 'text', 'New', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Status__c", 'text', 'Reserved-RS1', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Type", 'text', 'Standard', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Order__c", 'text', str(p1OrderNumber), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Parent_Order__c", 'text', '', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOC_Client_s_clip_rate__c", 'text', '10', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOCProjectCode__c", 'text', '7755001', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Artwork_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ClientListDueDate__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])                    
                    oppDetails['C_STATUS'] = 'Reserved-RS1'
                    oppDetails['C_Artwork_Due_Date'] = str(futureDate)
                    oppDetails['C_ClientListDueDate'] = str(futureDate)
                    oppDetails['C_Billed_based_on_Actual_Execution'] = True
                    oppDetails['C_ILocCategory'] = 'AIRTEL IDEA JIO VODAFONE TATA DOCOMO'
                    oppDetails['C_ILOC_Client_s_clip_rate'] = 10
                    oppDetails['C_ILOCProjectCode'] = 7755001
                print(child1OppDetails)
                if child1OppDetails["success"] == True:
                    print(f"Child1 Opp ID of {productLine}: ",
                          child1OppDetails['id'])
                    oppDetails['C_ID'] = child1OppDetails['id']
                    oppDetails['C_NAME'] = child1OpportunityName
                    oppDetails['C_ACC_ID'] = accountID
                    oppDetails['C_STAGE'] = 'Contract'

                    usProductLineMap[productLine]['PARENT_OPP_1']['CHILD1_OPP_ID']['ID'] = child1OppDetails['id']
                    usProductLineMap[productLine]['PARENT_OPP_1']['CHILD1_OPP_ID']['ORDER#'] = str(
                        p1OrderNumber)
                    childOpps.append(child1OppDetails['id'])
                    Messages.write_message(
                        f"Child 1 Opp ID of {productLine}: {child1OppDetails['id']}")
                    data_store.spec[f"P1_C1_ID_{productLine}"] = child1OppDetails['id']
                    data_store.spec[f"P1_C1_NAME_{productLine}"] = child1OpportunityName

            print("\nChild Opportunities\n", childOpps)
# -------------------------------------------------------------------------------------------------------------------------------------
            productsList = []
            productsList.clear()
            productName = None
            if productLine in ('InStore', 'InStore-Canada'):
                oppPLProductsSOQL = f"SELECT Id,Product_Line__c, Product_Line__r.Name,Product__c, Product__r.Name FROM Product_Junction__c where Product_Line__r.Name = 'InStore' and Product__r.IsActive = true and Product__c in (SELECT Product2Id from PricebookEntry where CurrencyIsoCode = '{currencyCode}' and IsActive = true)"
            else:
                oppPLProductsSOQL = f"SELECT Id,Product_Line__c, Product_Line__r.Name,Product__c, Product__r.Name FROM Product_Junction__c where Product_Line__r.Name = '{productLine}' and Product__r.IsActive = true and Product__c in (SELECT Product2Id from PricebookEntry where CurrencyIsoCode = '{currencyCode}' and IsActive = true)"
        #         oppPLProductsSOQL = f"SELECT Id,Product_Line__c, Product_Line__r.Name,Product__c, Product__r.Name FROM Product_Junction__c where Product_Line__r.Name = '{productLine}'"
            oppPLProductsResult = AccountOpportunityCreation.sf.query_all(
                query=oppPLProductsSOQL)
            oppPLProductsData = oppPLProductsResult['records']

            for oppPLProduct in oppPLProductsData:
                if oppPLProduct['Product__r'] != None:
                    if 'Name' in oppPLProduct['Product__r']:
                        productsList.append(oppPLProduct['Product__r']['Name'])
                        print("Found Product: ",
                              oppPLProduct['Product__r']['Name'])
            print(*productsList, sep="\n")
# -------------------------------------------------------------------------------------------------------------------------------------
            lineItemsMap = {}
            # pdb.set_trace()

            for idx, childOpp in enumerate(childOpps):
                IsActive = False
                while IsActive == False:
                    productName = random.choice(productsList)
                    productNameTemp = productName.replace("'", r"\'")

                    Messages.write_message(f"Product Name: {productName}")
                    data_store.spec[f"PRODUCT_NAME"] = productName
                    oppDetails['C_PRODUCT_NAME'] = productName

                    print("Product Name: ", productName)
                    # priceBookSOQL = f"SELECT Id, Name,IsActive FROM PricebookEntry where Name = '{productName}' and CurrencyIsoCode = '{currencyCode}' and IsActive = true"
                    priceBookSOQL = f"SELECT Id, IsActive FROM PricebookEntry WHERE CurrencyIsoCode = '{currencyCode}' AND IsActive = True AND Product2.Name = '{productNameTemp}' AND Product2Id in (SELECT Product__c FROM Product_Junction__c WHERE Product_Line__r.Name = '{productLine}')"
                    priceBookResult = AccountOpportunityCreation.sf.query_all(
                        query=priceBookSOQL)
                    priceBookData = priceBookResult['records']
                    if priceBookData[0]['IsActive'] != False:
                        priceBookId = priceBookData[0]['Id']
                        Messages.write_message(
                            f"Pricebook Entry ID: {priceBookId}")
                        print("Pricebook Entry ID: ", priceBookId)
                        data_store.spec["PRICE_BOOK_ID"] = priceBookId

                        if productLine in ('InStore', 'InStore-Canada'):
                            chargeDetailsSOQL = f"SELECT Id, Name, ProductLIne__c, Product__C, Charge_Type_2__c, Charge_Type_2__r.Commissionable__c, Charge_Type_2__r.Charge_Type_Categorty__c FROM Pricing_Detail__c where ProductLIne__c = 'Instore' and product__c = '{productNameTemp}' and CurrencyIsoCode = '{currencyCode}' and isActive__c = true and IsCurrent__c = true and RateCardIsCurrent__c = true"
                        elif productLine in ('Checkout 51'):
                            chargeDetailsSOQL = f"SELECT Id, Name, ProductLIne__c, Product__C, Charge_Type_2__c, Charge_Type_2__r.Commissionable__c, Charge_Type_2__r.Charge_Type_Categorty__c FROM Pricing_Detail__c where ProductLIne__c = '{productLine}' and FreedomOrNot__c != 'Freedom' and product__c = '{productNameTemp}' and CurrencyIsoCode = '{currencyCode}' and isActive__c = true and IsCurrent__c = true and RateCardIsCurrent__c = true"
                        elif productLine in ('Digital', 'Digital- Canada'):
                            chargeDetailsSOQL = f"SELECT Id, Name, ProductLIne__c, Product__C, Charge_Type_2__c, Charge_Type_2__r.Commissionable__c, Charge_Type_2__r.Charge_Type_Categorty__c FROM Pricing_Detail__c where ProductLIne__c = '{productLine}' and product__c = '{productNameTemp}' and CurrencyIsoCode = '{currencyCode}' and isActive__c = true and FreedomOrNot__c not in ('Freedom') and IsCurrent__c = true and RateCardIsCurrent__c = true"
                        else:
                            chargeDetailsSOQL = f"SELECT Id, Name, ProductLIne__c, Product__C, Charge_Type_2__c, Charge_Type_2__r.Commissionable__c, Charge_Type_2__r.Charge_Type_Categorty__c FROM Pricing_Detail__c where ProductLIne__c = '{productLine}' and product__c = '{productNameTemp}' and CurrencyIsoCode = '{currencyCode}' and isActive__c = true and IsCurrent__c = true and RateCardIsCurrent__c = true"
                        chargeDetailsResult = AccountOpportunityCreation.sf.query_all(
                            query=chargeDetailsSOQL)
                        chargeDetailsData = chargeDetailsResult['records']
                        print("\n", chargeDetailsSOQL, "\n")
                        if len(chargeDetailsData) > 2:
                            chargeTypeMap = {}
                            chargeTypeMap.clear()
                            for chargeDetails in chargeDetailsData:
                                chargeTypeMap[chargeDetails['Name']] = {"Id": chargeDetails['Id'], "ChargeType": chargeDetails[
                                    'Charge_Type_2__c'], 'Commissionable': chargeDetails['Charge_Type_2__r']['Commissionable__c'], 'Charge_Type_Category': chargeDetails['Charge_Type_2__r']['Charge_Type_Categorty__c']}
                            # print("\nCharge Types\n", chargeTypeMap)
                            [print(key, value)
                             for key, value in chargeTypeMap.items()]
                            Messages.write_message(
                                f"Charge Types MAP \n{chargeTypeMap}")
                            IsActive = True

                totalAmount = 0

                for cnt in range(1, 5):
                    chargeType = random.choice(list(chargeTypeMap))
                    chargeTypeTemp = chargeType.replace("'", r"\'")
                    soql = f"SELECT Charge_Type_2__r.Commissionable__c,Charge_Type__c, Charge_Type_Category__c, CurrencyIsoCode, Id, isActive__c, Name, Product__c, Quantity_1000_effective_from_date__c, Quantity_1000__c FROM Pricing_Detail__c WHERE Name = '{chargeTypeTemp}' and Product__c= '{productNameTemp}' and CurrencyIsoCode = '{currencyCode}' and isActive__c = True and RateCardIsCurrent__c = true and IsCurrent__c = true"
                    queryResult = AccountOpportunityCreation.sf.query_all(
                        query=soql)
                    recDetails = queryResult['records']
                    quantity = randint(1, 100)
                    salesPrice = round(random.uniform(1.1, 9.9), 2)
                    if recDetails[0]['Quantity_1000__c'] == True:
                        totalPrice = round((quantity / 1000) * salesPrice, 2)
                    else:
                        totalPrice = round((quantity * salesPrice), 2)
                    totalAmount = totalAmount + totalPrice

                    # pdb.set_trace()
                    lineItemsDetails = AccountOpportunityCreation.sf.OpportunityLineItem.create({'Charge_Type__c': chargeTypeMap[chargeType]["ChargeType"], 'Commissionable__c': chargeTypeMap[chargeType]["Commissionable"], 'pricebookentryid': priceBookId,
                                                                                                 'Pricing_Detail__c': chargeType, 'PricingDetail__c': chargeTypeMap[chargeType]["Id"], 'opportunityId': childOpp, 'Quantity': quantity, 'Sales_price__c': salesPrice, 'TotalPrice': totalPrice, 'Component__c': 'Video'})

                    # oppDetails[f'C_{cnt}_CHARGE_TYPE'] = chargeTypeMap[chargeType]["ChargeType"]
                    # oppDetails[f'C_{cnt}_CHARGE_TYPE_CATEGORY'] = chargeTypeMap[chargeType]["Charge_Type_Category"]
                    # oppDetails[f'C_{cnt}_COMMISSIONABLE'] = chargeTypeMap[chargeType]["Commissionable"]
                    # oppDetails[f'C_{cnt}_PRICING_DETAIL_NAME'] = chargeType
                    # oppDetails[f'C_{cnt}_PRICING_DETAIL_ID'] = lineItemsDetails["id"]
                    # oppDetails[f'C_{cnt}_QTY'] = quantity
                    # oppDetails[f'C_{cnt}_SALES_PRICE'] = salesPrice
                    # oppDetails[f'C_{cnt}_SUB_TOTAL'] = totalPrice

                    if lineItemsDetails["success"] == True:
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Id", 'id', lineItemsDetails["id"], lineItemsDetails["id"])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Opportunity__c", 'reference', child1OppDetails['id'], lineItemsDetails["id"])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "product__c", 'string', productNameTemp, lineItemsDetails["id"])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "ProductLIne__c", 'string', productLine, lineItemsDetails["id"])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "CurrencyIsoCode", 'text', currencyCode, lineItemsDetails["id"])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "priceBookId", 'string', priceBookId, lineItemsDetails["id"])
                        oppDetails[f'C_{cnt}_CHARGE_TYPE'] = chargeTypeMap[chargeType]["ChargeType"]
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Charge_Type__c", 'string', chargeTypeMap[chargeType]["ChargeType"], lineItemsDetails["id"])
                        oppDetails[f'C_{cnt}_CHARGE_TYPE_CATEGORY'] = chargeTypeMap[chargeType]["Charge_Type_Category"]
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Charge_Type_Category__c", 'string', chargeTypeMap[chargeType]["Charge_Type_Category"], lineItemsDetails["id"]) 
                        oppDetails[f'C_{cnt}_COMMISSIONABLE'] = chargeTypeMap[chargeType]["Commissionable"]
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Commissionable__c", 'text', chargeTypeMap[chargeType]["Commissionable"], lineItemsDetails["id"])
                        oppDetails[f'C_{cnt}_PRICING_DETAIL_NAME'] = chargeType
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Pricing_Detail__c", 'string', chargeType, lineItemsDetails["id"])
                        oppDetails[f'C_{cnt}_PRICING_DETAIL_ID'] = lineItemsDetails["id"]
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "PricingDetail__c", 'reference', chargeTypeMap[chargeType]["Id"], lineItemsDetails["id"])
                        oppDetails[f'C_{cnt}_QTY'] = quantity
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Quantity", 'double', quantity, lineItemsDetails["id"])
                        oppDetails[f'C_{cnt}_SALES_PRICE'] = salesPrice
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Sales_price__c", 'double', salesPrice, lineItemsDetails["id"])
                        oppDetails[f'C_{cnt}_SUB_TOTAL'] = totalPrice
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "TotalPrice", 'double', totalPrice, lineItemsDetails["id"])
                        if recDetails[0]['Quantity_1000__c'] == True:
                            Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Quantity_1000__c", 'text', recDetails[0]['Quantity_1000__c'], lineItemsDetails["id"])
                            Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Quantity_1000_effective_from_date__c", 'text', recDetails[0]['Quantity_1000_effective_from_date__c'], lineItemsDetails["id"])
                        print(
                            f"OPP[{childOpp}] Line Item {cnt}: {lineItemsDetails['id']} created..")
                        Messages.write_message(
                            f"OPP[{childOpp}] Line Item {cnt}: {lineItemsDetails['id']} created..")
                        lineItemsMap[childOpp +
                                     str(cnt)] = {cnt: lineItemsDetails['id']}
                oppDetails['C_TOTAL_AMOUNT'] = totalAmount
                # Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Total", 'currency', totalAmount, child1OppDetails['id'])
                if productLine in ('FSI') and productName != 'Remnant':
                    Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Total", 'currency', '8000', child1OppDetails['id'])
                else:
                    Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Total", 'currency', totalAmount, child1OppDetails['id'])

            columnCount = 0
            rowCount = 0

            df = pd.read_excel(opportunitiesDetailsFileName,
                               sheet_name=f"{productLine}_{currencyCode}")
            Messages.write_message(df.columns)
            columnList = df.columns.tolist()
            [print(key, value) for key, value in oppDetails.items()]
            rowCount = int(ws.max_row) + 1
            for key, value in oppDetails.items():
                #         pdb.set_trace()
                if key in columnList:
                    ws[f"{xlsxwriter.utility.xl_col_to_name(columnList.index(key))}{rowCount}"] = value
                else:
                    if len(columnList) == 0 and ws["A1"].value == None:
                        ws[f"{xlsxwriter.utility.xl_col_to_name(columnCount)}{1}"] = key
                        ws[f"{xlsxwriter.utility.xl_col_to_name(columnCount)}{2}"] = value
                    else:
                        columnCount = int(ws.max_column)
                        ws[f"{xlsxwriter.utility.xl_col_to_name(columnCount)}{1}"] = key
                        ws[f"{xlsxwriter.utility.xl_col_to_name(columnCount)}{rowCount}"] = value
                wb.save(filename=opportunitiesDetailsFileName)
            wb.save(filename=opportunitiesDetailsFileName)
# --------------------------------------------------------------------------------------------------------------------------
            usersVerified = []
            usersNotVerified = []

            # Verify Opportunities Team
            accountTeamMap = {}
            # '{accID}'"
            accountMembersSOQL = f"SELECT Account__r.Acc_Owner_Terr_Cat__c,Role_In_Territory__c ,TerritoryId__c,Territory_Category__c,User__c,User__r.Name FROM Account_Team__c where Account__c = '{accountID}'"
            accountMembersResult = AccountOpportunityCreation.sf.query_all(
                query=accountMembersSOQL)
            accountMembersData = accountMembersResult['records']
            if len(accountMembersData) > 0:
                for accountMembers in accountMembersData:
                    accountTeamMap[accountMembers["User__c"]] = {"Role_In_Territory__c": accountMembers["Role_In_Territory__c"], "TerritoryId__c": accountMembers["TerritoryId__c"],
                                                                 "Territory_Category__c": accountMembers["Territory_Category__c"], "Name": accountMembers["User__r"]["Name"], "Acc_Owner_Terr_Cat__c": accountMembers["Account__r"]["Acc_Owner_Terr_Cat__c"]}

            for childOpp in childOpps:
                opportunityMembersSOQL = f"SELECT Opportunity__r.RecordType.Name,Id,TeamMemberRole__c,TerritoryId__c,Territory_Category__c,User__c,User__r.Name FROM Opportunity_Team__c where Opportunity__c = '{childOpp}'"
                opportunityMembersResult = AccountOpportunityCreation.sf.query_all(
                    query=opportunityMembersSOQL)
                opportunityMembersData = opportunityMembersResult['records']
                if len(opportunityMembersData) > 0:
                    for opportunityMembers in opportunityMembersData:
                        isMemberVerified = False
                        for accountMembers in accountTeamMap:
                            if not isMemberVerified:
                                #                         pdb.set_trace()
                                print("\nVerifying Users", opportunityMembers['User__r']['Name'],
                                      "\t", accountTeamMap[opportunityMembers['User__c']]['Name'])
                #                 if (accountTeamMap['Role_In_Territory__c'] == territoryMembers['RoleInTerritory2']) and (accountTeamMap['TerritoryId__c'] == territoryMembers['Territory2Id']) and (accountTeamMap['User__c'] == territoryMembers['UserId']):
                                if opportunityMembers["User__r"]["Name"] == accountTeamMap[opportunityMembers["User__c"]]["Name"]:
                                    print(
                                        "\nMember Verified: ", accountTeamMap[opportunityMembers['User__c']]['Name'], "\t", opportunityMembers['User__r']['Name'])
                                    if accountTeamMap[opportunityMembers['User__c']]['Acc_Owner_Terr_Cat__c'] == 'Core':
                                        if opportunityMembers['Opportunity__r']['RecordType']['Name'] in ('FSI', 'InStore', 'SSMG'):
                                            if opportunityMembers['Territory_Category__c'] in ('Core') and opportunityMembers['TeamMemberRole__c'] in ('Primary'):
                                                print("\n Territory Category Verified: ", opportunityMembers['Opportunity__r'][
                                                    'RecordType']['Name'], "\t", opportunityMembers['TeamMemberRole__c'])
                                            elif opportunityMembers['Territory_Category__c'] in ('Checkout_51', 'Digital', 'Merchandising', 'SSD') and opportunityMembers['TeamMemberRole__c'] in ('Specialty', 'Integrated'):
                                                print("\n Territory Category Verified: ", opportunityMembers['Opportunity__r'][
                                                    'RecordType']['Name'], "\t", opportunityMembers['TeamMemberRole__c'])
                                    elif accountTeamMap[opportunityMembers['User__c']]['Acc_Owner_Terr_Cat__c'] == 'Checkout_51':
                                        if opportunityMembers['Opportunity__r']['RecordType']['Name'] in ('Checkout 51'):
                                            if opportunityMembers['Territory_Category__c'] in ('Checkout_51') and opportunityMembers['TeamMemberRole__c'] in ('Primary'):
                                                print("\n Territory Category Verified: ", opportunityMembers['Opportunity__r'][
                                                    'RecordType']['Name'], "\t", opportunityMembers['TeamMemberRole__c'])
                                    elif accountTeamMap[opportunityMembers['User__c']]['Acc_Owner_Terr_Cat__c'] == 'Digital':
                                        if opportunityMembers['Opportunity__r']['RecordType']['Name'] in ('Digital', 'Digital- Canada'):
                                            if opportunityMembers['Territory_Category__c'] in ('Digital') and opportunityMembers['TeamMemberRole__c'] in ('Primary'):
                                                print("\n Territory Category Verified: ", opportunityMembers['Opportunity__r'][
                                                    'RecordType']['Name'], "\t", opportunityMembers['TeamMemberRole__c'])
                                    elif accountTeamMap[opportunityMembers['User__c']]['Acc_Owner_Terr_Cat__c'] == 'Merchandising':
                                        if opportunityMembers['Opportunity__r']['RecordType']['Name'] in ('Merchandising'):
                                            if opportunityMembers['Territory_Category__c'] in ('Merchandising') and opportunityMembers['TeamMemberRole__c'] in ('Primary'):
                                                print("\n Territory Category Verified: ", opportunityMembers['Opportunity__r'][
                                                    'RecordType']['Name'], "\t", opportunityMembers['TeamMemberRole__c'])
                                    elif accountTeamMap[opportunityMembers['User__c']]['Acc_Owner_Terr_Cat__c'] == 'NTL':
                                        if opportunityMembers['Territory_Category__c'] in ('NTL') and opportunityMembers['TeamMemberRole__c'] in ('Primary'):
                                            print("\n Territory Category Verified: ", opportunityMembers['Opportunity__r'][
                                                'RecordType']['Name'], "\t", opportunityMembers['TeamMemberRole__c'])
                                    elif accountTeamMap[opportunityMembers['User__c']]['Acc_Owner_Terr_Cat__c'] == 'SM':
                                        if opportunityMembers['Territory_Category__c'] in ('SM') and opportunityMembers['TeamMemberRole__c'] in ('Primary'):
                                            print("\n Territory Category Verified: ", opportunityMembers['Opportunity__r'][
                                                'RecordType']['Name'], "\t", opportunityMembers['TeamMemberRole__c'])
                                    elif accountTeamMap[opportunityMembers['User__c']]['Acc_Owner_Terr_Cat__c'] == 'SSD':
                                        if opportunityMembers['Opportunity__r']['RecordType']['Name'] in ('SmartSource Direct', 'SmartSource Direct- Canada'):
                                            if opportunityMembers['Territory_Category__c'] in ('SSD') and opportunityMembers['TeamMemberRole__c'] in ('Primary'):
                                                print("\n Territory Category Verified: ", opportunityMembers['Opportunity__r'][
                                                    'RecordType']['Name'], "\t", opportunityMembers['TeamMemberRole__c'])
                                    elif accountTeamMap[opportunityMembers['User__c']]['Acc_Owner_Terr_Cat__c'] == 'SSMG':
                                        if opportunityMembers['Opportunity__r']['RecordType']['Name'] in ('SSMG'):
                                            if opportunityMembers['Territory_Category__c'] in ('SSMG') and opportunityMembers['TeamMemberRole__c'] in ('Primary'):
                                                print("\n Territory Category Verified: ", opportunityMembers['Opportunity__r']['RecordType'][
                                                    'Name'], "\t", accountTeamMap[opportunityMembers['User__c']]['Territory_Category__c'])

                                    if opportunityMembers['User__r']['Name'] not in usersVerified:
                                        usersVerified.append(
                                            opportunityMembers['User__r']['Name'])

                                    if opportunityMembers['User__r']['Name'] in usersNotVerified:
                                        usersNotVerified.remove(
                                            opportunityMembers['User__r']['Name'])
                                    isMemberVerified = True

                                else:
                                    #                 print("Member Not Verified: ", accountTeamMap['User__r']['Name'], "\t", accountTeamMap['Role_In_Territory__c'])
                                    if opportunityMembers['User__r']['Name'] not in usersNotVerified:
                                        usersNotVerified.append(
                                            opportunityMembers['User__r']['Name'])

                                    if opportunityMembers['User__r']['Name'] in usersVerified:
                                        usersVerified.remove(
                                            opportunityMembers['User__r']['Name'])
                                    isMemberVerified = False

                            print("\nMembers Verified: ", usersVerified)
                            print("\nMembers not verified: ", usersNotVerified)

            print("\nUS Product Lines\n", usProductLineMap)
            jsonData = json.dumps(usProductLineMap)
            with open("OpportunitiesDetails.json", 'w') as f:
                f.write(jsonData)
            print(
                f"{productLine} completed---------------------------------------------------------------------------\n")

    @step("Create <oppName> Opportunity for <productLineRegion> productline using <spQty>")
    def create_opportunity_for_productline(self, oppName, productLineRegion, spQty):

        today = date.today()
        oppDateToday = today.strftime("%Y%m%d")
        opportunityName = f"{oppName}{oppDateToday}"

        queryData = (AccountOpportunityCreation.sf.query(format_soql(
            "SELECT Id, Name FROM Opportunity WHERE Name = {} AND Opportunity_Category__c = 'Child'", opportunityName)))['records']

        if not len(queryData) > 0:
            region = productLineRegion.split("_")[1]
            productLine = productLineRegion.split("_")[0]
            currencyCode = region
            opportunitiesDetailsPath = Path(__file__).parents[1]
            opportunitiesDetailsFileName = str(
                opportunitiesDetailsPath) + "\\Data\\" + os.getenv("ILOC_OPPORTUNITY_DETAILS_FILE")
            wb = None
            ws = None

            if os.path.exists(opportunitiesDetailsFileName):
                wb = load_workbook(filename=opportunitiesDetailsFileName)
            else:
                wb = Workbook()
                wb.save(opportunitiesDetailsFileName)

            end = datetime.datetime.now()
            current_month = end.strftime('%m')
            current_day = end.strftime('%d')
            current_year_full = end.strftime('%Y')

            futureDate = datetime.datetime(int(current_year_full) + 1,
                                           int(current_month), int(current_day)).date()
            print("Future Date: ", futureDate)
            userData = AccountOpportunityCreation.sf.quick_search(
                f"{os.getenv('USER_ID')}")
            userId = userData['searchRecords'][0]['Id']
            print(userData['searchRecords'][0]['Id'])

            napiDateId, napiDate = AccountOpportunityCreation.getNAPIDate(
                self, AccountOpportunityCreation.sf, "Future")

            instoreCycle, instoreCycleName = AccountOpportunityCreation.getInstoreCycle(
                self, AccountOpportunityCreation.sf, f"{str(int(current_year_full)+1)}{current_month}", current_month, current_year_full)

            usProductLineMap = {}
            productLineList = None

            accountID = os.getenv(
                f'OPPORUNITIES_TO_CREATE_FOR_{region}_ACCOUNT_ID')

            opprtunityPLMap = {}
            opprtunityPLSOQL = f"select Id,Name from RecordType where sObjectType='Opportunity'"
            opprtunityPLResult = AccountOpportunityCreation.sf.query_all(
                query=opprtunityPLSOQL)
            opprtunityPLData = opprtunityPLResult['records']

            for opprtunityPL in opprtunityPLData:
                opprtunityPLMap[opprtunityPL['Name']] = opprtunityPL['Id']

            [print(key, value) for key, value in opprtunityPLMap.items()]

            childOpps = []
            oppDetails = {}
            oppDetails.clear()
            if f"{productLine}_{currencyCode}" in wb.sheetnames:
                ws = wb[f"{productLine}_{currencyCode}"]
            else:
                ws = wb.create_sheet(f"{productLine}_{currencyCode}", 1)

            wb.save(filename=opportunitiesDetailsFileName)

            df = pd.read_excel(opportunitiesDetailsFileName,
                               sheet_name=f"{productLine}_{currencyCode}")
            Messages.write_message(df.columns)
            columnList = df.columns.tolist()

            rowCount = int(df.shape[0])
            colCount = int(df.shape[1])
            # pdb.set_trace()
            print(
                f"\n{productLine} Started---------------------------------------------------------------------------\n")
            usProductLineMap[productLine] = {
                'PRODUCT_LINE_ID': opprtunityPLMap[productLine], 'PRODUCT_LINE_NAME': productLine, 'PARENT_OPP': ''}

            oppDateToday = today.strftime("%Y%m%d")
            parentOpportunityName = f"{oppName}{oppDateToday}#Parent"
            if productLine in ('InStore', 'InStore-Canada'):
                parentOppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine], 'Opportunity_Category__c': 'Parent', 'Name': parentOpportunityName,
                                                                                     'StageName': 'Contract', 'InStore_Cycle__c': str(instoreCycle), 'CloseDate': str(futureDate), 'AccountId': accountID, 'ILocArtworkDueDate__c': str(futureDate)})
            elif productLine in ('FSI'):
                parentOppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine], 'Opportunity_Category__c': 'Parent', 'Name': parentOpportunityName,
                                                                                     'StageName': 'Contract', 'NAPI_Insert_date__c': str(napiDateId), 'CloseDate': str(napiDate), 'AccountId': accountID})
        #     elif productLine in ('Digital','Digital- Canada'):
        #         parentOppDetails = sf.Opportunity.create({'RecordTypeId' : opprtunityPLMap[productLine],'Opportunity_Category__c' : 'Parent', 'Name' : parentOpportunityName,'StageName' : 'Contract','Insert_Date__c' : str(futureDate),'CloseDate' : str(futureDate), 'End_Date__c' : str(futureDate), 'AccountId': accountID})
            else:
                parentOppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine], 'Opportunity_Category__c': 'Parent', 'Name': parentOpportunityName,
                                                                                     'StageName': 'Contract', 'Insert_Date__c': str(futureDate), 'CloseDate': str(futureDate), 'End_Date__c': str(futureDate), 'AccountId': accountID})

            print(parentOppDetails)
            if parentOppDetails["success"] == True:
                print(f"Parent Opp ID of {productLine}: ",
                      parentOppDetails['id'])
                Messages.write_message(
                    f"Parent Opp ID of {productLine}: {parentOppDetails['id']}")

                data_store.spec[f"1_PO_ID_{productLine}"] = parentOppDetails['id']
                data_store.spec[f"1_PO_NAME_{productLine}"] = parentOpportunityName
                oppDetails['P_ID'] = parentOppDetails['id']
                oppDetails['P_NAME'] = parentOpportunityName

                usProductLineMap[productLine]['PARENT_OPP_1'] = {
                    'ID': parentOppDetails['id'], 'CHILD1_OPP_ID': '', 'CHILD2_OPP_ID': ''}
                usProductLineMap[productLine]['PARENT_OPP_1']['CHILD1_OPP_ID'] = {
                    'ID': '', 'ORDER#': '', 'PARENT_ORDER#': ''}
                usProductLineMap[productLine]['PARENT_OPP_1']['CHILD2_OPP_ID'] = {
                    'ID': '', 'ORDER#': '', 'PARENT_ORDER#': ''}

                # Write data to database                


            for i in range(1, 2):
                df = pd.read_excel(opportunitiesDetailsFileName,
                                   sheet_name=f"{productLine}_{currencyCode}")
                rowCount = int(df.shape[0])
                colCount = int(df.shape[1])
                today = date.today()
                oppDateToday = today.strftime("%Y%m%d")
                child1OpportunityName = f"{oppName}{oppDateToday}"
                p1OrderNumber = str(uuid.uuid4()).upper()[:10]

                if productLine in ('InStore', 'InStore-Canada'):
                    if productLine in ('InStore'):
                        projectCode = 300000
                    elif productLine in ('InStore-Canada'):
                        projectCode = 399999

                    child1OppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine],
                                                                                         'Opportunity_Category__c': 'Child',
                                                                                         'Name': child1OpportunityName,
                                                                                         'StageName': 'Contract',
                                                                                         'InStore_Cycle__c': str(instoreCycle),
                                                                                         'CloseDate': str(futureDate),
                                                                                         'End_Date__c': str(futureDate),
                                                                                         'Artwork_Due_Date__c': str(futureDate),
                                                                                         'AccountId': accountID,
                                                                                         'Parent_Opportunity__c': parentOppDetails['id'],
                                                                                         'Probability__c': '75',
                                                                                         'Estimated_Average_CPS__c': 1,
                                                                                         'Estimated_Store_Count__c': 1,
                                                                                         'Business_Type__c': 'New',
                                                                                         'Status__c': 'Reserved-RS1',
                                                                                         'Type': 'Tactical',
                                                                                         'Order__c': p1OrderNumber,
                                                                                         'Parent_Order__c': '',
                                                                                         'ILocArtworkDueDate__c': str(futureDate),
                                                                                         'ILocStoreCount__c': 10,
                                                                                         'ILocTradeClass__c': 'TRADE CLASS',
                                                                                         'ILocCategory__c': 'CATEGORY1, CATEGORY2, CATEGORY3',
                                                                                         'ILocBrand__c': 'BRAND',
                                                                                         'ILocLocType__c': 'TYPE',
                                                                                         'ILocProgram__c': 'PROGRAM',
                                                                                         'ILocGeography__c': '20',
                                                                                         'ILOCProjectCode__c': projectCode,
                                                                                         'Billed_based_on_Actual_Execution__c': True})
                    oppDetails['C_CYCLE'] = str(instoreCycleName)
                    oppDetails['C_STATUS'] = 'Reserved-RS1'
                    oppDetails['C_ILocStoreCount'] = 10
                    oppDetails['C_ILocTradeClass'] = 'TRADE CLASS'
                    oppDetails['C_ILocCategory'] = 'CATEGORY1, CATEGORY2, CATEGORY3'
                    oppDetails['C_ILocBrand'] = 'BRAND'
                    oppDetails['C_ILocLocType'] = 'TYPE'
                    oppDetails['C_ILocProgram'] = 'PROGRAM'
                    oppDetails['C_ILocGeography'] = '20'
                    oppDetails['C_ILOCProjectCode'] = projectCode
                    oppDetails['C_Billed_based_on_Actual_Execution'] = True
                    
                    if child1OppDetails["success"] == True:
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Id", 'id', child1OppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CurrencyIsoCode", 'text', currencyCode, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeId", 'id', str(opprtunityPLMap[productLine]), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeName", 'string', str(productLine), child1OppDetails['id'])                        
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Name", 'string', str(child1OpportunityName), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "PARENT_OPP_ID", 'id', parentOppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "StageName", 'text', 'Contract', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "InStore_Cycle__c", 'text', str(instoreCycle), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CloseDate", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "End_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Artwork_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "AccountId", 'reference', str(accountID), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Probability__c", 'text', '75', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Estimated_Average_CPS__c", 'currency', 1, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Estimated_Store_Count__c", 'currency', 1, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Business_Type__c", 'text', 'New', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Status__c", 'text', 'Reserved-RS1', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Type", 'text', 'Tactical', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Order__c", 'text', str(p1OrderNumber), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Parent_Order__c", 'text', '', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocArtworkDueDate__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocStoreCount__c", 'text', '1', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocTradeClass__c", 'text', 'TRADE CLASS', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocCategory__c", 'text', 'CATEGORY1, CATEGORY2, CATEGORY3', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocBrand__c", 'text', 'BRAND', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocLocType__c", 'text', 'TYPE', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocProgram__c", 'text', 'PROGRAM', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocGeography__c", 'text', '20', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOCProjectCode__c", 'text', str(projectCode), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])
                elif productLine in ('FSI'):
                    child1OppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine],
                                                                                         'Opportunity_Category__c': 'Child',
                                                                                         'Name': child1OpportunityName,
                                                                                         'StageName': 'Contract',
                                                                                         'NAPI_Insert_date__c': str(napiDateId),
                                                                                         'CloseDate': str(napiDate),
                                                                                         'End_Date__c': str(napiDate),
                                                                                         'AccountId': accountID,
                                                                                         'Parent_Opportunity__c': parentOppDetails['id'],
                                                                                         'Probability__c': '75',
                                                                                         'Estimated_Average_CPM__c': 1,
                                                                                         'Expected_Circulation__c': 1,
                                                                                         'Business_Type__c': 'New',
                                                                                         'Status__c': 'Reserved-RS1',
                                                                                         'Type': 'Standard',
                                                                                         'ILocCirculationCharges__c': '[{"ChargeType":"C","Charges":1000,"Description":"CIRCULATION CHARGE","Amount":1,"CirculationQty":1000}]',
                                                                                         'ILOCProductionCharges__c': '[{"ChargeType":"P","Charges":2000,"Description":"DISK HANDLING","Amount":1000,"CirculationQty":2}]',
                                                                                         'ILocOtherCharges__c': '[{"ChargeType":"o","Charges":5000,"Description":"OTHER CHARGE","Amount":50, "CirculationQty": 100}]',
                                                                                         'ILOCTotalProgramFee__c': 8000,
                                                                                         'Order__c': p1OrderNumber,
                                                                                         'Parent_Order__c': '',
                                                                                         'ILOCAdDescription__c': 'Description',
                                                                                         'ILocCategory__c': 'Category',
                                                                                         'ILocMarketListDueDate__c': str(napiDate),
                                                                                         'ILocArtworkDueDate__c': str(napiDate),
                                                                                         'Billed_based_on_Actual_Execution__c': True})
                    
                    if child1OppDetails["success"] == True: 
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Id", 'id', child1OppDetails['id'], child1OppDetails['id'])     
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CurrencyIsoCode", 'text', currencyCode, child1OppDetails['id'])    
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeId", 'id', str(opprtunityPLMap[productLine]), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeName", 'string', str(productLine), child1OppDetails['id'])                        
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Name", 'string', str(child1OpportunityName), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "PARENT_OPP_ID", 'id', parentOppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "StageName", 'text', 'Contract', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CloseDate", 'text', napiDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "End_Date__c", 'text', napiDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "NAPI_Insert_date__c", 'text', napiDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "AccountId", 'reference', str(accountID), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Probability__c", 'text', '75', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Estimated_Average_CPM__c", 'currency', 1, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Expected_Circulation__c", 'currency', 1, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Business_Type__c", 'text', 'New', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Status__c", 'text', 'Reserved-RS1', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Type", 'text', 'Standard', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocCirculationCharges__c", 'textarea', r'[{"ChargeType":"C","Charges":1000,"Description":"CIRCULATION CHARGE","Amount":1,"CirculationQty":1000}]', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOCProductionCharges__c", 'textarea', r'[{"ChargeType":"P","Charges":2000,"Description":"DISK HANDLING","Amount":1000,"CirculationQty":2}]', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocOtherCharges__c", 'textarea', r'[{"ChargeType":"o","Charges":5000,"Description":"OTHER CHARGE","Amount":50, "CirculationQty": 100}]', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOCTotalProgramFee__c", 'text', '8000', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Order__c", 'text', str(p1OrderNumber), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Parent_Order__c", 'text', '', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOCAdDescription__c", 'text', 'Description', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocCategory__c", 'text', 'Category', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocMarketListDueDate__c", 'text', napiDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocArtworkDueDate__c", 'text', napiDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])
                    oppDetails['C_NAPI_DATE'] = str(napiDateId)
                    oppDetails['C_STATUS'] = 'Reserved-RS1'
                    oppDetails['C_TOTAL_PROGRAM_FEE'] = 8000
                    oppDetails['C_ILOCAdDescription'] = 'Description'
                    oppDetails['C_ILocMarketListDueDate'] = str(
                        napiDateId)
                    oppDetails['C_ILocArtworkDueDate'] = str(napiDateId)
                    oppDetails['C_ILocCategory'] = 'CATEGORY1, CATEGORY2, CATEGORY3'
                    oppDetails['C_Billed_based_on_Actual_Execution'] = True
                elif productLine in ('SSMG'):
                    child1OppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine],
                                                                                         'Opportunity_Category__c': 'Child',
                                                                                         'Name': child1OpportunityName,
                                                                                         'StageName': 'Contract',
                                                                                         'Insert_Date__c': str(futureDate),
                                                                                         'CloseDate': str(futureDate),
                                                                                         'End_Date__c': str(futureDate),
                                                                                         'AccountId': accountID,
                                                                                         'Parent_Opportunity__c': parentOppDetails['id'],
                                                                                         'Probability__c': '75',
                                                                                         'Estimated_Average_CPM__c': 1,
                                                                                         'Expected_Circulation__c': 1,
                                                                                         'Business_Type__c': 'New',
                                                                                         'Type': 'Standard',
                                                                                         'Order__c': p1OrderNumber,
                                                                                         'Parent_Order__c': '',
                                                                                         'ILocMarketListDueDate__c': str(futureDate),
                                                                                         'ILoc_Material_Due_Date__c': str(futureDate),
                                                                                         'Billed_based_on_Actual_Execution__c': True})
                    if child1OppDetails["success"] == True:  
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Id", 'id', child1OppDetails['id'], child1OppDetails['id'])     
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CurrencyIsoCode", 'text', currencyCode, child1OppDetails['id'])    
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeId", 'id', str(opprtunityPLMap[productLine]), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeName", 'string', str(productLine), child1OppDetails['id'])                        
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Name", 'string', str(child1OpportunityName), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "PARENT_OPP_ID", 'id', parentOppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "StageName", 'text', 'Contract', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Insert_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CloseDate", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "End_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "AccountId", 'reference', str(accountID), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Probability__c", 'text', '75', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Estimated_Average_CPM__c", 'currency', 1, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Expected_Circulation__c", 'currency', 1, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Business_Type__c", 'text', 'New', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Status__c", 'text', 'Reserved-RS1', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Type", 'text', 'Standard', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Order__c", 'text', str(p1OrderNumber), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Parent_Order__c", 'text', '', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocMarketListDueDate__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Material_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])

                    oppDetails['C_ILocMarketListDueDate'] = str(
                        futureDate)
                    oppDetails['C_ILoc_Material_Due_Date'] = str(
                        futureDate)
                    oppDetails['C_Billed_based_on_Actual_Execution'] = True
                elif productLine in ('Checkout 51'):
                    child1OppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine],
                                                                                         'Opportunity_Category__c': 'Child',
                                                                                         'Name': child1OpportunityName,
                                                                                         'StageName': 'Contract',
                                                                                         'Insert_Date__c': str(futureDate),
                                                                                         'CloseDate': str(futureDate),
                                                                                         'End_Date__c': str(futureDate),
                                                                                         'AccountId': accountID,
                                                                                         'Parent_Opportunity__c': parentOppDetails['id'],
                                                                                         'Probability__c': '75',
                                                                                         'Business_Type__c': 'New',
                                                                                         'Order__c': p1OrderNumber,
                                                                                         'Parent_Order__c': '',
                                                                                         'ILoc_Assest_Due_Date__c': str(futureDate),
                                                                                         'Billed_based_on_Actual_Execution__c': True})
                    if child1OppDetails["success"] == True:   
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Id", 'id', child1OppDetails['id'], child1OppDetails['id'])     
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CurrencyIsoCode", 'text', currencyCode, child1OppDetails['id'])    
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeId", 'id', str(opprtunityPLMap[productLine]), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeName", 'string', str(productLine), child1OppDetails['id'])                        
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Name", 'string', str(child1OpportunityName), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "PARENT_OPP_ID", 'id', parentOppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "StageName", 'text', 'Contract', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Insert_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CloseDate", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "End_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "AccountId", 'reference', str(accountID), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Probability__c", 'text', '75', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Business_Type__c", 'text', 'New', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Order__c", 'text', str(p1OrderNumber), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Parent_Order__c", 'text', '', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Assest_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])
                    
                    oppDetails['C_ILoc_Assest_Due_Date'] = str(futureDate)
                    oppDetails['C_Billed_based_on_Actual_Execution'] = True
                elif productLine in ('Merchandising'):
                    projectCode = 600093
                    child1OppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine],
                                                                                         'Opportunity_Category__c': 'Child',
                                                                                         'Name': child1OpportunityName,
                                                                                         'StageName': 'Contract',
                                                                                         'Insert_Date__c': str(futureDate),
                                                                                         'CloseDate': str(futureDate),
                                                                                         'End_Date__c': str(futureDate),
                                                                                         'AccountId': accountID,
                                                                                         'Parent_Opportunity__c': parentOppDetails['id'],
                                                                                         'Probability__c': '75',
                                                                                         'Business_Type__c': 'New',
                                                                                         'Status__c': 'Reserved-RS1',
                                                                                         'Type': 'Subscription',
                                                                                         'Order__c': p1OrderNumber,
                                                                                         'Parent_Order__c': '',
                                                                                         'ILOCProjectCode__c': projectCode,
                                                                                         'ILOC_Store_Type__c': 'Store Type',
                                                                                         'ILOC_Tentative_geography__c': True,
                                                                                         'ILocStoreCount__c': 10,
                                                                                         'ILocTradeClass__c': 'Trade Class',
                                                                                         'ILocRemoveWave__c': True,
                                                                                         'ILocCategory__c': 'AIRTEL, IDEA, JIO, VODAFONE, TATA DOCOMO',
                                                                                         'Job_Description_Comments__c': 'INSTALLATION, MAINTENANACE, OFFERS, FITTING',
                                                                                         'Billed_based_on_Actual_Execution__c': True,
                                                                                         'ILocSignedLOCORStoreListDueDt__c': str(futureDate),
                                                                                         'POS_Materials_Produced_If_yes_above__c': 'NMCIS',
                                                                                         'ILocProgramDocumentDueDate__c': str(futureDate),
                                                                                         'POS_Material_Disposition__c': r"Return to Client at Client's expense",
                                                                                         'ILocArtworkDueDate__c': str(futureDate),
                                                                                         'Network_Retailer_List__c': 'AIRTEL, IDEA, JIO, VODAFONE, TATA DOCOMO',
                                                                                         'ILoc_Material_Due_Date__c': str(futureDate),
                                                                                         'Non_Network_Retailer_List__c': 'AIRTEL, IDEA, JIO, VODAFONE, TATA DOCOMO',
                                                                                         'Client_Supplied_Materials_Due_To__c': 'Other Vendor',
                                                                                         'Image_Request_Details__c': 'Subset List',
                                                                                         'Subset_List_Count__c': 46})
                    if child1OppDetails["success"] == True:   
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Id", 'id', child1OppDetails['id'], child1OppDetails['id'])     
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CurrencyIsoCode", 'text', currencyCode, child1OppDetails['id'])    
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeId", 'id', str(opprtunityPLMap[productLine]), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeName", 'string', str(productLine), child1OppDetails['id'])                        
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Name", 'string', str(child1OpportunityName), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "PARENT_OPP_ID", 'id', parentOppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "StageName", 'text', 'Contract', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Insert_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CloseDate", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "End_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "AccountId", 'reference', str(accountID), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Probability__c", 'text', '75', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Business_Type__c", 'text', 'New', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Status__c", 'text', 'Reserved-RS1', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Type", 'text', 'Subscription', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Order__c", 'text', str(p1OrderNumber), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Parent_Order__c", 'text', '', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Assest_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOC_Store_Type__c", 'text', 'Store Type', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOC_Tentative_geography__c", 'text', 'True', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocStoreCount__c", 'text', '10', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocTradeClass__c", 'text', 'Trade Class', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocRemoveWave__c", 'text', 'True', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocCategory__c", 'string', 'AIRTEL, IDEA, JIO, VODAFONE, TATA DOCOMO', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Job_Description_Comments__c", 'string', 'INSTALLATION, MAINTENANACE, OFFERS, FITTING', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocSignedLOCORStoreListDueDt__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "POS_Materials_Produced_If_yes_above__c", 'text', 'NMCIS', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocProgramDocumentDueDate__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "POS_Material_Disposition__c", 'string', r"Return to Client at Client expense", child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocArtworkDueDate__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Network_Retailer_List__c", 'string', 'AIRTEL, IDEA, JIO, VODAFONE, TATA DOCOMO', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Material_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Non_Network_Retailer_List__c", 'string', 'AIRTEL, IDEA, JIO, VODAFONE, TATA DOCOMO', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Client_Supplied_Materials_Due_To__c", 'text', 'Other Vendor', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Image_Request_Details__c", 'text', 'Image_Request_Details__c', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Subset_List_Count__c", 'text', '46', child1OppDetails['id'])
                    oppDetails['C_STATUS'] = 'Reserved-RS1'
                    oppDetails['C_ILOCProjectCode'] = projectCode
                    oppDetails['C_ILOC_Store_Type'] = 'Store Type'
                    oppDetails['C_ILOC_Tentative_geography'] = True
                    oppDetails['C_ILocStoreCount'] = 10
                    oppDetails['C_ILocTradeClass'] = 'Trade Class'
                    oppDetails['C_ILocRemoveWave'] = True
                    oppDetails['C_ILocCategory'] = 'AIRTEL IDEA JIO VODAFONE TATA DOCOMO'
                    oppDetails['C_Job_Description_Comments'] = 'INSTALLATION MAINTENANACE OFFERS FITTING'
                    oppDetails['C_Billed_based_on_Actual_Execution'] = True
                    oppDetails['C_ILocSignedLOCORStoreListDueDt'] = str(
                        futureDate)
                    oppDetails['C_POS_Materials_Produced_If_yes_above'] = 'NMCIS'
                    oppDetails['C_ILocProgramDocumentDueDate'] = str(
                        futureDate)
                    oppDetails['C_POS_Material_Disposition'] = r"Return to Client at Client's expense"
                    oppDetails['C_ILocArtworkDueDate'] = str(futureDate)
                    oppDetails['C_Network_Retailer_List'] = 'AIRTEL IDEA JIO VODAFONE TATA DOCOMO'
                    oppDetails['C_ILoc_Material_Due_Date'] = str(
                        futureDate)
                    oppDetails['C_Non_Network_Retailer_List'] = 'AIRTEL IDEA JIO VODAFONE TATA DOCOMO'
                    oppDetails['C_Client_Supplied_Materials_Due_To'] = 'Other Vendor'
                    oppDetails['C_Image_Request_Details'] = 'Subset List'
                    oppDetails['C_Subset_List_Count'] = 46
                elif productLine in ('SmartSource Direct', 'SmartSource Direct- Canada'):
                    if productLine in ('SmartSource Direct'):
                        projectCode = 549999
                    elif productLine in ('SmartSource Direct- Canada'):
                        projectCode = 599999
                    child1OppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine],
                                                                                         'Opportunity_Category__c': 'Child',
                                                                                         'Name': child1OpportunityName,
                                                                                         'StageName': 'Contract',
                                                                                         'Insert_Date__c': str(futureDate),
                                                                                         'CloseDate': str(futureDate),
                                                                                         'End_Date__c': str(futureDate),
                                                                                         'AccountId': accountID,
                                                                                         'Parent_Opportunity__c': parentOppDetails['id'],
                                                                                         'Probability__c': '75',
                                                                                         'Business_Type__c': 'New',
                                                                                         'Status__c': 'Reserved-RS1',
                                                                                         'Type': 'Standard',
                                                                                         'Order__c': p1OrderNumber,
                                                                                         'Parent_Order__c': '',
                                                                                         'ILOCProjectCode__c': projectCode,
                                                                                         'ILoc_Format__c': '4BF0D9BD 35F7 4FE6 B',
                                                                                         'ILoc_Targeting__c': 'C4C9F437 0703 4',
                                                                                         'ILocArtworkDueDate__c': str(futureDate),
                                                                                         'ILoc_Sample_Due_Date__c': str(futureDate),
                                                                                         'ILoc_Creative_Due_Date__c': str(futureDate),
                                                                                         'Billed_based_on_Actual_Execution__c': True,
                                                                                         'ILocCategory__c': '70888577, 4615, 4BDF, 8D4C, 350F21B2192D'})
                    if child1OppDetails["success"] == True: 
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Id", 'id', child1OppDetails['id'], child1OppDetails['id'])     
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CurrencyIsoCode", 'text', currencyCode, child1OppDetails['id'])    
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeId", 'id', str(opprtunityPLMap[productLine]), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeName", 'string', str(productLine), child1OppDetails['id'])                        
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Name", 'string', str(child1OpportunityName), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "PARENT_OPP_ID", 'id', parentOppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "StageName", 'text', 'Contract', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CloseDate", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "End_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Insert_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "AccountId", 'reference', str(accountID), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Probability__c", 'text', '75', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Business_Type__c", 'text', 'New', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Status__c", 'text', 'Reserved-RS1', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Type", 'text', 'Standard', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Order__c", 'text', str(p1OrderNumber), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Parent_Order__c", 'text', '', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Format__c", 'text',  '4BF0D9BD 35F7 4FE6 B', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOCProjectCode__c", 'text', str(projectCode), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Targeting__c", 'text', 'C4C9F437 0703 4', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocArtworkDueDate__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Sample_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Creative_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])                    
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocCategory__c", 'string', '70888577, 4615, 4BDF, 8D4C, 350F21B2192D', child1OppDetails['id'])

                    oppDetails['C_STATUS'] = 'Reserved-RS1'
                    oppDetails['C_ILOCProjectCode'] = projectCode
                    oppDetails['C_ILoc_Format'] = '4BF0D9BD 35F7 4FE6 B'
                    oppDetails['C_ILoc_Targeting'] = 'C4C9F437 0703 4'
                    oppDetails['C_ILocArtworkDueDate'] = str(
                        futureDate)
                    oppDetails['C_ILoc_Sample_Due_Date'] = str(
                        futureDate)
                    oppDetails['C_ILoc_Creative_Due_Date'] = str(
                        futureDate)
                    oppDetails['C_Billed_based_on_Actual_Execution'] = True
                    oppDetails['C_ILocCategory'] = '70888577 4615 4BDF 8D4C 350F21B2192D'
                elif productLine in ('Digital', 'Digital- Canada'):
                    child1OppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine],
                                                                                         'Opportunity_Category__c': 'Child',
                                                                                         'Name': child1OpportunityName,
                                                                                         'StageName': 'Contract',
                                                                                         'Insert_Date__c': str(futureDate),
                                                                                         'CloseDate': str(futureDate),
                                                                                         'End_Date__c': str(futureDate),
                                                                                         'AccountId': accountID,
                                                                                         'Parent_Opportunity__c': parentOppDetails['id'],
                                                                                         'Probability__c': '75',
                                                                                         'Business_Type__c': 'New',
                                                                                         'Status__c': 'Reserved-RS1',
                                                                                         'Type': 'Standard',
                                                                                         'Order__c': p1OrderNumber,
                                                                                         'Parent_Order__c': '',
                                                                                         'Artwork_Due_Date__c': str(futureDate),
                                                                                         'ClientListDueDate__c': str(futureDate),
                                                                                         'Billed_based_on_Actual_Execution__c': True,
                                                                                         'ILocCategory__c': 'AIRTEL IDEA JIO VODAFONE TATA DOCOMO',
                                                                                         'ILOC_Client_s_clip_rate__c': 10,
                                                                                         'ILOCProjectCode__c': 7755001})

                    if child1OppDetails["success"] == True: 
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Id", 'id', child1OppDetails['id'], child1OppDetails['id'])     
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CurrencyIsoCode", 'text', currencyCode, child1OppDetails['id'])    
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeId", 'id', str(opprtunityPLMap[productLine]), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeName", 'string', str(productLine), child1OppDetails['id'])                        
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Name", 'string', str(child1OpportunityName), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "PARENT_OPP_ID", 'id', parentOppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "StageName", 'text', 'Contract', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CloseDate", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "End_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Insert_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "AccountId", 'reference', str(accountID), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Probability__c", 'text', '75', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Business_Type__c", 'text', 'New', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Status__c", 'text', 'Reserved-RS1', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Type", 'text', 'Standard', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Order__c", 'text', str(p1OrderNumber), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Parent_Order__c", 'text', '', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOC_Client_s_clip_rate__c", 'text', '10', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOCProjectCode__c", 'text', '7755001', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Artwork_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ClientListDueDate__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])                    
                    oppDetails['C_STATUS'] = 'Reserved-RS1'
                    oppDetails['C_Artwork_Due_Date'] = str(futureDate)
                    oppDetails['C_ClientListDueDate'] = str(futureDate)
                    oppDetails['C_Billed_based_on_Actual_Execution'] = True
                    oppDetails['C_ILocCategory'] = 'AIRTEL IDEA JIO VODAFONE TATA DOCOMO'
                    oppDetails['C_ILOC_Client_s_clip_rate'] = 10
                    oppDetails['C_ILOCProjectCode'] = 7755001
                print(child1OppDetails)
                if child1OppDetails["success"] == True:
                    print(f"Child1 Opp ID of {productLine}: ",
                          child1OppDetails['id'])
                    oppDetails['C_ID'] = child1OppDetails['id']
                    oppDetails['C_NAME'] = child1OpportunityName
                    oppDetails['C_ACC_ID'] = accountID
                    oppDetails['C_STAGE'] = 'Contract'

                    usProductLineMap[productLine]['PARENT_OPP_1']['CHILD1_OPP_ID']['ID'] = child1OppDetails['id']
                    usProductLineMap[productLine]['PARENT_OPP_1']['CHILD1_OPP_ID']['ORDER#'] = str(
                        p1OrderNumber)
                    childOpps.append(child1OppDetails['id'])
                    Messages.write_message(
                        f"Child 1 Opp ID of {productLine}: {child1OppDetails['id']}")
                    data_store.spec[f"P1_C1_ID_{productLine}"] = child1OppDetails['id']
                    data_store.spec[f"P1_C1_NAME_{productLine}"] = child1OpportunityName

            print("\nChild Opportunities\n", childOpps)
    # -------------------------------------------------------------------------------------------------------------------------------------
            productsList = []
            productsList.clear()
            productName = None
            if productLine in ('InStore', 'InStore-Canada'):
                oppPLProductsSOQL = f"SELECT Id,Product_Line__c, Product_Line__r.Name,Product__c, Product__r.Name FROM Product_Junction__c where Product_Line__r.Name = 'InStore' and Product__r.IsActive = true and Product__c in (SELECT Product2Id from PricebookEntry where CurrencyIsoCode = '{currencyCode}' and IsActive = true)"
            else:
                oppPLProductsSOQL = f"SELECT Id,Product_Line__c, Product_Line__r.Name,Product__c, Product__r.Name FROM Product_Junction__c where Product_Line__r.Name = '{productLine}' and Product__r.IsActive = true and Product__c in (SELECT Product2Id from PricebookEntry where CurrencyIsoCode = '{currencyCode}' and IsActive = true)"
        #         oppPLProductsSOQL = f"SELECT Id,Product_Line__c, Product_Line__r.Name,Product__c, Product__r.Name FROM Product_Junction__c where Product_Line__r.Name = '{productLine}'"
            oppPLProductsResult = AccountOpportunityCreation.sf.query_all(
                query=oppPLProductsSOQL)
            oppPLProductsData = oppPLProductsResult['records']

            for oppPLProduct in oppPLProductsData:
                if oppPLProduct['Product__r'] != None:
                    if 'Name' in oppPLProduct['Product__r']:
                        productsList.append(oppPLProduct['Product__r']['Name'])
                        print("Found Product: ",
                              oppPLProduct['Product__r']['Name'])
            print(*productsList, sep="\n")
    # -------------------------------------------------------------------------------------------------------------------------------------
            lineItemsMap = {}
            # pdb.set_trace()
            productNameTemp = ''
            priceBookId = ''
            for idx, childOpp in enumerate(childOpps):
                IsActive = False
                while IsActive == False:
                    productName = random.choice(productsList)
                    productNameTemp = productName.replace("'", r"\'")

                    Messages.write_message(f"Product Name: {productName}")
                    data_store.spec[f"PRODUCT_NAME"] = productName
                    oppDetails['C_PRODUCT_NAME'] = productName

                    print("Product Name: ", productName)
                    # priceBookSOQL = f"SELECT Id, Name,IsActive FROM PricebookEntry where Name = '{productName}' and CurrencyIsoCode = '{currencyCode}' and IsActive = true"
                    priceBookSOQL = f"SELECT Id, IsActive FROM PricebookEntry WHERE CurrencyIsoCode = '{currencyCode}' AND IsActive = True AND Product2.Name = '{productNameTemp}' AND Product2Id in (SELECT Product__c FROM Product_Junction__c WHERE Product_Line__r.Name = '{productLine}')"
                    priceBookResult = AccountOpportunityCreation.sf.query_all(
                        query=priceBookSOQL)
                    priceBookData = priceBookResult['records']
                    if priceBookData[0]['IsActive'] != False:
                        priceBookId = priceBookData[0]['Id']
                        Messages.write_message(
                            f"Pricebook Entry ID: {priceBookId}")
                        print("Pricebook Entry ID: ", priceBookId)
                        data_store.spec["PRICE_BOOK_ID"] = priceBookId

                        if productLine in ('InStore', 'InStore-Canada'):
                            chargeDetailsSOQL = f"SELECT Id, Name, ProductLIne__c, Product__C, Charge_Type_2__c, Charge_Type_2__r.Commissionable__c, Charge_Type_2__r.Charge_Type_Categorty__c FROM Pricing_Detail__c where ProductLIne__c = 'Instore' and product__c = '{productNameTemp}' and CurrencyIsoCode = '{currencyCode}' and isActive__c = true and IsCurrent__c = true and RateCardIsCurrent__c = true"
                        elif productLine in ('Checkout 51'):
                            chargeDetailsSOQL = f"SELECT Id, Name, ProductLIne__c, Product__C, Charge_Type_2__c, Charge_Type_2__r.Commissionable__c, Charge_Type_2__r.Charge_Type_Categorty__c FROM Pricing_Detail__c where ProductLIne__c = '{productLine}' and FreedomOrNot__c != 'Freedom' and product__c = '{productNameTemp}' and CurrencyIsoCode = '{currencyCode}' and isActive__c = true and IsCurrent__c = true and RateCardIsCurrent__c = true"
                        elif productLine in ('Digital', 'Digital- Canada'):
                            chargeDetailsSOQL = f"SELECT Id, Name, ProductLIne__c, Product__C, Charge_Type_2__c, Charge_Type_2__r.Commissionable__c, Charge_Type_2__r.Charge_Type_Categorty__c FROM Pricing_Detail__c where ProductLIne__c = '{productLine}' and product__c = '{productNameTemp}' and CurrencyIsoCode = '{currencyCode}' and isActive__c = true and FreedomOrNot__c not in ('Freedom') and IsCurrent__c = true  and RateCardIsCurrent__c = true"
                        else:
                            chargeDetailsSOQL = f"SELECT Id, Name, ProductLIne__c, Product__C, Charge_Type_2__c, Charge_Type_2__r.Commissionable__c, Charge_Type_2__r.Charge_Type_Categorty__c FROM Pricing_Detail__c where ProductLIne__c = '{productLine}' and product__c = '{productNameTemp}' and CurrencyIsoCode = '{currencyCode}' and isActive__c = true and IsCurrent__c = true  and RateCardIsCurrent__c = true"
                        chargeDetailsResult = AccountOpportunityCreation.sf.query_all(
                            query=chargeDetailsSOQL)
                        chargeDetailsData = chargeDetailsResult['records']
                        print("\n", chargeDetailsSOQL, "\n")
                        if len(chargeDetailsData) > 2:
                            chargeTypeMap = {}
                            chargeTypeMap.clear()
                            for chargeDetails in chargeDetailsData:
                                chargeTypeMap[chargeDetails['Name']] = {"Id": chargeDetails['Id'], "ChargeType": chargeDetails[
                                    'Charge_Type_2__c'], 'Commissionable': chargeDetails['Charge_Type_2__r']['Commissionable__c'], 'Charge_Type_Category': chargeDetails['Charge_Type_2__r']['Charge_Type_Categorty__c']}
                            # print("\nCharge Types\n", chargeTypeMap)
                            [print(key, value)
                                for key, value in chargeTypeMap.items()]
                            Messages.write_message(
                                f"Charge Types MAP \n{chargeTypeMap}")
                            IsActive = True

                totalAmount = 0

                for cnt in range(1, 5):
                    chargeType = random.choice(list(chargeTypeMap))
                    chargeTypeTemp = chargeType.replace("'", r"\'")
                    soql = f"SELECT Charge_Type_2__r.Commissionable__c,Charge_Type__c, Charge_Type_Category__c, CurrencyIsoCode, Id, isActive__c, Name, Product__c, Quantity_1000_effective_from_date__c, Quantity_1000__c FROM Pricing_Detail__c WHERE Name = '{chargeTypeTemp}' and Product__c= '{productNameTemp}' and CurrencyIsoCode = '{currencyCode}' and isActive__c = True and RateCardIsCurrent__c = true and IsCurrent__c = true"
                    queryResult = AccountOpportunityCreation.sf.query_all(
                        query=soql)
                    recDetails = queryResult['records']
                    if "CLIENT_ORDER_LIMIT_GT_ILOC_TOTAL" in oppName:
                        quantity = randint(10, 10)
                        salesPrice = round(random.uniform(10, 10), 2)
                    elif "CLIENT_ORDER_LIMIT_LT_ILOC_TOTAL" in oppName:
                        quantity = randint(5000, 5000)
                        salesPrice = round(random.uniform(50, 50), 2)
                    else:
                        spQty
                        quantity = int(spQty.split(":")[1].split("=")[1])
                        salesPrice = int(spQty.split(":")[0].split("=")[1])

                    if recDetails[0]['Quantity_1000__c'] == True:
                        totalPrice = round((quantity / 1000) * salesPrice, 2)
                    else:
                        totalPrice = round((quantity * salesPrice), 2)
                    totalAmount = totalAmount + totalPrice

                    # pdb.set_trace()
                    lineItemsDetails = AccountOpportunityCreation.sf.OpportunityLineItem.create({'Charge_Type__c': chargeTypeMap[chargeType]["ChargeType"], 'Commissionable__c': chargeTypeMap[chargeType]["Commissionable"], 'pricebookentryid': priceBookId,
                                                                                                 'Pricing_Detail__c': chargeType, 'PricingDetail__c': chargeTypeMap[chargeType]["Id"], 'opportunityId': childOpp, 'Quantity': quantity, 'Sales_price__c': salesPrice, 'TotalPrice': totalPrice, 'Component__c': 'Video'})

                    # oppDetails[f'C_{cnt}_CHARGE_TYPE'] = chargeTypeMap[chargeType]["ChargeType"]
                    # oppDetails[f'C_{cnt}_CHARGE_TYPE_CATEGORY'] = chargeTypeMap[chargeType]["Charge_Type_Category"]
                    # oppDetails[f'C_{cnt}_COMMISSIONABLE'] = chargeTypeMap[chargeType]["Commissionable"]
                    # oppDetails[f'C_{cnt}_PRICING_DETAIL_NAME'] = chargeType
                    # oppDetails[f'C_{cnt}_PRICING_DETAIL_ID'] = lineItemsDetails["id"]
                    # oppDetails[f'C_{cnt}_QTY'] = quantity
                    # oppDetails[f'C_{cnt}_SALES_PRICE'] = salesPrice
                    # oppDetails[f'C_{cnt}_SUB_TOTAL'] = totalPrice

                    if lineItemsDetails["success"] == True:
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Id", 'id', lineItemsDetails["id"], lineItemsDetails["id"])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Opportunity__c", 'reference', child1OppDetails['id'], lineItemsDetails["id"])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "product__c", 'string', productNameTemp, lineItemsDetails["id"])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "ProductLIne__c", 'string', productLine, lineItemsDetails["id"])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "CurrencyIsoCode", 'text', currencyCode, lineItemsDetails["id"])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "priceBookId", 'string', priceBookId, lineItemsDetails["id"])
                        oppDetails[f'C_{cnt}_CHARGE_TYPE'] = chargeTypeMap[chargeType]["ChargeType"]
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Charge_Type__c", 'string', chargeTypeMap[chargeType]["ChargeType"], lineItemsDetails["id"])
                        oppDetails[f'C_{cnt}_CHARGE_TYPE_CATEGORY'] = chargeTypeMap[chargeType]["Charge_Type_Category"]
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Charge_Type_Category__c", 'string', chargeTypeMap[chargeType]["Charge_Type_Category"], lineItemsDetails["id"]) 
                        oppDetails[f'C_{cnt}_COMMISSIONABLE'] = chargeTypeMap[chargeType]["Commissionable"]
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Commissionable__c", 'text', chargeTypeMap[chargeType]["Commissionable"], lineItemsDetails["id"])
                        oppDetails[f'C_{cnt}_PRICING_DETAIL_NAME'] = chargeType
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Pricing_Detail__c", 'string', chargeType, lineItemsDetails["id"])
                        oppDetails[f'C_{cnt}_PRICING_DETAIL_ID'] = lineItemsDetails["id"]
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "PricingDetail__c", 'reference', chargeTypeMap[chargeType]["Id"], lineItemsDetails["id"])
                        oppDetails[f'C_{cnt}_QTY'] = quantity
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Quantity", 'double', quantity, lineItemsDetails["id"])
                        oppDetails[f'C_{cnt}_SALES_PRICE'] = salesPrice
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Sales_price__c", 'double', salesPrice, lineItemsDetails["id"])
                        oppDetails[f'C_{cnt}_SUB_TOTAL'] = totalPrice
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "TotalPrice", 'double', totalPrice, lineItemsDetails["id"])
                        if recDetails[0]['Quantity_1000__c'] == True:
                            Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Quantity_1000__c", 'text', recDetails[0]['Quantity_1000__c'], lineItemsDetails["id"])
                            Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Quantity_1000_effective_from_date__c", 'text', recDetails[0]['Quantity_1000_effective_from_date__c'], lineItemsDetails["id"])
                        print(
                            f"OPP[{childOpp}] Line Item {cnt}: {lineItemsDetails['id']} created..")
                        Messages.write_message(
                            f"OPP[{childOpp}] Line Item {cnt}: {lineItemsDetails['id']} created..")
                        lineItemsMap[childOpp +
                                     str(cnt)] = {cnt: lineItemsDetails['id']}
                oppDetails['C_TOTAL_AMOUNT'] = totalAmount
                # Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Total", 'currency', totalAmount, child1OppDetails['id'])
                if productLine in ('FSI') and productName != 'Remnant':
                    Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Total", 'currency', '8000', child1OppDetails['id'])
                else:
                    Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Total", 'currency', totalAmount, child1OppDetails['id'])                

            columnCount = 0
            rowCount = 0

            df = pd.read_excel(opportunitiesDetailsFileName,
                               sheet_name=f"{productLine}_{currencyCode}")
            Messages.write_message(df.columns)
            columnList = df.columns.tolist()
            [print(key, value) for key, value in oppDetails.items()]
            rowCount = int(ws.max_row) + 1
            for key, value in oppDetails.items():
                #         pdb.set_trace()
                if key in columnList:
                    ws[f"{xlsxwriter.utility.xl_col_to_name(columnList.index(key))}{rowCount}"] = value
                else:
                    if len(columnList) == 0 and ws["A1"].value == None:
                        ws[f"{xlsxwriter.utility.xl_col_to_name(columnCount)}{1}"] = key
                        ws[f"{xlsxwriter.utility.xl_col_to_name(columnCount)}{2}"] = value
                    else:
                        columnCount = int(ws.max_column)
                        ws[f"{xlsxwriter.utility.xl_col_to_name(columnCount)}{1}"] = key
                        ws[f"{xlsxwriter.utility.xl_col_to_name(columnCount)}{rowCount}"] = value
                wb.save(filename=opportunitiesDetailsFileName)
            wb.save(filename=opportunitiesDetailsFileName)

    @step("Create <oppName> Opportunity for <productLineRegion> productline")
    def create_opportunity_for_productline(self, oppName, productLineRegion):

        today = date.today()
        oppDateToday = today.strftime("%Y%m%d")
        opportunityName = f"{oppName}{oppDateToday}"

        queryData = (AccountOpportunityCreation.sf.query(format_soql(
            "SELECT Id, Name FROM Opportunity WHERE Name = {} AND Opportunity_Category__c = 'Child'", opportunityName)))['records']

        if not len(queryData) > 0:
            region = productLineRegion.split("_")[1]
            productLine = productLineRegion.split("_")[0]
            currencyCode = region
            opportunitiesDetailsPath = Path(__file__).parents[1]
            opportunitiesDetailsFileName = str(
                opportunitiesDetailsPath) + "\\Data\\" + os.getenv("ILOC_OPPORTUNITY_DETAILS_FILE")
            wb = None
            ws = None

            if os.path.exists(opportunitiesDetailsFileName):
                wb = load_workbook(filename=opportunitiesDetailsFileName)
            else:
                wb = Workbook()
                wb.save(opportunitiesDetailsFileName)

            end = datetime.datetime.now()
            current_month = end.strftime('%m')
            current_day = end.strftime('%d')
            current_year_full = end.strftime('%Y')

            futureDate = datetime.datetime(int(current_year_full) + 1,
                                           int(current_month), int(current_day)).date()
            print("Future Date: ", futureDate)
            userData = AccountOpportunityCreation.sf.quick_search(
                f"{os.getenv('USER_ID')}")
            userId = userData['searchRecords'][0]['Id']
            print(userData['searchRecords'][0]['Id'])

            napiDateId, napiDate = AccountOpportunityCreation.getNAPIDate(
                self, AccountOpportunityCreation.sf, "Future")

            instoreCycle, instoreCycleName = AccountOpportunityCreation.getInstoreCycle(
                self, AccountOpportunityCreation.sf, f"{str(int(current_year_full)+1)}{current_month}", current_month, current_year_full)

            usProductLineMap = {}
            productLineList = None

            accountID = os.getenv(
                f'OPPORUNITIES_TO_CREATE_FOR_{region}_ACCOUNT_ID')

            opprtunityPLMap = {}
            opprtunityPLSOQL = f"select Id,Name from RecordType where sObjectType='Opportunity'"
            opprtunityPLResult = AccountOpportunityCreation.sf.query_all(
                query=opprtunityPLSOQL)
            opprtunityPLData = opprtunityPLResult['records']

            for opprtunityPL in opprtunityPLData:
                opprtunityPLMap[opprtunityPL['Name']] = opprtunityPL['Id']

            [print(key, value) for key, value in opprtunityPLMap.items()]

            childOpps = []
            oppDetails = {}
            oppDetails.clear()
            if f"{productLine}_{currencyCode}" in wb.sheetnames:
                ws = wb[f"{productLine}_{currencyCode}"]
            else:
                ws = wb.create_sheet(f"{productLine}_{currencyCode}", 1)

            wb.save(filename=opportunitiesDetailsFileName)

            df = pd.read_excel(opportunitiesDetailsFileName,
                               sheet_name=f"{productLine}_{currencyCode}")
            Messages.write_message(df.columns)
            columnList = df.columns.tolist()

            rowCount = int(df.shape[0])
            colCount = int(df.shape[1])
            # pdb.set_trace()
            print(
                f"\n{productLine} Started---------------------------------------------------------------------------\n")
            usProductLineMap[productLine] = {
                'PRODUCT_LINE_ID': opprtunityPLMap[productLine], 'PRODUCT_LINE_NAME': productLine, 'PARENT_OPP': ''}

            parentOpportunityName = f"{productLine}#P#{rowCount + 1}"
            if productLine in ('InStore', 'InStore-Canada'):
                parentOppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine], 'Opportunity_Category__c': 'Parent', 'Name': parentOpportunityName,
                                                                                     'StageName': 'Contract', 'InStore_Cycle__c': str(instoreCycle), 'CloseDate': str(futureDate), 'AccountId': accountID, 'ILocArtworkDueDate__c': str(futureDate)})
            elif productLine in ('FSI'):
                parentOppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine], 'Opportunity_Category__c': 'Parent', 'Name': parentOpportunityName,
                                                                                     'StageName': 'Contract', 'NAPI_Insert_date__c': str(napiDateId), 'CloseDate': str(napiDate), 'AccountId': accountID})
        #     elif productLine in ('Digital','Digital- Canada'):
        #         parentOppDetails = sf.Opportunity.create({'RecordTypeId' : opprtunityPLMap[productLine],'Opportunity_Category__c' : 'Parent', 'Name' : parentOpportunityName,'StageName' : 'Contract','Insert_Date__c' : str(futureDate),'CloseDate' : str(futureDate), 'End_Date__c' : str(futureDate), 'AccountId': accountID})
            else:
                parentOppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine], 'Opportunity_Category__c': 'Parent', 'Name': parentOpportunityName,
                                                                                     'StageName': 'Contract', 'Insert_Date__c': str(futureDate), 'CloseDate': str(futureDate), 'End_Date__c': str(futureDate), 'AccountId': accountID})

            print(parentOppDetails)
            if parentOppDetails["success"] == True:
                print(f"Parent Opp ID of {productLine}: ",
                      parentOppDetails['id'])
                Messages.write_message(
                    f"Parent Opp ID of {productLine}: {parentOppDetails['id']}")

                data_store.spec[f"1_PO_ID_{productLine}"] = parentOppDetails['id']
                data_store.spec[f"1_PO_NAME_{productLine}"] = parentOpportunityName
                oppDetails['P_ID'] = parentOppDetails['id']
                oppDetails['P_NAME'] = parentOpportunityName

                usProductLineMap[productLine]['PARENT_OPP_1'] = {
                    'ID': parentOppDetails['id'], 'CHILD1_OPP_ID': '', 'CHILD2_OPP_ID': ''}
                usProductLineMap[productLine]['PARENT_OPP_1']['CHILD1_OPP_ID'] = {
                    'ID': '', 'ORDER#': '', 'PARENT_ORDER#': ''}
                usProductLineMap[productLine]['PARENT_OPP_1']['CHILD2_OPP_ID'] = {
                    'ID': '', 'ORDER#': '', 'PARENT_ORDER#': ''}

                # Write data to database                


            for i in range(1, 2):
                df = pd.read_excel(opportunitiesDetailsFileName,
                                   sheet_name=f"{productLine}_{currencyCode}")
                rowCount = int(df.shape[0])
                colCount = int(df.shape[1])
                today = date.today()
                oppDateToday = today.strftime("%Y%m%d")
                child1OpportunityName = f"{oppName}{oppDateToday}"
                p1OrderNumber = str(uuid.uuid4()).upper()[:10]

                if productLine in ('InStore', 'InStore-Canada'):
                    if productLine in ('InStore'):
                        projectCode = 300000
                    elif productLine in ('InStore-Canada'):
                        projectCode = 399999

                    child1OppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine],
                                                                                         'Opportunity_Category__c': 'Child',
                                                                                         'Name': child1OpportunityName,
                                                                                         'StageName': 'Contract',
                                                                                         'InStore_Cycle__c': str(instoreCycle),
                                                                                         'CloseDate': str(futureDate),
                                                                                         'End_Date__c': str(futureDate),
                                                                                         'Artwork_Due_Date__c': str(futureDate),
                                                                                         'AccountId': accountID,
                                                                                         'Parent_Opportunity__c': parentOppDetails['id'],
                                                                                         'Probability__c': '75',
                                                                                         'Estimated_Average_CPS__c': 1,
                                                                                         'Estimated_Store_Count__c': 1,
                                                                                         'Business_Type__c': 'New',
                                                                                         'Status__c': 'Reserved-RS1',
                                                                                         'Type': 'Tactical',
                                                                                         'Order__c': p1OrderNumber,
                                                                                         'Parent_Order__c': '',
                                                                                         'ILocArtworkDueDate__c': str(futureDate),
                                                                                         'ILocStoreCount__c': 10,
                                                                                         'ILocTradeClass__c': 'TRADE CLASS',
                                                                                         'ILocCategory__c': 'CATEGORY1, CATEGORY2, CATEGORY3',
                                                                                         'ILocBrand__c': 'BRAND',
                                                                                         'ILocLocType__c': 'TYPE',
                                                                                         'ILocProgram__c': 'PROGRAM',
                                                                                         'ILocGeography__c': '20',
                                                                                         'ILOCProjectCode__c': projectCode,
                                                                                         'Billed_based_on_Actual_Execution__c': True})
                    oppDetails['C_CYCLE'] = str(instoreCycleName)
                    oppDetails['C_STATUS'] = 'Reserved-RS1'
                    oppDetails['C_ILocStoreCount'] = 10
                    oppDetails['C_ILocTradeClass'] = 'TRADE CLASS'
                    oppDetails['C_ILocCategory'] = 'CATEGORY1, CATEGORY2, CATEGORY3'
                    oppDetails['C_ILocBrand'] = 'BRAND'
                    oppDetails['C_ILocLocType'] = 'TYPE'
                    oppDetails['C_ILocProgram'] = 'PROGRAM'
                    oppDetails['C_ILocGeography'] = '20'
                    oppDetails['C_ILOCProjectCode'] = projectCode
                    oppDetails['C_Billed_based_on_Actual_Execution'] = True
                    
                    if child1OppDetails["success"] == True:
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Id", 'id', child1OppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CurrencyIsoCode", 'text', currencyCode, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeId", 'id', str(opprtunityPLMap[productLine]), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeName", 'string', str(productLine), child1OppDetails['id'])                        
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Name", 'string', str(child1OpportunityName), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "PARENT_OPP_ID", 'id', parentOppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "StageName", 'text', 'Contract', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "InStore_Cycle__c", 'text', str(instoreCycle), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CloseDate", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "End_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Artwork_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "AccountId", 'reference', str(accountID), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Probability__c", 'text', '75', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Estimated_Average_CPS__c", 'currency', 1, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Estimated_Store_Count__c", 'currency', 1, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Business_Type__c", 'text', 'New', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Status__c", 'text', 'Reserved-RS1', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Type", 'text', 'Tactical', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Order__c", 'text', str(p1OrderNumber), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Parent_Order__c", 'text', '', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocArtworkDueDate__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocStoreCount__c", 'text', '1', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocTradeClass__c", 'text', 'TRADE CLASS', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocCategory__c", 'text', 'CATEGORY1, CATEGORY2, CATEGORY3', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocBrand__c", 'text', 'BRAND', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocLocType__c", 'text', 'TYPE', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocProgram__c", 'text', 'PROGRAM', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocGeography__c", 'text', '20', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOCProjectCode__c", 'text', str(projectCode), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])
                elif productLine in ('FSI'):
                    child1OppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine],
                                                                                         'Opportunity_Category__c': 'Child',
                                                                                         'Name': child1OpportunityName,
                                                                                         'StageName': 'Contract',
                                                                                         'NAPI_Insert_date__c': str(napiDateId),
                                                                                         'CloseDate': str(napiDate),
                                                                                         'End_Date__c': str(napiDate),
                                                                                         'AccountId': accountID,
                                                                                         'Parent_Opportunity__c': parentOppDetails['id'],
                                                                                         'Probability__c': '75',
                                                                                         'Estimated_Average_CPM__c': 1,
                                                                                         'Expected_Circulation__c': 1,
                                                                                         'Business_Type__c': 'New',
                                                                                         'Status__c': 'Reserved-RS1',
                                                                                         'Type': 'Standard',
                                                                                         'ILocCirculationCharges__c': '[{"ChargeType":"C","Charges":1000,"Description":"CIRCULATION CHARGE","Amount":1,"CirculationQty":1000}]',
                                                                                         'ILOCProductionCharges__c': '[{"ChargeType":"P","Charges":2000,"Description":"DISK HANDLING","Amount":1000,"CirculationQty":2}]',
                                                                                         'ILocOtherCharges__c': '[{"ChargeType":"o","Charges":5000,"Description":"OTHER CHARGE","Amount":50, "CirculationQty": 100}]',
                                                                                         'ILOCTotalProgramFee__c': 8000,
                                                                                         'Order__c': p1OrderNumber,
                                                                                         'Parent_Order__c': '',
                                                                                         'ILOCAdDescription__c': 'Description',
                                                                                         'ILocCategory__c': 'Category',
                                                                                         'ILocMarketListDueDate__c': str(napiDate),
                                                                                         'ILocArtworkDueDate__c': str(napiDate),
                                                                                         'Billed_based_on_Actual_Execution__c': True})
                    
                    if child1OppDetails["success"] == True: 
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Id", 'id', child1OppDetails['id'], child1OppDetails['id'])     
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CurrencyIsoCode", 'text', currencyCode, child1OppDetails['id'])    
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeId", 'id', str(opprtunityPLMap[productLine]), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeName", 'string', str(productLine), child1OppDetails['id'])                        
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Name", 'string', str(child1OpportunityName), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "PARENT_OPP_ID", 'id', parentOppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "StageName", 'text', 'Contract', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CloseDate", 'text', napiDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "End_Date__c", 'text', napiDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "NAPI_Insert_date__c", 'text', napiDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "AccountId", 'reference', str(accountID), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Probability__c", 'text', '75', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Estimated_Average_CPM__c", 'currency', 1, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Expected_Circulation__c", 'currency', 1, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Business_Type__c", 'text', 'New', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Status__c", 'text', 'Reserved-RS1', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Type", 'text', 'Standard', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocCirculationCharges__c", 'textarea', r'[{"ChargeType":"C","Charges":1000,"Description":"CIRCULATION CHARGE","Amount":1,"CirculationQty":1000}]', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOCProductionCharges__c", 'textarea', r'[{"ChargeType":"P","Charges":2000,"Description":"DISK HANDLING","Amount":1000,"CirculationQty":2}]', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocOtherCharges__c", 'textarea', r'[{"ChargeType":"o","Charges":5000,"Description":"OTHER CHARGE","Amount":50, "CirculationQty": 100}]', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOCTotalProgramFee__c", 'text', '8000', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Order__c", 'text', str(p1OrderNumber), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Parent_Order__c", 'text', '', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOCAdDescription__c", 'text', 'Description', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocCategory__c", 'text', 'Category', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocMarketListDueDate__c", 'text', napiDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocArtworkDueDate__c", 'text', napiDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])
                    oppDetails['C_NAPI_DATE'] = str(napiDateId)
                    oppDetails['C_STATUS'] = 'Reserved-RS1'
                    oppDetails['C_TOTAL_PROGRAM_FEE'] = 8000
                    oppDetails['C_ILOCAdDescription'] = 'Description'
                    oppDetails['C_ILocMarketListDueDate'] = str(
                        napiDateId)
                    oppDetails['C_ILocArtworkDueDate'] = str(napiDateId)
                    oppDetails['C_ILocCategory'] = 'CATEGORY1, CATEGORY2, CATEGORY3'
                    oppDetails['C_Billed_based_on_Actual_Execution'] = True
                elif productLine in ('SSMG'):
                    child1OppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine],
                                                                                         'Opportunity_Category__c': 'Child',
                                                                                         'Name': child1OpportunityName,
                                                                                         'StageName': 'Contract',
                                                                                         'Insert_Date__c': str(futureDate),
                                                                                         'CloseDate': str(futureDate),
                                                                                         'End_Date__c': str(futureDate),
                                                                                         'AccountId': accountID,
                                                                                         'Parent_Opportunity__c': parentOppDetails['id'],
                                                                                         'Probability__c': '75',
                                                                                         'Estimated_Average_CPM__c': 1,
                                                                                         'Expected_Circulation__c': 1,
                                                                                         'Business_Type__c': 'New',
                                                                                         'Type': 'Standard',
                                                                                         'Order__c': p1OrderNumber,
                                                                                         'Parent_Order__c': '',
                                                                                         'ILocMarketListDueDate__c': str(futureDate),
                                                                                         'ILoc_Material_Due_Date__c': str(futureDate),
                                                                                         'Billed_based_on_Actual_Execution__c': True})
                    if child1OppDetails["success"] == True:  
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Id", 'id', child1OppDetails['id'], child1OppDetails['id'])     
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CurrencyIsoCode", 'text', currencyCode, child1OppDetails['id'])    
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeId", 'id', str(opprtunityPLMap[productLine]), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeName", 'string', str(productLine), child1OppDetails['id'])                        
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Name", 'string', str(child1OpportunityName), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "PARENT_OPP_ID", 'id', parentOppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "StageName", 'text', 'Contract', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Insert_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CloseDate", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "End_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "AccountId", 'reference', str(accountID), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Probability__c", 'text', '75', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Estimated_Average_CPM__c", 'currency', 1, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Expected_Circulation__c", 'currency', 1, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Business_Type__c", 'text', 'New', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Status__c", 'text', 'Reserved-RS1', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Type", 'text', 'Standard', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Order__c", 'text', str(p1OrderNumber), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Parent_Order__c", 'text', '', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocMarketListDueDate__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Material_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])

                    oppDetails['C_ILocMarketListDueDate'] = str(
                        futureDate)
                    oppDetails['C_ILoc_Material_Due_Date'] = str(
                        futureDate)
                    oppDetails['C_Billed_based_on_Actual_Execution'] = True
                elif productLine in ('Checkout 51'):
                    child1OppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine],
                                                                                         'Opportunity_Category__c': 'Child',
                                                                                         'Name': child1OpportunityName,
                                                                                         'StageName': 'Contract',
                                                                                         'Insert_Date__c': str(futureDate),
                                                                                         'CloseDate': str(futureDate),
                                                                                         'End_Date__c': str(futureDate),
                                                                                         'AccountId': accountID,
                                                                                         'Parent_Opportunity__c': parentOppDetails['id'],
                                                                                         'Probability__c': '75',
                                                                                         'Business_Type__c': 'New',
                                                                                         'Order__c': p1OrderNumber,
                                                                                         'Parent_Order__c': '',
                                                                                         'ILoc_Assest_Due_Date__c': str(futureDate),
                                                                                         'Billed_based_on_Actual_Execution__c': True})
                    if child1OppDetails["success"] == True:   
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Id", 'id', child1OppDetails['id'], child1OppDetails['id'])     
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CurrencyIsoCode", 'text', currencyCode, child1OppDetails['id'])    
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeId", 'id', str(opprtunityPLMap[productLine]), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeName", 'string', str(productLine), child1OppDetails['id'])                        
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Name", 'string', str(child1OpportunityName), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "PARENT_OPP_ID", 'id', parentOppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "StageName", 'text', 'Contract', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Insert_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CloseDate", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "End_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "AccountId", 'reference', str(accountID), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Probability__c", 'text', '75', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Business_Type__c", 'text', 'New', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Order__c", 'text', str(p1OrderNumber), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Parent_Order__c", 'text', '', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Assest_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])
                    
                    oppDetails['C_ILoc_Assest_Due_Date'] = str(futureDate)
                    oppDetails['C_Billed_based_on_Actual_Execution'] = True
                elif productLine in ('Merchandising'):
                    projectCode = 600093
                    child1OppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine],
                                                                                         'Opportunity_Category__c': 'Child',
                                                                                         'Name': child1OpportunityName,
                                                                                         'StageName': 'Contract',
                                                                                         'Insert_Date__c': str(futureDate),
                                                                                         'CloseDate': str(futureDate),
                                                                                         'End_Date__c': str(futureDate),
                                                                                         'AccountId': accountID,
                                                                                         'Parent_Opportunity__c': parentOppDetails['id'],
                                                                                         'Probability__c': '75',
                                                                                         'Business_Type__c': 'New',
                                                                                         'Status__c': 'Reserved-RS1',
                                                                                         'Type': 'Subscription',
                                                                                         'Order__c': p1OrderNumber,
                                                                                         'Parent_Order__c': '',
                                                                                         'ILOCProjectCode__c': projectCode,
                                                                                         'ILOC_Store_Type__c': 'Store Type',
                                                                                         'ILOC_Tentative_geography__c': True,
                                                                                         'ILocStoreCount__c': 10,
                                                                                         'ILocTradeClass__c': 'Trade Class',
                                                                                         'ILocRemoveWave__c': True,
                                                                                         'ILocCategory__c': 'AIRTEL, IDEA, JIO, VODAFONE, TATA DOCOMO',
                                                                                         'Job_Description_Comments__c': 'INSTALLATION, MAINTENANACE, OFFERS, FITTING',
                                                                                         'Billed_based_on_Actual_Execution__c': True,
                                                                                         'ILocSignedLOCORStoreListDueDt__c': str(futureDate),
                                                                                         'POS_Materials_Produced_If_yes_above__c': 'NMCIS',
                                                                                         'ILocProgramDocumentDueDate__c': str(futureDate),
                                                                                         'POS_Material_Disposition__c': r"Return to Client at Client's expense",
                                                                                         'ILocArtworkDueDate__c': str(futureDate),
                                                                                         'Network_Retailer_List__c': 'AIRTEL, IDEA, JIO, VODAFONE, TATA DOCOMO',
                                                                                         'ILoc_Material_Due_Date__c': str(futureDate),
                                                                                         'Non_Network_Retailer_List__c': 'AIRTEL, IDEA, JIO, VODAFONE, TATA DOCOMO',
                                                                                         'Client_Supplied_Materials_Due_To__c': 'Other Vendor',
                                                                                         'Image_Request_Details__c': 'Subset List',
                                                                                         'Subset_List_Count__c': 46})
                    if child1OppDetails["success"] == True:   
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Id", 'id', child1OppDetails['id'], child1OppDetails['id'])     
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CurrencyIsoCode", 'text', currencyCode, child1OppDetails['id'])    
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeId", 'id', str(opprtunityPLMap[productLine]), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeName", 'string', str(productLine), child1OppDetails['id'])                        
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Name", 'string', str(child1OpportunityName), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "PARENT_OPP_ID", 'id', parentOppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "StageName", 'text', 'Contract', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Insert_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CloseDate", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "End_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "AccountId", 'reference', str(accountID), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Probability__c", 'text', '75', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Business_Type__c", 'text', 'New', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Status__c", 'text', 'Reserved-RS1', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Type", 'text', 'Subscription', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Order__c", 'text', str(p1OrderNumber), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Parent_Order__c", 'text', '', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Assest_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOC_Store_Type__c", 'text', 'Store Type', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOC_Tentative_geography__c", 'text', 'True', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocStoreCount__c", 'text', '10', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocTradeClass__c", 'text', 'Trade Class', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocRemoveWave__c", 'text', 'True', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocCategory__c", 'string', 'AIRTEL, IDEA, JIO, VODAFONE, TATA DOCOMO', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Job_Description_Comments__c", 'string', 'INSTALLATION, MAINTENANACE, OFFERS, FITTING', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocSignedLOCORStoreListDueDt__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "POS_Materials_Produced_If_yes_above__c", 'text', 'NMCIS', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocProgramDocumentDueDate__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "POS_Material_Disposition__c", 'string', r"Return to Client at Client expense", child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocArtworkDueDate__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Network_Retailer_List__c", 'string', 'AIRTEL, IDEA, JIO, VODAFONE, TATA DOCOMO', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Material_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Non_Network_Retailer_List__c", 'string', 'AIRTEL, IDEA, JIO, VODAFONE, TATA DOCOMO', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Client_Supplied_Materials_Due_To__c", 'text', 'Other Vendor', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Image_Request_Details__c", 'text', 'Image_Request_Details__c', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Subset_List_Count__c", 'text', '46', child1OppDetails['id'])
                    oppDetails['C_STATUS'] = 'Reserved-RS1'
                    oppDetails['C_ILOCProjectCode'] = projectCode
                    oppDetails['C_ILOC_Store_Type'] = 'Store Type'
                    oppDetails['C_ILOC_Tentative_geography'] = True
                    oppDetails['C_ILocStoreCount'] = 10
                    oppDetails['C_ILocTradeClass'] = 'Trade Class'
                    oppDetails['C_ILocRemoveWave'] = True
                    oppDetails['C_ILocCategory'] = 'AIRTEL IDEA JIO VODAFONE TATA DOCOMO'
                    oppDetails['C_Job_Description_Comments'] = 'INSTALLATION MAINTENANACE OFFERS FITTING'
                    oppDetails['C_Billed_based_on_Actual_Execution'] = True
                    oppDetails['C_ILocSignedLOCORStoreListDueDt'] = str(
                        futureDate)
                    oppDetails['C_POS_Materials_Produced_If_yes_above'] = 'NMCIS'
                    oppDetails['C_ILocProgramDocumentDueDate'] = str(
                        futureDate)
                    oppDetails['C_POS_Material_Disposition'] = r"Return to Client at Client's expense"
                    oppDetails['C_ILocArtworkDueDate'] = str(futureDate)
                    oppDetails['C_Network_Retailer_List'] = 'AIRTEL IDEA JIO VODAFONE TATA DOCOMO'
                    oppDetails['C_ILoc_Material_Due_Date'] = str(
                        futureDate)
                    oppDetails['C_Non_Network_Retailer_List'] = 'AIRTEL IDEA JIO VODAFONE TATA DOCOMO'
                    oppDetails['C_Client_Supplied_Materials_Due_To'] = 'Other Vendor'
                    oppDetails['C_Image_Request_Details'] = 'Subset List'
                    oppDetails['C_Subset_List_Count'] = 46
                elif productLine in ('SmartSource Direct', 'SmartSource Direct- Canada'):
                    if productLine in ('SmartSource Direct'):
                        projectCode = 549999
                    elif productLine in ('SmartSource Direct- Canada'):
                        projectCode = 599999
                    child1OppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine],
                                                                                         'Opportunity_Category__c': 'Child',
                                                                                         'Name': child1OpportunityName,
                                                                                         'StageName': 'Contract',
                                                                                         'Insert_Date__c': str(futureDate),
                                                                                         'CloseDate': str(futureDate),
                                                                                         'End_Date__c': str(futureDate),
                                                                                         'AccountId': accountID,
                                                                                         'Parent_Opportunity__c': parentOppDetails['id'],
                                                                                         'Probability__c': '75',
                                                                                         'Business_Type__c': 'New',
                                                                                         'Status__c': 'Reserved-RS1',
                                                                                         'Type': 'Standard',
                                                                                         'Order__c': p1OrderNumber,
                                                                                         'Parent_Order__c': '',
                                                                                         'ILOCProjectCode__c': projectCode,
                                                                                         'ILoc_Format__c': '4BF0D9BD 35F7 4FE6 B',
                                                                                         'ILoc_Targeting__c': 'C4C9F437 0703 4',
                                                                                         'ILocArtworkDueDate__c': str(futureDate),
                                                                                         'ILoc_Sample_Due_Date__c': str(futureDate),
                                                                                         'ILoc_Creative_Due_Date__c': str(futureDate),
                                                                                         'Billed_based_on_Actual_Execution__c': True,
                                                                                         'ILocCategory__c': '70888577, 4615, 4BDF, 8D4C, 350F21B2192D'})
                    if child1OppDetails["success"] == True: 
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Id", 'id', child1OppDetails['id'], child1OppDetails['id'])     
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CurrencyIsoCode", 'text', currencyCode, child1OppDetails['id'])    
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeId", 'id', str(opprtunityPLMap[productLine]), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeName", 'string', str(productLine), child1OppDetails['id'])                        
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Name", 'string', str(child1OpportunityName), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "PARENT_OPP_ID", 'id', parentOppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "StageName", 'text', 'Contract', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CloseDate", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "End_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Insert_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "AccountId", 'reference', str(accountID), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Probability__c", 'text', '75', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Business_Type__c", 'text', 'New', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Status__c", 'text', 'Reserved-RS1', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Type", 'text', 'Standard', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Order__c", 'text', str(p1OrderNumber), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Parent_Order__c", 'text', '', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Format__c", 'text',  '4BF0D9BD 35F7 4FE6 B', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOCProjectCode__c", 'text', str(projectCode), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Targeting__c", 'text', 'C4C9F437 0703 4', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocArtworkDueDate__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Sample_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILoc_Creative_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])                    
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILocCategory__c", 'string', '70888577, 4615, 4BDF, 8D4C, 350F21B2192D', child1OppDetails['id'])

                    oppDetails['C_STATUS'] = 'Reserved-RS1'
                    oppDetails['C_ILOCProjectCode'] = projectCode
                    oppDetails['C_ILoc_Format'] = '4BF0D9BD 35F7 4FE6 B'
                    oppDetails['C_ILoc_Targeting'] = 'C4C9F437 0703 4'
                    oppDetails['C_ILocArtworkDueDate'] = str(
                        futureDate)
                    oppDetails['C_ILoc_Sample_Due_Date'] = str(
                        futureDate)
                    oppDetails['C_ILoc_Creative_Due_Date'] = str(
                        futureDate)
                    oppDetails['C_Billed_based_on_Actual_Execution'] = True
                    oppDetails['C_ILocCategory'] = '70888577 4615 4BDF 8D4C 350F21B2192D'
                elif productLine in ('Digital', 'Digital- Canada'):
                    child1OppDetails = AccountOpportunityCreation.sf.Opportunity.create({'RecordTypeId': opprtunityPLMap[productLine],
                                                                                         'Opportunity_Category__c': 'Child',
                                                                                         'Name': child1OpportunityName,
                                                                                         'StageName': 'Contract',
                                                                                         'Insert_Date__c': str(futureDate),
                                                                                         'CloseDate': str(futureDate),
                                                                                         'End_Date__c': str(futureDate),
                                                                                         'AccountId': accountID,
                                                                                         'Parent_Opportunity__c': parentOppDetails['id'],
                                                                                         'Probability__c': '75',
                                                                                         'Business_Type__c': 'New',
                                                                                         'Status__c': 'Reserved-RS1',
                                                                                         'Type': 'Standard',
                                                                                         'Order__c': p1OrderNumber,
                                                                                         'Parent_Order__c': '',
                                                                                         'Artwork_Due_Date__c': str(futureDate),
                                                                                         'ClientListDueDate__c': str(futureDate),
                                                                                         'Billed_based_on_Actual_Execution__c': True,
                                                                                         'ILocCategory__c': 'AIRTEL IDEA JIO VODAFONE TATA DOCOMO',
                                                                                         'ILOC_Client_s_clip_rate__c': 10,
                                                                                         'ILOCProjectCode__c': 7755001})

                    if child1OppDetails["success"] == True: 
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Id", 'id', child1OppDetails['id'], child1OppDetails['id'])     
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CurrencyIsoCode", 'text', currencyCode, child1OppDetails['id'])    
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeId", 'id', str(opprtunityPLMap[productLine]), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "RecordTypeName", 'string', str(productLine), child1OppDetails['id'])                        
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Name", 'string', str(child1OpportunityName), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "PARENT_OPP_ID", 'id', parentOppDetails['id'], child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "StageName", 'text', 'Contract', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "CloseDate", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "End_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Insert_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "AccountId", 'reference', str(accountID), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Probability__c", 'text', '75', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Business_Type__c", 'text', 'New', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Status__c", 'text', 'Reserved-RS1', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Type", 'text', 'Standard', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Order__c", 'text', str(p1OrderNumber), child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Parent_Order__c", 'text', '', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOC_Client_s_clip_rate__c", 'text', '10', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ILOCProjectCode__c", 'text', '7755001', child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Artwork_Due_Date__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "ClientListDueDate__c", 'text', futureDate, child1OppDetails['id'])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Billed_based_on_Actual_Execution__c", 'text', 'True', child1OppDetails['id'])                    
                    oppDetails['C_STATUS'] = 'Reserved-RS1'
                    oppDetails['C_Artwork_Due_Date'] = str(futureDate)
                    oppDetails['C_ClientListDueDate'] = str(futureDate)
                    oppDetails['C_Billed_based_on_Actual_Execution'] = True
                    oppDetails['C_ILocCategory'] = 'AIRTEL IDEA JIO VODAFONE TATA DOCOMO'
                    oppDetails['C_ILOC_Client_s_clip_rate'] = 10
                    oppDetails['C_ILOCProjectCode'] = 7755001
                print(child1OppDetails)
                if child1OppDetails["success"] == True:
                    print(f"Child1 Opp ID of {productLine}: ",
                          child1OppDetails['id'])
                    oppDetails['C_ID'] = child1OppDetails['id']
                    oppDetails['C_NAME'] = child1OpportunityName
                    oppDetails['C_ACC_ID'] = accountID
                    oppDetails['C_STAGE'] = 'Contract'

                    usProductLineMap[productLine]['PARENT_OPP_1']['CHILD1_OPP_ID']['ID'] = child1OppDetails['id']
                    usProductLineMap[productLine]['PARENT_OPP_1']['CHILD1_OPP_ID']['ORDER#'] = str(
                        p1OrderNumber)
                    childOpps.append(child1OppDetails['id'])
                    Messages.write_message(
                        f"Child 1 Opp ID of {productLine}: {child1OppDetails['id']}")
                    data_store.spec[f"P1_C1_ID_{productLine}"] = child1OppDetails['id']
                    data_store.spec[f"P1_C1_NAME_{productLine}"] = child1OpportunityName

            print("\nChild Opportunities\n", childOpps)
    # -------------------------------------------------------------------------------------------------------------------------------------
            productsList = []
            productsList.clear()
            productName = None
            if productLine in ('InStore', 'InStore-Canada'):
                oppPLProductsSOQL = f"SELECT Id,Product_Line__c, Product_Line__r.Name,Product__c, Product__r.Name FROM Product_Junction__c where Product_Line__r.Name = 'InStore' and Product__r.IsActive = true and Product__c in (SELECT Product2Id from PricebookEntry where CurrencyIsoCode = '{currencyCode}' and IsActive = true)"
            else:
                oppPLProductsSOQL = f"SELECT Id,Product_Line__c, Product_Line__r.Name,Product__c, Product__r.Name FROM Product_Junction__c where Product_Line__r.Name = '{productLine}' and Product__r.IsActive = true and Product__c in (SELECT Product2Id from PricebookEntry where CurrencyIsoCode = '{currencyCode}' and IsActive = true)"
        #         oppPLProductsSOQL = f"SELECT Id,Product_Line__c, Product_Line__r.Name,Product__c, Product__r.Name FROM Product_Junction__c where Product_Line__r.Name = '{productLine}'"
            oppPLProductsResult = AccountOpportunityCreation.sf.query_all(
                query=oppPLProductsSOQL)
            oppPLProductsData = oppPLProductsResult['records']

            for oppPLProduct in oppPLProductsData:
                if oppPLProduct['Product__r'] != None:
                    if 'Name' in oppPLProduct['Product__r']:
                        productsList.append(oppPLProduct['Product__r']['Name'])
                        print("Found Product: ",
                              oppPLProduct['Product__r']['Name'])
            print(*productsList, sep="\n")
    # -------------------------------------------------------------------------------------------------------------------------------------
            lineItemsMap = {}
            # pdb.set_trace()
            productNameTemp = ''
            priceBookId = ''
            for idx, childOpp in enumerate(childOpps):
                IsActive = False
                while IsActive == False:
                    productName = random.choice(productsList)
                    productNameTemp = productName.replace("'", r"\'")

                    Messages.write_message(f"Product Name: {productName}")
                    data_store.spec[f"PRODUCT_NAME"] = productName
                    oppDetails['C_PRODUCT_NAME'] = productName

                    print("Product Name: ", productName)
                    # priceBookSOQL = f"SELECT Id, Name,IsActive FROM PricebookEntry where Name = '{productName}' and CurrencyIsoCode = '{currencyCode}' and IsActive = true"
                    priceBookSOQL = f"SELECT Id, IsActive FROM PricebookEntry WHERE CurrencyIsoCode = '{currencyCode}' AND IsActive = True AND Product2.Name = '{productNameTemp}' AND Product2Id in (SELECT Product__c FROM Product_Junction__c WHERE Product_Line__r.Name = '{productLine}')"
                    priceBookResult = AccountOpportunityCreation.sf.query_all(
                        query=priceBookSOQL)
                    priceBookData = priceBookResult['records']
                    if priceBookData[0]['IsActive'] != False:
                        priceBookId = priceBookData[0]['Id']
                        Messages.write_message(
                            f"Pricebook Entry ID: {priceBookId}")
                        print("Pricebook Entry ID: ", priceBookId)
                        data_store.spec["PRICE_BOOK_ID"] = priceBookId

                        if productLine in ('InStore', 'InStore-Canada'):
                            chargeDetailsSOQL = f"SELECT Id, Name, ProductLIne__c, Product__C, Charge_Type_2__c, Charge_Type_2__r.Commissionable__c, Charge_Type_2__r.Charge_Type_Categorty__c FROM Pricing_Detail__c where ProductLIne__c = 'Instore' and product__c = '{productNameTemp}' and CurrencyIsoCode = '{currencyCode}' and isActive__c = true and IsCurrent__c = true and RateCardIsCurrent__c = true"
                        elif productLine in ('Checkout 51'):
                            chargeDetailsSOQL = f"SELECT Id, Name, ProductLIne__c, Product__C, Charge_Type_2__c, Charge_Type_2__r.Commissionable__c, Charge_Type_2__r.Charge_Type_Categorty__c FROM Pricing_Detail__c where ProductLIne__c = '{productLine}' and FreedomOrNot__c != 'Freedom' and product__c = '{productNameTemp}' and CurrencyIsoCode = '{currencyCode}' and isActive__c = true and IsCurrent__c = true and RateCardIsCurrent__c = true"
                        elif productLine in ('Digital', 'Digital- Canada'):
                            chargeDetailsSOQL = f"SELECT Id, Name, ProductLIne__c, Product__C, Charge_Type_2__c, Charge_Type_2__r.Commissionable__c, Charge_Type_2__r.Charge_Type_Categorty__c FROM Pricing_Detail__c where ProductLIne__c = '{productLine}' and product__c = '{productNameTemp}' and CurrencyIsoCode = '{currencyCode}' and isActive__c = true and FreedomOrNot__c not in ('Freedom') and IsCurrent__c = true  and RateCardIsCurrent__c = true"
                        else:
                            chargeDetailsSOQL = f"SELECT Id, Name, ProductLIne__c, Product__C, Charge_Type_2__c, Charge_Type_2__r.Commissionable__c, Charge_Type_2__r.Charge_Type_Categorty__c FROM Pricing_Detail__c where ProductLIne__c = '{productLine}' and product__c = '{productNameTemp}' and CurrencyIsoCode = '{currencyCode}' and isActive__c = true and IsCurrent__c = true  and RateCardIsCurrent__c = true"
                        chargeDetailsResult = AccountOpportunityCreation.sf.query_all(
                            query=chargeDetailsSOQL)
                        chargeDetailsData = chargeDetailsResult['records']
                        print("\n", chargeDetailsSOQL, "\n")
                        if len(chargeDetailsData) > 2:
                            chargeTypeMap = {}
                            chargeTypeMap.clear()
                            for chargeDetails in chargeDetailsData:
                                chargeTypeMap[chargeDetails['Name']] = {"Id": chargeDetails['Id'], "ChargeType": chargeDetails[
                                    'Charge_Type_2__c'], 'Commissionable': chargeDetails['Charge_Type_2__r']['Commissionable__c'], 'Charge_Type_Category': chargeDetails['Charge_Type_2__r']['Charge_Type_Categorty__c']}
                            # print("\nCharge Types\n", chargeTypeMap)
                            [print(key, value)
                                for key, value in chargeTypeMap.items()]
                            Messages.write_message(
                                f"Charge Types MAP \n{chargeTypeMap}")
                            IsActive = True

                totalAmount = 0

                for cnt in range(1, 5):
                    chargeType = random.choice(list(chargeTypeMap))
                    chargeTypeTemp = chargeType.replace("'", r"\'")
                    soql = f"SELECT Charge_Type_2__r.Commissionable__c,Charge_Type__c, Charge_Type_Category__c, CurrencyIsoCode, Id, isActive__c, Name, Product__c, Quantity_1000_effective_from_date__c, Quantity_1000__c FROM Pricing_Detail__c WHERE Name = '{chargeTypeTemp}' and Product__c= '{productNameTemp}' and CurrencyIsoCode = '{currencyCode}' and isActive__c = True and RateCardIsCurrent__c = true and IsCurrent__c = true"
                    queryResult = AccountOpportunityCreation.sf.query_all(
                        query=soql)
                    recDetails = queryResult['records']
                    if "CLIENT_ORDER_LIMIT_GT_ILOC_TOTAL" in oppName:
                        quantity = randint(10, 10)
                        salesPrice = round(random.uniform(10, 10), 2)
                    elif "CLIENT_ORDER_LIMIT_LT_ILOC_TOTAL" in oppName:
                        quantity = randint(5000, 5000)
                        salesPrice = round(random.uniform(50, 50), 2)
                    else:
                        quantity = randint(1, 100)
                        salesPrice = round(random.uniform(1.1, 9.9), 2)

                    if recDetails[0]['Quantity_1000__c'] == True:
                        totalPrice = round((quantity / 1000) * salesPrice, 2)
                    else:
                        totalPrice = round((quantity * salesPrice), 2)
                    totalAmount = totalAmount + totalPrice

                    # pdb.set_trace()
                    lineItemsDetails = AccountOpportunityCreation.sf.OpportunityLineItem.create({'Charge_Type__c': chargeTypeMap[chargeType]["ChargeType"], 'Commissionable__c': chargeTypeMap[chargeType]["Commissionable"], 'pricebookentryid': priceBookId,
                                                                                                 'Pricing_Detail__c': chargeType, 'PricingDetail__c': chargeTypeMap[chargeType]["Id"], 'opportunityId': childOpp, 'Quantity': quantity, 'Sales_price__c': salesPrice, 'TotalPrice': totalPrice, 'Component__c': 'Video'})

                    if lineItemsDetails["success"] == True:
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Id", 'id', lineItemsDetails["id"], lineItemsDetails["id"])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Opportunity__c", 'reference', child1OppDetails['id'], lineItemsDetails["id"])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "product__c", 'string', productNameTemp, lineItemsDetails["id"])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "ProductLIne__c", 'string', productLine, lineItemsDetails["id"])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "CurrencyIsoCode", 'text', currencyCode, lineItemsDetails["id"])
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "priceBookId", 'string', priceBookId, lineItemsDetails["id"])
                        oppDetails[f'C_{cnt}_CHARGE_TYPE'] = chargeTypeMap[chargeType]["ChargeType"]
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Charge_Type__c", 'string', chargeTypeMap[chargeType]["ChargeType"], lineItemsDetails["id"])
                        oppDetails[f'C_{cnt}_CHARGE_TYPE_CATEGORY'] = chargeTypeMap[chargeType]["Charge_Type_Category"]
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Charge_Type_Category__c", 'string', chargeTypeMap[chargeType]["Charge_Type_Category"], lineItemsDetails["id"]) 
                        oppDetails[f'C_{cnt}_COMMISSIONABLE'] = chargeTypeMap[chargeType]["Commissionable"]
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Commissionable__c", 'text', chargeTypeMap[chargeType]["Commissionable"], lineItemsDetails["id"])
                        oppDetails[f'C_{cnt}_PRICING_DETAIL_NAME'] = chargeType
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Pricing_Detail__c", 'string', chargeType, lineItemsDetails["id"])
                        oppDetails[f'C_{cnt}_PRICING_DETAIL_ID'] = lineItemsDetails["id"]
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "PricingDetail__c", 'reference', chargeTypeMap[chargeType]["Id"], lineItemsDetails["id"])
                        oppDetails[f'C_{cnt}_QTY'] = quantity
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Quantity", 'double', quantity, lineItemsDetails["id"])
                        oppDetails[f'C_{cnt}_SALES_PRICE'] = salesPrice
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Sales_price__c", 'double', salesPrice, lineItemsDetails["id"])
                        oppDetails[f'C_{cnt}_SUB_TOTAL'] = totalPrice
                        Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "TotalPrice", 'double', totalPrice, lineItemsDetails["id"])
                        if recDetails[0]['Quantity_1000__c'] == True:
                            Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Quantity_1000__c", 'text', recDetails[0]['Quantity_1000__c'], lineItemsDetails["id"])
                            Common_Steps.CommonSteps.write_data_to_table_column(self, "OLI", "Quantity_1000_effective_from_date__c", 'text', recDetails[0]['Quantity_1000_effective_from_date__c'], lineItemsDetails["id"])
                        print(
                            f"OPP[{childOpp}] Line Item {cnt}: {lineItemsDetails['id']} created..")
                        Messages.write_message(
                            f"OPP[{childOpp}] Line Item {cnt}: {lineItemsDetails['id']} created..")
                        lineItemsMap[childOpp +
                                     str(cnt)] = {cnt: lineItemsDetails['id']}
                oppDetails['C_TOTAL_AMOUNT'] = totalAmount
                if productLine in ('FSI') and productName != 'Remnant':
                    Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Total", 'currency', '8000', child1OppDetails['id'])
                else:
                    Common_Steps.CommonSteps.write_data_to_table_column(self, "Opportunity", "Total", 'currency', totalAmount, child1OppDetails['id'])
                # pdb.set_trace()

            columnCount = 0
            rowCount = 0

            df = pd.read_excel(opportunitiesDetailsFileName,
                               sheet_name=f"{productLine}_{currencyCode}")
            Messages.write_message(df.columns)
            columnList = df.columns.tolist()
            [print(key, value) for key, value in oppDetails.items()]
            rowCount = int(ws.max_row) + 1
            for key, value in oppDetails.items():
                #         pdb.set_trace()
                if key in columnList:
                    ws[f"{xlsxwriter.utility.xl_col_to_name(columnList.index(key))}{rowCount}"] = value
                else:
                    if len(columnList) == 0 and ws["A1"].value == None:
                        ws[f"{xlsxwriter.utility.xl_col_to_name(columnCount)}{1}"] = key
                        ws[f"{xlsxwriter.utility.xl_col_to_name(columnCount)}{2}"] = value
                    else:
                        columnCount = int(ws.max_column)
                        ws[f"{xlsxwriter.utility.xl_col_to_name(columnCount)}{1}"] = key
                        ws[f"{xlsxwriter.utility.xl_col_to_name(columnCount)}{rowCount}"] = value
                wb.save(filename=opportunitiesDetailsFileName)
            wb.save(filename=opportunitiesDetailsFileName)

    @after_step
    def after_step_hook(self, context):
        if context.step.is_failing == True:
            Messages.write_message(context.step.text)
            # Messages.write_message(context.step.message)
            Screenshots.capture_screenshot()

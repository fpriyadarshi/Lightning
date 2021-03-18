from getgauge.python import data_store, step, Messages, after_spec, after_step
from getgauge.python import Screenshots
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import os
import sys
import random
import string
from datetime import date
from step_impl import Drivers
from selenium.webdriver.support.select import Select
from random import randint
from time import sleep
from step_impl import Utils
from selenium.common.exceptions import NoSuchElementException
import shutil
from selenium.webdriver.common.action_chains import ActionChains
from pathlib import Path
import pandas as pd
import xlsxwriter
import csv
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl import Workbook
from xlrd import open_workbook
import pdb
from decimal import Decimal
from selenium.common.exceptions import TimeoutException, UnexpectedAlertPresentException
import json
import re
import sqlite3
from selenium.webdriver.support.wait import WebDriverWait


class CommonSteps:

    def fetch_record_details(self, searchSOQL, fetchSOQL):
        Drivers.dbConn.row_factory = sqlite3.Row        
        cur = Drivers.dbConn.cursor()
        recordCount = 0
        tableRecords = None
        tableRecords = cur.execute(searchSOQL)
        recordCount = tableRecords.fetchone()[0]
        print("Record(s) Found: ", recordCount)      
        if recordCount  == 1:
            tableRecords = cur.execute(fetchSOQL)            
        return tableRecords

    def write_data_to_table_column(self, tableName, columnName, columnType, columnValue, IdValue):
        createTable = f"CREATE TABLE IF NOT EXISTS {tableName}(Id STRING(20))"
        tableData  = Drivers.dbCursor.execute(createTable)
        # print("Table Created", tableData)
        # Messages.write_message(f"Table Created {tableData}")
        
        tableinfo = [i[1] for i in Drivers.dbCursor.execute(f"PRAGMA table_info('{tableName}')")]
        print(tableinfo)
        # Messages.write_message(f"Table Info {tableinfo}")
        
        tableRecords = Drivers.dbCursor.execute(f"SELECT EXISTS(SELECT 1 FROM {tableName} WHERE Id = '{IdValue}')")
        recordCount = tableRecords.fetchone()[0]
        print("Record(s) Found: ", recordCount)      
        if recordCount  == 0:
            insertValue = Drivers.dbCursor.execute(f"INSERT INTO {tableName} (Id) VALUES ('{IdValue}')")
            Drivers.dbConn.commit()

        if columnName not in tableinfo:
            if columnType in ('id','reference'):
                columnSet = f'{columnName} STRING(20)'
            elif columnType in ('picklist', 'string', 'text'):
                columnSet = f'{columnName} STRING(50)'
            elif columnType in ('text'):
                columnSet = f'{columnName} STRING(20)'
            elif columnType in ('textarea', 'multipicklist'):
                columnSet = f'{columnName} VARCHAR(200)'
            elif columnType in ('double', 'currency'):
                columnSet = f'{columnName} DOUBLE(20,2)'
            elif columnType in ('boolean'):
                columnSet = f'{columnName} BOOLEAN'
            elif columnType in ('date'):
                columnSet = f'{columnName} DATE'

            tableUpdate = Drivers.dbCursor.execute(f"ALTER TABLE {tableName} ADD COLUMN {columnSet}")
            
            if columnType in ('textarea', 'reference', 'datetime', 'picklist', 'string', 'id', 'multipicklist', 'text'):
                columnValue = str(re.sub('[^A-Z-a-z#0-9._, ]+', '', str(columnValue)))
                updateRecord = Drivers.dbCursor.execute(f"UPDATE {tableName} SET {columnName} = '{columnValue}' WHERE Id = '{IdValue}'")
            elif columnType in ('date', 'boolean', 'double', 'currency'):
                updateRecord = Drivers.dbCursor.execute(f"UPDATE {tableName} SET {columnName} = {columnValue} WHERE Id = '{IdValue}'")

            print(f"Inserting column {columnName} as {columnValue} where {IdValue}")
            # Messages.write_message(f"Inserting column {columnName} as {columnValue} where {IdValue}")
            Drivers.dbConn.commit()                       
        else:                        
            
            print(f"Updating column {columnName} as {columnValue} where {IdValue}")
            # Messages.write_message(f"Updating column {columnName} as {columnValue} where {IdValue}")
            if columnType in ('textarea', 'reference', 'datetime', 'picklist', 'string', 'id', 'multipicklist', 'text'):
                columnValue = str(re.sub('[^A-Z-a-z#0-9._, ]+', '', str(columnValue)))
                updateRecord = Drivers.dbCursor.execute(f"UPDATE {tableName} SET {columnName} = '{columnValue}' WHERE Id = '{IdValue}'")
            elif columnType in ('date', 'boolean', 'double', 'currency'):
                updateRecord = Drivers.dbCursor.execute(f"UPDATE {tableName} SET {columnName} = {columnValue} WHERE Id = '{IdValue}'")
            Drivers.dbConn.commit()                   
        # tableinfo = [i[1] for i in Drivers.dbCursor.execute(f"PRAGMA table_info('{tableName}')")]
        # print(tableinfo)
        # Messages.write_message(f"Table Info {tableinfo}")
        
    def write_data_to_table(self, objectName, objectFieldInfo, tableName, dataToWrite):
        createTable = f"CREATE TABLE IF NOT EXISTS '{tableName}'(Id STRING(20))"
        tableData  = Drivers.dbCursor.execute(createTable)
        print("Table Created", tableData)
        # Messages.write_message(f"Table Created {tableData}")
        
        tableinfo = [i[1] for i in Drivers.dbCursor.execute(f"PRAGMA table_info('{tableName}')")]
        print(tableinfo)
        # Messages.write_message(f"Table Info {tableinfo}")
        
        tableRecords = Drivers.dbCursor.execute(f"SELECT EXISTS(SELECT 1 FROM Form_Header WHERE Id = '{dataToWrite['Id']}')")
        recordCount = tableRecords.fetchone()[0]
        print("Record(s) Found: ", recordCount)        
        if recordCount  == 0:
            insertValue = Drivers.dbCursor.execute(f"INSERT INTO {tableName} (Id) VALUES ('{dataToWrite['Id']}')")
            Drivers.dbConn.commit()
        
        objectColumn = []
        objectValue = []
        for key, value in dataToWrite.items():
            print(f"Column {key} as {value} where {dataToWrite['Id']}")
            # Messages.write_message(f"Column {key} as {value} where {dataToWrite['Id']}")
            if key in objectFieldInfo and value != None:
                columnSet = ''
                objectColumn.append(key)
                objectValue.append(value)
                
                if key not in tableinfo:
                    if objectFieldInfo[key] in ('id','reference'):
                        columnSet = f'{key} STRING(20)'
                    elif objectFieldInfo[key] in ('picklist', 'string', 'text'):
                        columnSet = f'{key} STRING(50)'
                    elif objectFieldInfo[key] in ('textarea', 'multipicklist'):
                        columnSet = f'{key} VARCHAR(200)'
                    elif objectFieldInfo[key] in ('double', 'currency'):
                        columnSet = f'{key} DOUBLE(20,2)'
                    elif objectFieldInfo[key] in ('boolean'):
                        columnSet = f'{key} BOOLEAN'
                    elif objectFieldInfo[key] in ('date'):
                        columnSet = f'{key} DATE'

                    tableUpdate = Drivers.dbCursor.execute(f"ALTER TABLE {tableName} ADD COLUMN {columnSet}")
                    if objectFieldInfo[key] in ('textarea', 'reference', 'datetime', 'picklist', 'string', 'id', 'multipicklist', 'text'):
                        value = str(re.sub('[^A-Z-a-z#0-9._, ]+', '', str(value)))
                        updateRecord = Drivers.dbCursor.execute(f"UPDATE {tableName} SET {key} = '{value}' WHERE Id = '{dataToWrite['Id']}'")
                    elif objectFieldInfo[key] in ('date', 'boolean', 'double', 'currency'):
                        updateRecord = Drivers.dbCursor.execute(f"UPDATE {tableName} SET {key} = {value} WHERE Id = '{dataToWrite['Id']}'")

                    print(f"Inserting column {key} as {value} where {dataToWrite['Id']}")
                    # Messages.write_message(f"Inserting column {key} as {value} where {dataToWrite['Id']}")
                    Drivers.dbConn.commit()                       
                else:
                    print(f"Updating column {key} as {value} where {dataToWrite['Id']}")
                    # Messages.write_message(f"Updating column {key} as {value} where {dataToWrite['Id']}")
                    if objectFieldInfo[key] in ('textarea', 'reference', 'datetime', 'picklist', 'string', 'id', 'multipicklist', 'text'):
                        value = str(re.sub('[^A-Z-a-z#0-9._, ]+', '', str(value)))
                        updateRecord = Drivers.dbCursor.execute(f"UPDATE {tableName} SET {key} = '{value}' WHERE Id = '{dataToWrite['Id']}'")
                    elif objectFieldInfo[key] in ('date', 'boolean', 'double', 'currency'):
                        updateRecord = Drivers.dbCursor.execute(f"UPDATE {tableName} SET {key} = {value} WHERE Id = '{dataToWrite['Id']}'")
                    Drivers.dbConn.commit()                   
        # tableinfo = [i[1] for i in Drivers.dbCursor.execute(f"PRAGMA table_info('{tableName}')")]
        # print(tableinfo)
        # Messages.write_message(f"Table Info {tableinfo}")
        # columnString = ', '.join("{0}".format(accs) for accs in objectColumn)
        # valueString = ', '.join("{0}".format(accs) for accs in objectValue)
        # print(columnString, " ", len(objectColumn))
        # print(valueString, " ", len(objectValue))
        
    def get_column_index(self):
        labelsMap = {}
        try:
            tableHeader = Drivers.driver.find_elements_by_xpath(
                "//table[@id='SelectedProductType']/thead/tr/th")
            columnIndex = 1
            for theader in tableHeader:
                labelXpath = "//table[@id='SelectedProductType']/thead/tr/th[" + \
                    str(columnIndex) + "]/div"
                # labelCount = "//table[@id='SelectedProductType']/thead/tr/th[count(*)=1][" + str(columnIndex) + "]"
                # labelName = Drivers.driver.find_element_by_xpath(labelXpath)
                # print("Label Count: ", labelCount)
                # columnName = theader.find_element_by_css_selector(".slds-truncate").get_attribute("title")
                columnName = theader.text.title()
                labelsMap[columnName] = columnIndex
                Messages.write_message(
                    columnName + " : " + str(columnIndex))
                columnIndex = columnIndex + 1
            return labelsMap
        except NoSuchElementException:
            print("element not found")

    def pricing_details_table(self, productName, table):
        rows_data = table.rows
        totalRows = len(rows_data)
        Messages.write_message("Total Rows : " + str(totalRows))

        tableData = {}
        columnNames = table.get_column_names()

        if "Week Start Date" in columnNames:
            weekStartdateColumnData = table.get_column_values_with_name(
                "Week Start Date")
            tableData["Week Start Date"] = weekStartdateColumnData

        if "Week End Date" in columnNames:
            weekEndDateColumnData = table.get_column_values_with_name(
                "Week End Date")
            tableData["Week End Date"] = weekEndDateColumnData

        if "Offer" in columnNames:
            offerColumnData = table.get_column_values_with_name("Offer")
            tableData["Offer"] = offerColumnData
            print(tableData["Offer"][0])

        if "Applicable Partner" in columnNames:
            applicablePartnerColumnData = table.get_column_values_with_name(
                "Applicable Partner")
            tableData["Applicable Partner"] = applicablePartnerColumnData

        if "Charge Description" in columnNames:
            chargeDescColumnData = table.get_column_values_with_name(
                "Charge Description")
            tableData["Charge Description"] = chargeDescColumnData

        if "Description" in columnNames:
            descColumnData = table.get_column_values_with_name("Description")
            tableData["Description"] = descColumnData

        if "Sales Price" in columnNames:
            salesPriceColumnData = table.get_column_values_with_name(
                "Sales Price")
            tableData["Sales Price"] = salesPriceColumnData

        if "Quantity" in columnNames:
            quantityColumnData = table.get_column_values_with_name("Quantity")
            tableData["Quantity"] = quantityColumnData

        if "% Of Redemption" in columnNames:
            percentageOfRedemptionColumnData = table.get_column_values_with_name(
                "% Of Redemption")
            tableData["% Of Redemption"] = percentageOfRedemptionColumnData

        if "Placement Number" in columnNames:
            placementNumColumnData = table.get_column_values_with_name(
                "Placement Number")
            tableData["Placement Number"] = placementNumColumnData

        if "Quantity Per Store" in columnNames:
            quantityPerStoreColumnData = table.get_column_values_with_name(
                "Quantity Per Store")
            tableData["Quantity Per Store"] = quantityPerStoreColumnData
        Messages.write_message(tableData)
        CommonSteps.pricing_details(self, totalRows, productName, tableData)

    def pricing_details(self, totalLineItem, productName, data=None):
        opportunitiesDetails = {}
        opportunitiesDetails = data_store.spec.get(productName)

        pricingDictionary = {}
        pricingDetails = {}
        sleep(2)

        columnList = CommonSteps.get_column_index(self)
        print(columnList)
        Messages.write_message(columnList)
        pricingDetails["TOTAL LINE ITEMS"] = totalLineItem
        totalLineItem = int(totalLineItem) + 1

        selectedPartner = ""
        selectedCharge = ""

        oppCurrency = None

        if productName not in ("SMARTSOURCE_DIRECT_US") and str(productName.split("_")[1]) == "US":
            oppCurrency = "USD"
        elif productName not in ("SMARTSOURCE_DIRECT_CA") and str(productName.split("_")[1]) == "CA":
            oppCurrency = "CAD"
        elif productName in ("SMARTSOURCE_DIRECT_US") and str(productName.split("_")[2]) == "US":
            oppCurrency = "USD"
        elif productName in ("SMARTSOURCE_DIRECT_CA") and str(productName.split("_")[2]) == "CA":
            oppCurrency = "CAD"

        if (productName == "DIGITAL_US" or productName == "DIGITAL_CA" or productName == "SMARTSOURCE_DIRECT_US" or productName == "SMARTSOURCE_DIRECT_CA"):
            Drivers.driverWait.until(
                EC.visibility_of_element_located((By.XPATH, "//div[span[text()='Component']]/select")))
            dropComponent = Select(
                Drivers.driver.find_element_by_xpath("//div[span[text()='Component']]/select"))
            dropDownOptionCnt = str(
                len(Drivers.driver.find_elements_by_xpath("//div[span[text()='Component']]/select/option")))
            dropComponent.select_by_index(
                randint(2, (int(dropDownOptionCnt) - 1)))
        totalAmount = 0
        for cnt in range(1, totalLineItem):
            redemption = 0
            subTotalAmount = 0
            salesPrice = 0
            quantity = 0

            print("Counter: ", cnt)
            currentDate = date.today()
            currentDate = currentDate.strftime("%m/%d/%Y")

            if "Week Start Date" in columnList:
                quantityColIndex = columnList["Week Start Date"]
                quantityXpath = "//table[@id='SelectedProductType']/tbody/tr[" + str(
                    cnt) + "]/td[" + str(quantityColIndex) + "]//input"
                Drivers.driverWait.until(
                    EC.visibility_of_element_located((By.XPATH, quantityXpath)))
                txtBoxField = Drivers.driver.find_element_by_xpath(
                    quantityXpath)
                txtBoxField.clear()
                if not data:
                    txtBoxField.send_keys(currentDate)
                else:
                    txtBoxField.send_keys(currentDate)
                pricingDetails["WEEK START DATE-" + str(cnt)] = currentDate

            if "Week End Date" in columnList:
                quantityColIndex = columnList["Week End Date"]
                quantityXpath = "//table[@id='SelectedProductType']/tbody/tr[" + str(
                    cnt) + "]/td[" + str(quantityColIndex) + "]//input"
                Drivers.driverWait.until(
                    EC.visibility_of_element_located((By.XPATH, quantityXpath)))
                txtBoxField = Drivers.driver.find_element_by_xpath(
                    quantityXpath)
                txtBoxField.clear()
                if not data:
                    txtBoxField.send_keys(currentDate)
                else:
                    txtBoxField.send_keys(currentDate)
                pricingDetails["WEEK END DATE-" + str(cnt)] = currentDate

            if "Offer" in columnList:
                offerColIndex = columnList["Offer"]
                offerXpath = "//table[@id='SelectedProductType']/tbody/tr[" + str(
                    cnt) + "]/td[" + str(offerColIndex) + "]//input"
                Drivers.driverWait.until(
                    EC.visibility_of_element_located((By.XPATH, offerXpath)))
                lower_upper_alphabet = string.ascii_letters
                txtBoxOffer = Drivers.driver.find_element_by_xpath(offerXpath)
                if not data:
                    offer = random.choice(lower_upper_alphabet)
                else:
                    offer = None
                    if "Offer" in data:
                        offer = data["Offer"][cnt - 1]
                txtBoxOffer.send_keys(offer)
                pricingDetails["OFFER-" + str(cnt)] = offer

            if "Applicable Partner" in columnList:
                applicablePartnerColIndex = columnList["Applicable Partner"]
                applicablePartnerXpath = "//table[@id='SelectedProductType']/tbody/tr[" + str(
                    cnt) + "]/td[" + str(applicablePartnerColIndex) + "]//select"
                Drivers.driverWait.until(
                    EC.visibility_of_element_located((By.XPATH, applicablePartnerXpath)))
                optionsXpath = "//table[@id='SelectedProductType']/tbody/tr[" + str(
                    cnt) + "]/td[" + str(applicablePartnerColIndex) + "]//select/option"
                dropDownProductLine = Select(
                    Drivers.driver.find_element_by_xpath(applicablePartnerXpath))
                dropDownOptionCnt = str(
                    len(Drivers.driver.find_elements_by_xpath(optionsXpath)))
                if not data:
                    dropDownProductLine.select_by_index(
                        randint(2, (int(dropDownOptionCnt) - 1)))
                else:
                    if "Applicable Partner" in data:
                        dropDownProductLine.select_by_value(
                            data["Applicable Partner"][cnt - 1])
                selectedPartner = dropDownProductLine.first_selected_option.text
                pricingDetails["APPLICABLE PARTNER-" +
                               str(cnt)] = selectedPartner

            if "Quantity Per Store" in columnList:
                quantityColIndex = columnList["Quantity Per Store"]
                quantityXpath = "//table[@id='SelectedProductType']/tbody/tr[" + str(
                    cnt) + "]/td[" + str(quantityColIndex) + "]//input"
                Drivers.driverWait.until(
                    EC.visibility_of_element_located((By.XPATH, quantityXpath)))
                txtBoxQuantity = Drivers.driver.find_element_by_xpath(
                    quantityXpath)
                txtBoxQuantity.clear()
                if not data:
                    quantityPerStore = randint(1, 99)
                else:
                    if "Quantity Per Store" in data:
                        quantityPerStore = data["Quantity Per Store"][cnt - 1]
                    else:
                        quantityPerStore = 0
                txtBoxQuantity.send_keys(str(quantityPerStore))
                pricingDetails["QUANTITY PER STORE-" +
                               str(cnt)] = quantityPerStore

            if "Placement Number" in columnList:
                placementNumberColIndex = columnList["Placement Number"]
                placementNumberXpath = "//table[@id='SelectedProductType']/tbody/tr[" + str(
                    cnt) + "]/td[" + str(placementNumberColIndex) + "]//input"
                Drivers.driverWait.until(
                    EC.visibility_of_element_located((By.XPATH, placementNumberXpath)))
                txtBoxPlacementNumber = Drivers.driver.find_element_by_xpath(
                    placementNumberXpath)
                txtBoxPlacementNumber.clear()
                if not data:
                    placementNumber = randint(1, 99)
                else:
                    placementNumber = randint(1, 99)
                    if "Placement Number" in data:
                        placementNumber = data["Placement Number"][cnt - 1]
                txtBoxPlacementNumber.send_keys(str(placementNumber))
                pricingDetails["PLACEMENT NUMBER-" +
                               str(cnt)] = placementNumber

            if "Charge Description" in columnList:
                chargeDescriptionColIndex = columnList["Charge Description"]
                chargeDescriptionXpath = "//table[@id='SelectedProductType']/tbody/tr[" + str(
                    cnt) + "]/td[" + str(chargeDescriptionColIndex) + "]//select"
                Drivers.driverWait.until(
                    EC.visibility_of_element_located((By.XPATH, chargeDescriptionXpath)))
                optionsXpath = "//table[@id='SelectedProductType']/tbody/tr[" + str(
                    cnt) + "]/td[" + str(chargeDescriptionColIndex) + "]//select/option"
                
                dropDownProductLine = Select(Drivers.driver.find_element_by_xpath(chargeDescriptionXpath))
                dropDownOptionCnt = str(
                    len(Drivers.driver.find_elements_by_xpath(optionsXpath)))
                if not data:
                    dropDownProductLine.select_by_index(
                        randint(1, (int(dropDownOptionCnt) - 1)))
                else:
                    #                     if data['Charge Description'][cnt - 1].find("'") > 0:
                    selectedCharge = data['Charge Description'][cnt - 1]
                    print(f"Adding Charge {selectedCharge}")
                    soql = f"SELECT Id,Name,Product__c,CurrencyIsoCode FROM Pricing_Detail__c WHERE Name = '{selectedCharge}' and Product__c = '{opportunitiesDetails['PRODUCT NAME']}' and CurrencyIsoCode = '{oppCurrency}'"
                    queryResult = Drivers.sf.query_all(query=soql)
                    recDetails = queryResult['records']
                    chargeDescription = recDetails[0]['Name']
                    # dropDownOptions = dropDownProductLine.options
                    # for dropDownOption in dropDownOptions:
                    dropDownProductLine.select_by_value(chargeDescription)
                selectedCharge = dropDownProductLine.first_selected_option.text.replace(
                    "'", r"\'")
                pricingDetails["CHARGE DESCRIPTION-" +
                               str(cnt)] = dropDownProductLine.first_selected_option.text
#             if selectedCharge.find("'") > 0:
#                 selectedCharge = selectedCharge.replace("'","\'")
            quantityBy1000Logic = False

            parentPath = Path(__file__).parents[1]
            quantityBy1000FileName = str(
                parentPath) + "\\Data\\" + os.getenv("QUANTITY_BY_1000_DETAILS_FILE")
            quantityBy1000Sheet = os.getenv("QUANTITY_BY_1000_DETAILS_SHEET")

            soql = f"SELECT Charge_Type_2__r.Commissionable__c,Charge_Type__c, Charge_Type_Category__c, CurrencyIsoCode, Id, isActive__c, Name, Product__c, Quantity_1000_effective_from_date__c, Quantity_1000__c FROM Pricing_Detail__c WHERE Name = '{selectedCharge}' and Product__c='{opportunitiesDetails['PRODUCT NAME']}' and CurrencyIsoCode = '{oppCurrency}' and isActive__c = True"
            queryResult = Drivers.sf.query_all(query=soql)
            recDetails = queryResult['records']
#             pdb.set_trace()

            isQtyBy1000LogicSetForCharge = False
            rowNo = 0
            for rec in recDetails:
                pricingDetails["CHARGE ID-" + str(cnt)] = rec['Id']
                pricingDetails["CHARGE NAME-" +
                               str(cnt)] = rec['Charge_Type_Category__c']
                pricingDetails["CHARGE TYPE CATEGORY-" +
                               str(cnt)] = rec['Charge_Type_Category__c']
                pricingDetails["CURRENCY CODE-" +
                               str(cnt)] = rec['CurrencyIsoCode']
                pricingDetails["COMMISSIONABLE?-" +
                               str(cnt)] = rec['Charge_Type_2__r']['Commissionable__c']
                pricingDetails["QUANTITY/1000 EFFECTIVE DATE-" +
                               str(cnt)] = rec['Quantity_1000_effective_from_date__c']
                pricingDetails["QUANTITY/1000?-" +
                               str(cnt)] = rec['Quantity_1000__c']

                if rec['Quantity_1000__c'] == True:
                    isQtyBy1000LogicSetForCharge = True

            if isQtyBy1000LogicSetForCharge == False:
                rowNo = Utils.get_row_col_num_from_xlsx(
                    quantityBy1000Sheet, quantityBy1000FileName, opportunitiesDetails['PRODUCT NAME'], opp_currency=oppCurrency)

            if rowNo > 0:
                df = pd.read_excel(quantityBy1000FileName,
                                   sheet_name=quantityBy1000Sheet)
                xlsxColumnList = df.columns.tolist()

                wb = load_workbook(
                    filename=quantityBy1000FileName, data_only=True)
                ws = wb[quantityBy1000Sheet]

                productName = ws[f"{xlsxwriter.utility.xl_col_to_name(xlsxColumnList.index('Product'))}{rowNo}"].value
                chargeName = ws[f"{xlsxwriter.utility.xl_col_to_name(xlsxColumnList.index('Charge'))}{rowNo}"].value
                effectiveDate = ws[f"{xlsxwriter.utility.xl_col_to_name(xlsxColumnList.index('Effective_Date'))}{rowNo}"].value
                produtCurrency = ws[
                    f"{xlsxwriter.utility.xl_col_to_name(xlsxColumnList.index('Currency'))}{rowNo}"].value
#                 soql = f"SELECT Id,CurrencyIsoCode,Quantity_1000_effective_from_date__c,Quantity_1000__c FROM Pricing_Detail__c WHERE Name = '{chargeName}' and Product__c='{productName}'"
#                 queryResult = Drivers.sf.query_all(query=soql)
#                 recDetails = queryResult['records']

                for rec in recDetails:
                    Messages.write_message(
                        f"Updating effective date to {effectiveDate} for record {rec['Id']}:{chargeName} of product {productName}")
                    updateRecord = Drivers.sf.Pricing_Detail__c.update(str(rec['Id']), {
                                                                       'Quantity_1000__c': True, 'Quantity_1000_effective_from_date__c': str(effectiveDate)})
                    Messages.write_message(
                        "Record update status: " + str(updateRecord))
                    isQtyBy1000LogicSetForCharge = True

            if "Description" in columnList:
                descColIndex = columnList["Description"]
                descXpath = "//table[@id='SelectedProductType']/tbody/tr[" + str(
                    cnt) + "]/td[" + str(descColIndex) + "]//textarea"
                Drivers.driverWait.until(
                    EC.visibility_of_element_located((By.XPATH, descXpath)))
                txtBoxDesc = Drivers.driver.find_element_by_xpath(
                    descXpath)
                txtBoxDesc.clear()
                # if not data:
                txtBoxDesc.send_keys(selectedCharge)
                pricingDetails["DESCRIPTION-" + str(cnt)] = selectedCharge
                # else:
                #     txtBoxDesc.send_keys(data["Description"][cnt - 1])
                #     pricingDetails["DESCRIPTION-" +
                #                    str(cnt)] = data["Description"][cnt - 1]

            if "Sales Price" in columnList:
                salesPriceColIndex = columnList["Sales Price"]
                salesPriceXpath = "//table[@id='SelectedProductType']/tbody/tr[" + str(
                    cnt) + "]/td[" + str(salesPriceColIndex) + "]//input"
                Drivers.driverWait.until(
                    EC.visibility_of_element_located((By.XPATH, salesPriceXpath)))
                txtBoxSalesPrice = Drivers.driver.find_element_by_xpath(
                    salesPriceXpath)
                if not data:
                    salesPrice = round(Decimal(random.uniform(1.10, 99.99)), 2)
                else:
                    salesPrice = round(
                        Decimal(data["Sales Price"][cnt - 1]), 2)
                txtBoxSalesPrice.send_keys(str(salesPrice))
                pricingDetails[f"SALES PRICE ({oppCurrency})-" +
                               str(cnt)] = salesPrice

            if "Quantity" in columnList:
                quantityColIndex = columnList["Quantity"]
                quantityXpath = "//table[@id='SelectedProductType']/tbody/tr[" + str(
                    cnt) + "]/td[" + str(quantityColIndex) + "]//input"
                Drivers.driverWait.until(
                    EC.visibility_of_element_located((By.XPATH, quantityXpath)))
                txtBoxQuantity = Drivers.driver.find_element_by_xpath(
                    quantityXpath)
                txtBoxQuantity.clear()
                if not data:
                    quantity = Decimal(randint(1, 99))
                else:
                    quantity = Decimal(data["Quantity"][cnt - 1])
                txtBoxQuantity.send_keys(str(quantity))
                pricingDetails["QUANTITY-" + str(cnt)] = quantity

            if selectedPartner == "SavingStar" and selectedCharge == "Face Value Fee":
                if "% Of Redemption" in columnList:
                    redemptionColIndex = columnList["% Of Redemption"]
                    redemptionXpath = "//table[@id='SelectedProductType']/tbody/tr[" + str(
                        cnt) + "]/td[" + str(redemptionColIndex) + "]//input"
                    Drivers.driverWait.until(
                        EC.visibility_of_element_located((By.XPATH, redemptionXpath)))
                    txtBoxRedemption = Drivers.driver.find_element_by_xpath(
                        redemptionXpath)
                    if not data:
                        redemption = Decimal(randint(1, 100))
                    else:
                        redemption = Decimal(data["% Of Redemption"][cnt - 1])
                    txtBoxRedemption.clear()
                    txtBoxRedemption.send_keys(str(redemption))
                    pricingDetails["% OF REDEMPTION-" + str(cnt)] = redemption

#             salesPrice = float(salesPrice)
#             quantity = int(quantity)
#             redemption = int(redemption)
            if selectedPartner == "SavingStar" and selectedCharge == "Face Value Fee" and quantityBy1000Logic == False:
                subTotalAmount = round(
                    Decimal(salesPrice * quantity * (redemption / 100)), 2)
            elif selectedPartner == "SavingStar" and selectedCharge == "Face Value Fee" and quantityBy1000Logic == True:
                subTotalAmount = round(
                    Decimal(salesPrice * (quantity / 1000) * (redemption / 100)), 2)
            elif isQtyBy1000LogicSetForCharge == True:
                pricingDetails["Quantity/1000-" + str(cnt)] = (quantity / 1000)
                subTotalAmount = round(
                    Decimal(salesPrice * (quantity / 1000)), 2)
            elif isQtyBy1000LogicSetForCharge == False:
                pricingDetails["Quantity/1000-" + str(cnt)] = "-"
                subTotalAmount = round(Decimal(salesPrice * quantity), 2)
            totalAmount = totalAmount + subTotalAmount
            pricingDetails[f"SUB TOTAL ({oppCurrency})-" +
                           str(cnt)] = subTotalAmount
            pricingDetails["TOTAL AMOUNT"] = totalAmount

            if totalLineItem != 1 and cnt < (totalLineItem - 1):
                actionColIndex = columnList["Action"]
                actionXpath = "//table[@id='SelectedProductType']/tbody/tr[" + str(
                    cnt) + "]/td[" + str(actionColIndex) + "]//input[@class='btn addButton']"
                Drivers.driverWait.until(
                    EC.visibility_of_element_located((By.XPATH, actionXpath)))
                buttonAddRow = Drivers.driver.find_element_by_xpath(
                    actionXpath)
                buttonAddRow.click()
                sleep(1)
            opportunitiesDetails.update(pricingDetails)
            print("\n", "LineItems Detail: ")
            print(opportunitiesDetails)
            data_store.spec[productName] = opportunitiesDetails

    # --------------------------
    # Gauge step implementations
    # --------------------------
    @step("Open <tabName> Tab")
    def open_tab(self, tabName):
        try:
            Drivers.driverWait.until(
                EC.visibility_of_element_located((By.XPATH, "//div[@class='slds-icon-waffle']")))
            btnAllTabs = Drivers.driver.find_element_by_xpath(
                "//div[@class='slds-icon-waffle']")
            btnAllTabs.click()

            Drivers.driverWait.until(
                EC.visibility_of_element_located((By.XPATH, "//input[@type='search' and @placeholder='Search apps and items...']")))
            inputBoxSearch = Drivers.driver.find_element_by_xpath(
                "//input[@type='search' and @placeholder='Search apps and items...']")
            inputBoxSearch.clear()
            inputBoxSearch.send_keys(str(tabName))
            
            Drivers.driverWait.until(
                EC.visibility_of_element_located((By.XPATH, f"//p[@class='slds-truncate']/b[text()='{tabName}']")))
            tabLink = Drivers.driver.find_element_by_xpath(
                f"//p[@class='slds-truncate']/b[text()='{tabName}']")
            tabLink.click()
            
            # Drivers.driverWait.until(
            #     EC.visibility_of_element_located((By.LINK_TEXT, tabName)))
            # linkTabName = Drivers.driver.find_element_by_link_text(tabName)
            # linkTabName.click()
            Messages.write_message("Clicked on " + tabName + " tab")
            # Drivers.driverWait.until(
            #     EC.visibility_of_element_located((By.ID, "fcf")))

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
            # logging.info(f"{exc_type} == {fname} == {exc_tb.tb_lineno}")
            # self.sendErrorMail(f"{exc_type} == {fname} == {exc_tb.tb_lineno}")

    @step("Click on <buttonName> button")
    def click_button(self, buttonName):
        try:
            errorMsg = None
            
            xPath = f"//a[div[@title='{buttonName}']]"
            Drivers.driverWait.until(
                EC.visibility_of_element_located((By.XPATH, xPath)))
            # pdb.set_trace()
            buttonToClick = Drivers.driver.find_element_by_xpath(xPath)
            # Drivers.driver.execute_script("scroll(250, 0)")
            buttonToClick.click()
            sleep(5)

            if buttonName == "Create ILOC":
                multi_window = Drivers.driver.window_handles
                data_store.spec.clear
                data_store.spec['BASE_WINDOW'] = Drivers.driver.current_window_handle
                parentPath = Path(__file__).parents[1]
                ilocObjectRepositoryFileName = str(
                    parentPath) + "\\ObjectRepository\\" + os.getenv("ILOC_OBJECT_REPOSITORY_FILE")
                ilocObjectRepositorySheet = os.getenv(
                    "ILOC_OBJECT_REPOSITORY_SHEET")
                ilocObjectRepositoryJsonFileName = str(
                    parentPath) + "\\ObjectRepository\\" + f"{ilocObjectRepositorySheet}.json"

                if os.path.exists(ilocObjectRepositoryFileName):
                    Utils.set_json_from_object_repository(
                        ilocObjectRepositorySheet, ilocObjectRepositoryFileName)

                if os.path.exists(ilocObjectRepositoryJsonFileName):
                    f = open(ilocObjectRepositoryJsonFileName)
                    data_store.spec[ilocObjectRepositorySheet] = json.load(f)
                if len(multi_window) < 2:
                    buttonToClick.click()
                    sleep(5)
            if buttonName == "Send for Signature":
                try:
                    alert_obj = Drivers.driverWait.until(EC.alert_is_present())
                    # alert_obj = Drivers.driver.switch_to_alert()
                    alert_obj.accept()
                except UnexpectedAlertPresentException:
                    driverWait = WebDriverWait(Drivers.driver, 60)
                    alert = driverWait.until(EC.alert_is_present())
                    # alert_obj = Drivers.driver.switch_to_alert()
                    alert.accept()
                    print("Alert Accepted")
                    Messages.write_message("Alert Accepted")                    
            # errorImage = Drivers.driver.find_element_by_xpath("//img[@title='ERROR']")
            errorMsg = Drivers.driver.find_element_by_xpath(
                "//td[@class='messageCell']").text
            Messages.write_message(errorMsg)
            if errorMsg != None:
                raise Exception(errorMsg)

        except NoSuchElementException:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)

    @step("Click on button <buttonName>")
    def click_on_button(self, buttonName):
        try:
            xPath = f"//input[contains(@value,'{buttonName}')]"
            buttonToClick = Drivers.driver.find_element_by_xpath(xPath)
            buttonToClick.click()
            sleep(5)

        except NoSuchElementException:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
            
    @step("<buttonName> the ILOC")
    def click_on_button(self, buttonName):
        try:
            parentPath = Path(__file__).parents[1]
            ilocObjectRepositoryFileName = str(
                parentPath) + "\\ObjectRepository\\" + os.getenv("ILOC_OBJECT_REPOSITORY_FILE")
            ilocObjectRepositorySheet = os.getenv(
                "ILOC_OBJECT_REPOSITORY_SHEET")
            ilocObjectRepositoryJsonFileName = str(
                parentPath) + "\\ObjectRepository\\" + f"{ilocObjectRepositorySheet}.json"

            if os.path.exists(ilocObjectRepositoryFileName):
                Utils.set_json_from_object_repository(
                    ilocObjectRepositorySheet, ilocObjectRepositoryFileName)

            if os.path.exists(ilocObjectRepositoryJsonFileName):
                f = open(ilocObjectRepositoryJsonFileName)
                data_store.spec[ilocObjectRepositorySheet] = json.load(f)
            Drivers.driver.switch_to.window(Drivers.driver.window_handles[1])

            buttonToClick = Utils.get_element(buttonName, sheet_name=ilocObjectRepositorySheet,
                                              file_path=ilocObjectRepositoryFileName, json_data=data_store.spec.get(ilocObjectRepositorySheet))
            buttonToClick.click()

            # errorImage = Drivers.driver.find_element_by_xpath("//img[@title='ERROR']")
            errorMsg = Drivers.driver.find_element_by_xpath(
                "//td[@class='messageCell']").text
            Messages.write_message(errorMsg)

        except NoSuchElementException:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)

    @step("Send report email")
    def send_report_email(self):
        shutil.copy(os.getenv("XML_REPORT_FILE_NAME"),
                    os.getenv("XML_REPORT_FILE_COPY"))
        xmlFileName = os.getenv("XML_REPORT_FILE_COPY")
        xlxsFile = os.getenv("XLSX_RESULT_FILE_NAME")
        csvFile = os.getenv("CSV_RESULT_FILE_NAME")
        Utils.read_xml_copy_to_xlxs(xmlFileName, xlxsFile, csvFile)
        Utils.sendExecutionStatusMail()

    # # ---------------
    # # Execution Hooks
    # # ---------------
    # @after_spec
    # def after_spec_hook(self):
    #     Drivers.driver.quit()

    @step("User Logout from SalesForce")
    def user_loout_from_salesforce(self):
        Drivers.driverWait.until(
            EC.visibility_of_element_located((By.ID, "userNavButton")))
        userMenu = Drivers.driver.find_element_by_id("userNavButton")
        userMenu.click()

        Drivers.driverWait.until(
            EC.visibility_of_element_located((By.LINK_TEXT, "Logout")))
        linkLogout = Drivers.driver.find_element_by_link_text("Logout")
        linkLogout.click()

    @after_step
    def after_step_hook(self, context):
        if context.step.is_failing == True:
            Messages.write_message(context.step.text)
            # Messages.write_message(context.step.message)
            Screenshots.capture_screenshot()
		